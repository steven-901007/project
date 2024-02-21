from openpyxl import Workbook, load_workbook,workbook
from openpyxl.styles import Font
from turtle import title
from numpy import append

Awb = input('被比較的期數(小的數字)')
Bwb =  input('比較的期數(大的數字)')

wb = load_workbook ('C:/Users/steve/Desktop/學校/大三上/策略管理/mbs/mbs ' + Bwb + ' 期.xlsx')
ws = wb['資產負債表']
sheet = wb.active
a = 13
A=[]
i=0 
while i < a:
    i = i+1
    A.insert(i, ws.cell (i,2).value)

ws = wb['現金流量表']
sheet = wb.active
a = 34
B=[]
i=0
while i < a:
    i = i+1
    B.insert(i, ws.cell (i,2).value)

ws = wb['損益表']
sheet = wb.active
a = 50
C=[]
i=0
while i < a:
    i = i+1
    C.insert(i, ws.cell (i,2).value)

#以上是複製





wb = load_workbook('C:/Users/steve/Desktop/學校/大三上/策略管理/mbs/mbs ' + Awb + ' 期.xlsx')
ws = wb['資產負債表']

ws.insert_rows(1)
ws['B1'].value = Awb
ws['C1'].value = Bwb
ws['D1'].value = '比較(' + Bwb + '-' + Awb + ')'
ws.column_dimensions['D'].width = 12.0

for i in range(3,8):
    ws['D' + str(i)].value = '=C' + str(i) +'-B' + str(i)
for i in range(10,15):
    ws['D' + str(i)].value = '=C' + str(i) +'-B' + str(i)

sheet = wb.active
i = 0
a = 13
while i < a:
    ws.cell (i+2,3).value = A[i]
    i = i+1
# ws.move_range()

ws = wb['現金流量表']

ws.insert_rows(1)
ws['B1'].value = Awb
ws['C1'].value = Bwb
ws['D1'].value = '比較(' + Bwb + '-' + Awb + ')'
ws.column_dimensions['D'].width = 12.0

for i in range(3,7):
    ws['D' + str(i)].value = '=C' + str(i) +'-B' + str(i)
ws['D9'].value = "=C9-B9"
for i in range(12,16):
    ws['D' + str(i)].value = '=C' + str(i) +'-B' + str(i)
for i in range(17,20):
    ws['D' + str(i)].value = '=C' + str(i) +'-B' + str(i)
for i in range(22,36):
    ws['D' + str(i)].value = '=C' + str(i) +'-B' + str(i)

sheet = wb.active
i = 0
a = 34
while i < a:
    ws.cell (i+2,3).value = B[i]
    i = i+1


ws = wb['損益表']

ws.insert_rows(1)
ws['B1'].value = Awb
ws['C1'].value = Bwb
ws['D1'].value = '比較(' + Bwb + '-' + Awb + ')'
ws.column_dimensions['D'].width = 12.0

for i in range(3,8):
    ws['D' + str(i)].value = '=C' + str(i) +'-B' + str(i)
for i in range(10,20):
    ws['D' + str(i)].value = '=C' + str(i) +'-B' + str(i)
ws['D21'].value = "=C21-B21"
for i in range(26,33):
    ws['D' + str(i)].value = '=C' + str(i) +'-B' + str(i)
for i in range(35,42):
    ws['D' + str(i)].value = '=C' + str(i) +'-B' + str(i)
ws['D43'].value = "=C43-B43"
for i in range(46,52):
    ws['D' + str(i)].value = '=C' + str(i) +'-B' + str(i)

sheet = wb.active
i = 0
a = 50
while i < a:
    ws.cell (i+2,3).value = C[i]
    i = i+1


wb.save('C:/Users/steve/Desktop/學校/大三上/策略管理/mbs/mbs' + Awb + '期vs' + Bwb + '期.xlsx')
