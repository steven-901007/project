from turtle import title
from webbrowser import get
from numpy import append
from pkg_resources import working_set
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import ExcelWriter

path = "C:/Users/steve/Downloads/chromedriver_win32/chromedriver.exe"
driver = webdriver.Chrome(path)

driver.get("http://mbs.top-boss.com/login/a22?lang=zh_tw")

driver.find_element(By.ID,"cuid").send_keys("23413677")
driver.find_element(By.XPATH,'//*[@id="app"]/main/div/div/div/div[2]/div/div[6]/div/input').send_keys("mbs123")
driver.find_element(By.CLASS_NAME,"btn").click()

element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME,"card-top")))

driver.find_element(By.CLASS_NAME,"hamburger").click()
time.sleep(1)
driver.find_element(By.XPATH,'//*[@id="sidebar"]/div/ul/li[6]').click()
time.sleep(1)
driver.find_element(By.XPATH,'//*[@id="jixiao"]/li[1]/a').click()

element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="sunyi"]/div/div[1]/h5')))

day = driver.find_element(By.CLASS_NAME,'form-check-label').text

wb = Workbook()
ws = wb.active

ws.title = '資產負債表'
wb.create_sheet('現金流量表')
wb.create_sheet('損益表')

ws = wb["損益表"]

ws.append(["資產"])
ws.append(["現金"])
ws.append(["製成品存貨價值"])
ws.append(["原物料存貨價值"])
ws.append(["生產設備帳面價值"])
ws.append(["資產總額"])

ws.append([""])

ws.append(["負債及業主權益"])
ws.append(["借款"])
ws.append(["非正常負債"])
ws.append(["負債總額"])
ws.append(["業主權益"])
ws.append(["負債及業主權益總額"])




ws.append(["營業活動"])
ws.append(["銷售收益"])
ws.append(["現今費用支出"])
ws.append(["購料支出"])
ws.append(["營利事業所得稅"])

ws.append([""])

ws.append(["投資活動"])
ws.append(["設備投資支出"])

ws.append([""])

ws.append(["融資活動"])
ws.append(["還款"])
ws.append(["借款"])
ws.append(["非正常負債"])
ws.append(["股利支出"])

ws.append([""])

ws.append(["本期現金流量"])
ws.append(["期出現金餘額"])
ws.append(["期末現金餘額"])

ws.append([""])

ws.append(["現金費用支出項明細"])
ws.append(["行銷費用"])
ws.append(["研發費用"])
ws.append(["維護費用"])
ws.append(["人工費用"])
ws.append(["管理費用"])
ws.append(["原物料持有成本"])
ws.append(["製成品持有成本"])
ws.append(["設備投資相關費用"])
ws.append(["財務費用負債利息"])
ws.append(["訂購成本"])
ws.append(["工作班次變換成本"])
ws.append(["雜項費用"])
ws.append(["運費"])
ws.append(["情報交易收取費用"])



ws.append(["銷售收益"])
ws.append(["市場1"])
ws.append(["市場2"])
ws.append(["市場3"])
ws.append(["市場4"])
ws.append(["合計"])

ws.append([""])

ws.append(["營業成本"])
ws.append(["人工費用"])
ws.append(["材料耗用"])
ws.append(["銷貨成本修正額"])
ws.append(["維護費用"])
ws.append(["緊急採購成本"])
ws.append(["折舊"])
ws.append(["訂購成本"])
ws.append(["班次變換成本"])
ws.append(["設備投資相關費用"])
ws.append(["合計"])

ws.append([""])

ws.append(["毛利"])

ws.append([""])

ws.append(["營業費用"])

ws.append([""])

ws.append(["行銷費用"])
ws.append(["市場1"])
ws.append(["市場2"])
ws.append(["市場3"])
ws.append(["市場4"])
ws.append(["小計"])
ws.append(["運費"])
ws.append(["合計"])

ws.append([""])

ws.append(["管理費用"])
ws.append(["研究發展費用"])
ws.append(["管理費用及雜項費用"])
ws.append(["情報費用"])
ws.append(["製成品存貨持有成本"])
ws.append(["原物料持有成本"])
ws.append(["財務費用及利息支出"])
ws.append(["合計"])

ws.append([""])

ws.append(["稅前淨利(EBT)"])

ws.append([""])

ws.append(["*備註"])
ws.append(["稅前淨利(EBT)"])
ws.append(["財務費用及利息支出"])
ws.append(["稅前息前淨利(EBIT)"])
ws.append(["財務費用及利息支出"])
ws.append(["營利事業所得稅"])
ws.append(["稅後損益"])


for e in driver.find_elements(By.CLASS_NAME,'text-right'): #尋找多個class name
    ws.append([e.text])
    # print(e.text)  #列印出每個e的text 

#以上是複製資料

ws.move_range('A99:A103',rows=-97,cols=1)
ws.move_range('A104:A108',rows=-95,cols=1)
ws.move_range('A109:A112',rows=-94,cols=1)
ws.move_range('A113',rows=-92,cols=1)
ws.move_range('A114:A117',rows=-90,cols=1)
ws.move_range('A118:A120',rows=-89,cols=1)
ws.move_range('A121:A134',rows=-87,cols=1)
ws.move_range('A135:A139',rows=-86,cols=1)
ws.move_range('A140:A149',rows=-84,cols=1)
ws.move_range('A150',rows=-83,cols=1)
ws.move_range('A151:A157',rows=-79,cols=1)
ws.move_range('A158:A164',rows=-77,cols=1)
ws.move_range('A165',rows=-76,cols=1)
ws.move_range('A166:A171',rows=-74,cols=1)

wb.save('C:/Users/steve/Desktop/學校/大三上/策略管理/mbs/mbs '+day+'.xlsx')

#以上是整理格式

wb = load_workbook ('C:/Users/steve/Desktop/學校/大三上/策略管理/mbs/mbs '+day+'.xlsx')
ws = wb['損益表']
sheet = wb.active
a = 47
A = []
B = []
i=0
while i < a:
    i = i+1

    A.insert(i, ws.cell (i,1).value)
    B.insert(i, ws.cell (i,2).value)

for j in range(1,48):
    ws.delete_rows(1)


ws = wb['現金流量表']
sheet = wb.active
i = 0
while i < a:
    ws.cell (i+1,1).value = A[i]
    ws.cell (i+1,2).value = B[i]
    i = i+1

ws = wb['現金流量表']
sheet = wb.active
a = 13
A = []
B = []
i=0
while i < a:
    i = i+1

    A.insert(i, ws.cell (i,1).value)
    B.insert(i, ws.cell (i,2).value)

for j in range(1,14):
    ws.delete_rows(1)

ws = wb['資產負債表']
sheet = wb.active
i = 0
while i < a:
    ws.cell (i+1,1).value = A[i]
    ws.cell (i+1,2).value = B[i]
    i = i+1

wb.save('C:/Users/steve/Desktop/學校/大三上/策略管理/mbs/mbs '+day+'.xlsx')
#以上是複製資料至所應至位置

wb = load_workbook('C:/Users/steve/Desktop/學校/大三上/策略管理/mbs/mbs '+day+'.xlsx')

ws = wb["資產負債表"]
ws["A1"].font = Font (bold = True , color= "000000FF")#改style
ws["A8"].font = Font (bold = True , color= "000000FF")
ws.column_dimensions['A'].width = 20.0
ws.column_dimensions['B'].width = 12.0


ws = wb["現金流量表"]
ws["A1"].font = Font (bold = True , color= "000000FF")#改style
ws["A7"].font = Font (bold = True , color= "000000FF")
ws["A10"].font = Font (bold = True , color= "000000FF")
ws["A16"].font = Font (bold = True )
ws["A17"].font = Font (bold = True )
ws["A18"].font = Font (bold = True )
ws["A20"].font = Font (bold = True , color= "000000FF")
ws.column_dimensions['A'].width = 21.0
ws.column_dimensions['B'].width = 10.50

ws = wb["損益表"]
ws["A1"].font = Font (bold = True , color= "000000FF")#改style
ws["A8"].font = Font (bold = True , color= "000000FF")
ws["A20"].font = Font (bold = True )
ws["A22"].font = Font (bold = True , color= "000000FF")
ws["A29"].font = Font (bold = True )
ws["A31"].font = Font (bold = True )
ws["A33"].font = Font (bold = True , color= "000000FF")
ws["A40"].font = Font (bold = True )
ws["A42"].font = Font (bold = True )
ws["A44"].font = Font (bold = True , color= "000000FF")
ws["A45"].font = Font (bold = True )
ws["A47"].font = Font (bold = True )
ws["A50"].font = Font (bold = True )
ws.column_dimensions['A'].width = 20.20
ws.column_dimensions['B'].width = 10.7


wb.save('C:/Users/steve/Desktop/學校/大三上/策略管理/mbs/mbs '+day+'.xlsx')

#以上是改style


driver.close()






