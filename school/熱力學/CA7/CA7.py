import netCDF4 as nc
import re
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

# 定位最接近指定值的索引
def locate(list, number):
    return np.abs(list - number).argmin()

# 讀取 NetCDF 文件，獲取指定經緯度位置的索引
def map_coordinates(path, latitude, longitude):
    map_file = nc.Dataset(path)
    lat = map_file.variables["lat"][:]
    lon = map_file.variables["lon"][:]
    y = locate(lat, latitude)  # south_north
    x = locate(lon, longitude)  # west_east
    return x, y

# 讀取密度資料
def read_density(path):
    delimiter_pattern = re.compile(r'\s+|\n|\t')
    density_data = []
    with open(path, 'r') as file:
        for line in file:
            data = re.split(delimiter_pattern, line.strip())
            try:
                density_data.append(float(data[1]))
            except:
                pass
    # print(density_data)
    return density_data

# 讀取壓力資料
def read_pressure(path):
    delimiter_pattern = re.compile(r'\s+|\n|\t')
    pressure_data = []
    with open(path, 'r') as file:
        for line in file:
            data = re.split(delimiter_pattern, line.strip())
            try:
                pressure_data.append(float(data[1]))
            except:
                pass
    # print(pressure_data)
    return pressure_data

# 讀取溫度資料
def read_temperature(excel_path):
    wb = load_workbook(excel_path)
    ws = wb['Temperature']
    temp_data = [ws.cell(row=4+i, column=3).value for i in range(ws.max_row-3)]
    # print(temp_data)
    return temp_data
    

# 讀取和處理指定時間的 thermodynamics 和 surface 數據
def read_thermodynamics_data(thermo_path, surface_path, x, y, time_index):
    thermo_data = nc.Dataset(thermo_path)
    surface_data = nc.Dataset(surface_path)

    # 提取需要的變量
    zc = thermo_data.variables['zc'][:]  # 垂直高度
    th = thermo_data.variables['th'][time_index, :, y, x]  # 位溫
    qv = thermo_data.variables['qv'][time_index, :, y, x]  # 水汽混合比
    qc = thermo_data.variables['qc'][time_index, :, y, x]  # 雲水混合比
    zc = zc[2:]
    # print(zc)
    th = th[2:]
    qv = qv[2:]
    qc = qc[2:]
    # 計算等效位溫 (theta_e)
    Lv = 2.5e6  # 水的汽化潛熱 (J/kg)
    cp = 1004  # 空氣的比熱 (J/kg·K)
    theta_e = th * np.exp((Lv * qv) / (cp * th))

    # 計算飽和等效位溫 (theta_es)
    es = 6.112 * np.exp((17.67 * (th - 273.15)) / (th - 29.65))
    ws = (0.622 * es) / (pressure_data[0] - es)
    theta_es = th * np.exp((Lv * ws) / (cp * th))

    # 計算液態水混合比 (ql) 和總水混合比 (qt)
    ql = qc  # 液態水混合比等於雲水混合比
    qt = qv + ql  # 總水混合比

    return zc, th, theta_e, theta_es, qv, ql, qt, qc


#ytick
ytick = [150]
for i in range(1000,20000,1000):
    ytick.append(i)

# 繪製垂直剖面圖
def plot_vertical_profiles(zc, th, theta_e, theta_es, qv, ql, qt, qc):
    plt.rcParams['font.sans-serif'] = [u'MingLiu'] #細明體
    plt.rcParams['axes.unicode_minus'] = False 
    
    plt.figure(figsize=(10, 6))
    plt.plot(th, zc, label='$\\theta$')
    plt.plot(theta_e, zc, label='$\\theta_e$')
    plt.plot(theta_es, zc, label='$\\theta_{es}$')
    plt.plot(qv, zc, label='$q_v$')
    plt.plot(ql, zc, label='$q_l$')
    plt.plot(qt, zc, label='$q_t$')
    
    # 標記 Zm 和 Zcm
    Zm = zc[np.argmax(qt)]  # 假設 Zm 是總水混合比的最大值高度
    Zcm = zc[np.argmax(qc)]  # 假設 Zcm 是雲水混合比的最大值高度
    plt.axhline(Zm, color='red', linestyle='--', label='$Z_m$')
    plt.axhline(Zcm, color='blue', linestyle='--', label='$Z_{cm}$')
    plt.ylim(150)
    plt.xlabel('Variable Value')
    plt.ylabel('Height (m)')
    plt.yticks(ytick)
    plt.legend()
    plt.title('tpe20110802cln surver(140.112.66.200)\ntime :000000\n苗栗站(24.56457N,120.82458E)')
    plt.show()

# 主程序

# 定義文件路徑
density_path = "C:/Users/steve/python_data/thermodynamics/density.txt"
pressure_path = "C:/Users/steve/python_data/thermodynamics/pressure.txt"
excel_path = "C:/Users/steve/python_data/thermodynamics/CA1/Temperature.xlsx"
map_file_path = "C:/Users/steve/python_data/thermodynamics/TOPO.nc"
thermo_path = "C:/Users/steve/python_data/thermodynamics/Thermodynamic/tpe20110802cln.L.Thermodynamic-000000.nc"
surface_path = "C:/Users/steve/python_data/thermodynamics/Surface/exp.C.Surface-000144.nc"

# 讀取密度、壓力和溫度數據
density_data = read_density(density_path)
pressure_data = read_pressure(pressure_path)
temp_data = read_temperature(excel_path)

# 定位經緯度對應的索引
latitude, longitude = 24.82705, 121.01422  # 替換為你需要的座標
x, y = map_coordinates(map_file_path, latitude, longitude)

# 讀取 thermodynamics 和 surface 數據
time_index = 0  # 假設取第一個時間點的數據
zc, th, theta_e, theta_es, qv, ql, qt, qc = read_thermodynamics_data(thermo_path, surface_path, x, y, time_index)

# 繪製垂直剖面圖
plot_vertical_profiles(zc, th, theta_e, theta_es, qv, ql, qt, qc)
