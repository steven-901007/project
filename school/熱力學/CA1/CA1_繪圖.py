import matplotlib as mpl
import matplotlib.cm as cm
import matplotlib.colors as mcolors
import matplotlib.pyplot as plt
from openpyxl import load_workbook

wb = load_workbook("C:/Users/steve/python_data/thermodynamics/CA1/Temperature.xlsx")
ws = wb['Temperature']
X = ws.max_column-1
Y = ws.max_row-1
# print(X,Y)

DATA = []
P = []
H = []
TIME = []

#有效資料認證
for h in range(Y-1):
    data_list = []
    for t in range(X-2):
        data = ws.cell(h+2,t+3).value
        # print(data)        
        data_list.append(data)
    if data_list.count(-999) == 0:
        DATA.append(data_list)
        P.append(ws.cell(h+2,1).value/100)
        H.append(ws.cell(h+2,2).value/1000)
for t in range(X-1):
    TIME.append(ws.cell(1,t+3).value)
# print(DATA) #[T] DATA[[same hight]]
# print(P) #[hpa]
# print(H) #[km]
# print(TIME)


data_ave = []
for h in range(len(DATA)):
    ave = 0
    for t in range(len(DATA[h])):
           ave += DATA[h][t]
    data_ave.append(ave/len(DATA[h]))
    
# print(data_ave)    




#繪圖區

#散佈圖
plt.rcParams['font.sans-serif'] = [u'MingLiu'] #細明體
plt.rcParams['axes.unicode_minus'] = False 
fig = plt.figure()
ax = fig.add_subplot()


time_to_print = ['06:00','12:00','18:00']
            
for i in range(len(DATA[0])):
    data = []
    if time_to_print.count(TIME[i]) != 0:
        for h in range(len(DATA)):
            data.append(DATA[h][i])
    # print(data)
        ax.plot(data,H,lw = 2,label = TIME[i]) #繪製不同圖要改第二個變數



#等高座標繪圖
H_tick_set = [H[0]]
for i in range(round(max(H))-round(min(H))+1):
    H_tick_set.append(i+1)
# print(H_tick_set)
ax.plot(data_ave,H,color ='black',lw = 1,label = 'average')
ax.set_ylim(H[0])
ax.set_yticks(H_tick_set)
plt.ylabel('Height[km]',fontsize = 20)
        

#等壓座標作圖
# P_tick_set = []
# for i in range(round(min(P)),round(max(P))-round(min(P))+1,100):
#     P_tick_set.append(i+1)
# P_tick_set.append(max(P))
# P_tick_set.reverse()
# print(P_tick_set)
# ax.plot(data_ave,P,color ='black',lw = 1,label = 'average')
# ax.invert_yaxis()
# ax.set_ylim(P[len(P)-1])
# ax.set_yticks(P_tick_set)
# plt.ylabel('Pressure[mb]',fontsize = 20)



ax.tick_params(axis='x', labelsize=12) 
ax.tick_params(axis='y', labelsize=12) 

plt.legend()
plt.xlabel('Temperature[°C]',fontsize = 20)

plt.title('tpe20110802cln\n苗栗站(24.56457N,120.82458E)',fontsize = 20)
plt.show()