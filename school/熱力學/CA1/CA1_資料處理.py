import netCDF4 as nc
import glob
import re
from openpyxl import Workbook
import numpy as np
import os
# map file
def locate(list,number):
    return np.abs(list-number).argmin()

def map(path,latitude,longitude):

    map_file_path =path
    map = nc.Dataset(map_file_path)
    # print(map)



    # map data
    lat = map.variables["lat"][:]
    lon = map.variables["lon"][:]

    # 指定經緯度的值
    latitude = latitude
    longitude = longitude

    # 找到最接近指定經緯度值的索引
    y = locate(lat, latitude)  # south_north
    x = locate(lon, longitude)  # west_east
    # print(lat[y],lon[x])
    # print(lat)
    return x , y

map_file_path = "C:/Users/steve/python_data/thermodynamics/TOPO.nc"
x,y = map(map_file_path,24.56457,120.82458)    #苗栗座標
# x,y = map(map_file_path,24.95930,121.52000)    #座標
# print(x,y)    


#pressure file
def pressure(path):
    pressure_file_path = path

    delimiter_pattern = re.compile(r'\s+|\n|\   ')
    pressure_data = []
    with open(pressure_file_path, 'r') as file:
        for line in file:
            data = re.split(delimiter_pattern, line.strip())
            # print(data) #以列表顯示
            try:
                pressure_data.append(data[1])
            except:pass
    pressure_data.remove(pressure_data[0])
    # print(len(pressure_data))
    # print(pressure_data)    #從第一層到第70層
    return pressure_data
P = pressure("C:/Users/steve/python_data/thermodynamics/pressure.txt")  #[pa]
# print(P)    


#thermodynamics data file

Temperature = []

thermodynamics_folder = "C:/Users/steve/python_data/thermodynamics/archive/"

result  =glob.glob(thermodynamics_folder+'**')
for f in result:
    # f = os.path.basename(f) #單純輸出檔案名稱
    print(f)
    thermodynamics_file_path = f
    # thermodynamics_file_path = thermodynamics_folder+'tpe20110802cln.L.Thermodynamic-000029.nc'
    thrermodynamics_data = nc.Dataset(thermodynamics_file_path)
    # print(thrermodynamics_data)
    # print(thrermodynamics_data.variables["zc"][:])
    potential_temps = thrermodynamics_data.variables["th"][0]
    specific_humidity = thrermodynamics_data.variables["qv"][0]
    # print(potential_temps[0][y][x])  #potential temp
    temperature = []
    for i in range(len(potential_temps)):
        pt = potential_temps[i][y][x]
        qv = specific_humidity[i][y][x]
        if qv!=0:
            T = round(pt*((float(P[i])/100000)**(0.286))-273.15,5)
        else:
            T = -999
        # print(qv)
        temperature.append(T)
    Temperature.append(temperature)

# print(Temperature)

wb = Workbook()
ws = wb.active
ws.title = 'Temperature'

#寫入data
for i in range(len(Temperature)):
    for j in range(len(Temperature[0])):
        if Temperature[i][j] != -999:
            ws.cell(j+2,i+3).value = Temperature[i][j]
        else:
            ws.cell(j+2,i+3).value = -999

#加上高度
hight = thrermodynamics_data.variables["zc"]
for i in range(len(hight)):
    qv = specific_humidity[i][y][x]
    ws.cell(i+2,2).value =round(float(hight[i]))
ws.cell(1,2).value = 'hight'
#加上壓力
for i in range(len(P)):
    ws.cell(i+2,1).value =round(float(P[i]))
ws.cell(1,1).value = 'pressure'
#加上時間
lc = 3
for i in range(0,24): 
    for j in range(0,51,10):    
        time = str(i).zfill(2)+':'+str(j).zfill(2)
        ws.cell(1,lc).value = time
        lc +=1
ws.cell(1,lc).value = '24:00'
            
wb.save("C:/Users/steve/python_data/thermodynamics/CA1/Temperature.xlsx")
