import netCDF4 as nc
import glob
import re
from openpyxl import load_workbook
import numpy as np
import matplotlib as mpl
import matplotlib.cm as cm
import matplotlib.colors as mcolors
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
# map file
def locate(list,number):
    return np.abs(list-number).argmin()

def map(path,latitude,longitude):

    map_file_path =path
    map = nc.Dataset(map_file_path)
    # print(map)



    # map data
    lat = map.variables["lat"][:]
    lon = map.variables["lon"][:]

    # 指定經緯度的值
    latitude = latitude
    longitude = longitude

    # 找到最接近指定經緯度值的索引
    y = locate(lat, latitude)  # south_north
    x = locate(lon, longitude)  # west_east
    # print(lat[y],lon[x])
    # print(lat)
    return x , y

map_file_path = "C:/Users/steve/python_data/thermodynamics/TOPO.nc"
# x,y = map(map_file_path,24.56457,120.82458)    #苗栗座標
x,y = map(map_file_path,23.96,120.30586)    #座標
# print(x,y)


#pressure file
def pressure(path):
    pressure_file_path = path

    delimiter_pattern = re.compile(r'\s+|\n|\   ')
    pressure_data = []
    with open(pressure_file_path, 'r') as file:
        for line in file:
            data = re.split(delimiter_pattern, line.strip())
            # print(data) #以列表顯示
            try:
                pressure_data.append(data[1])
            except:pass
    pressure_data.remove(pressure_data[0])
    # print(len(pressure_data))
    # print(pressure_data)    #從第一層到第70層
    return pressure_data
P = pressure("C:/Users/steve/python_data/thermodynamics/pressure.txt")  #[pa]
# print(P[2])


# wth(surface flux of potential temperature)
wth_time_list = []
wth_folder = "C:/Users/steve/python_data/thermodynamics/Surface/"
result  =glob.glob(wth_folder+'**')
for wth_file_path in result:
    # print(wth_file_path)
# wth_file_path = "C:/Users/steve/python_data/thermodynamics/Surface/tpe20110802cln.C.Surface-000000.nc"
    wth_data = nc.Dataset(wth_file_path)
    # print(wth_data)
    wth = wth_data.variables['wth'][0][x][y]
    # print(wth)
    wth_time_list.append(wth)
# print(wth_time_list)


#Cp(speccific heat at constant pressure)
Cp = 1004


#Q(sensible heat flux)[J/(m^2*s)]
Q_time_list = []
for i in range(len(wth_time_list)):
    Q = float(wth_time_list[i])*1004*((float(P[2])/100000)**0.286)
    Q = round(Q,2)
    Q_time_list.append(Q)
print(Q_time_list)
add_Q_time_list = [0]
for i in range(len(Q_time_list)):
    add_Q_time_list.append(add_Q_time_list[i]+Q_time_list[i])

add_Q_time_list.remove(add_Q_time_list[0])
# print(add_Q_time_list)


# T0 [C] 模式氣溫
T0_time_list= []
wb = load_workbook("C:/Users/steve/python_data/thermodynamics/CA1/Temperature.xlsx")
ws = wb['Temperature']
for i in range(ws.max_column-2):
    T = round(ws.cell(4,i+3).value,2)
    T0_time_list.append(T) 
# print(T0_time_list)
add_T0_time_list = [0]
for i in range(len(T0_time_list)):
    add_T0_time_list.append(add_T0_time_list[i]+T0_time_list[i])

add_T0_time_list.remove(add_T0_time_list[0])   

# dt [s]
dt = 600


#T 
T_time_list = []
for i in range(len(T0_time_list)):
    T0 =  T0_time_list[i]
    q = Q_time_list[i]
    T = round(T0+q/1004*600,2)
    T_time_list.append(T)
# print(T_time_list)
add_T_time_list = [0]
for i in range(len(T_time_list)):
    add_T_time_list.append(add_T_time_list[i]+T_time_list[i])

add_T_time_list.remove(add_T_time_list[0])


#time
Time = []
for i in range(0,24): 
    for j in range(0,51,10):    
        time = str(i).zfill(2)+':'+str(j).zfill(2)
        Time.append(time)
Time.append('24:00')
# print(Time)

time_tick = []
time_count =[]
Y = []
for i in range(0,len(Time)):
    Y.append(i)
    if Time[i][3:] == '00':
        time_tick.append(Time[i])
        time_count.append(i)
        
    # else:
        # print(Time[3:])
        # time_tick.append(' ')
print(len(time_tick))
print(len(Y))


#繪圖
plt.rcParams['font.sans-serif'] = [u'MingLiu'] #細明體
plt.rcParams['axes.unicode_minus'] = False 
fig = plt.figure()
a1 = fig.add_subplot()
a1.plot(Y,add_Q_time_list,color =  'r',linestyle = '-',label ='sensible heat flux')
a1.set_xticks(time_count,time_tick,fontsize = 15,rotation = 60)
a1.set_xlabel('Time [s]',fontsize = 20)
a1.set_ylabel('sensible heat flux [J/(m^2*s)]',fontsize = 20)
plt.legend(fontsize = 15)
plt.yticks(fontsize = 20)
a1.set_ylim(0)
plt.title('tpe20110802cln surver(140.112.66.200)\n苗栗站(24.56457N,120.82458E)',fontsize = 20)

fig = plt.figure()
a2 = fig.add_subplot()
a2.plot(Y,add_T0_time_list,color =  'g',linestyle = '-',label = 'near-surface air temperature')
a2.plot(Y,add_T_time_list,color =  'y',linestyle = '-',label = 'estimation by the sensible heat')
plt.fill_between(Y,add_T0_time_list, add_T_time_list, color= (1.0, 0.647, 0.0), alpha=0.3)
plt.fill_between(Y,add_T0_time_list, color= (0,1,0), alpha=0.3)
a2.set_xticks(time_count,time_tick,fontsize = 15,rotation = 60)
a2.set_xlabel('Time [s]',fontsize = 20)
a2.set_ylabel('temperature [°C]',fontsize = 20)
plt.yticks(fontsize = 20)
a2.set_ylim(min(T0_time_list))
plt.title('tpe20110802cln surver(140.112.66.200)\n苗栗站(24.56457N,120.82458E)',fontsize = 20)
plt.legend(fontsize = 15)
plt.show()