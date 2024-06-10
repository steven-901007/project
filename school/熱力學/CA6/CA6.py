import numpy as np
import netCDF4 as nc
import re
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# time = 49
time = 121

# 地圖文件
def locate(lst, number):
    return np.abs(lst - number).argmin()

def map_coords(path, latitude, longitude):
    map_file_path = path
    map_data = nc.Dataset(map_file_path)
    lat = map_data.variables["lat"][:]
    lon = map_data.variables["lon"][:]
    
    y = locate(lat, latitude)  # south_north
    x = locate(lon, longitude)  # west_east
    return x, y

map_file_path = "C:/Users/steve/python_data/thermodynamics/TOPO.nc"
x, y = map_coords(map_file_path, 24.56457, 120.82458)  # 苗栗座標
# print(x, y)

# 读取温度数据
file_path_temp = "C:/Users/steve/python_data/thermodynamics/CA1/Temperature.xlsx"
temp_profile = []
wb = load_workbook(file_path_temp)
ws = wb['Temperature']
t = ws.max_column - 1
h = ws.max_row - 1
for i in range(h - 2):
    temp_profile.append(ws.cell(i + 4, time + 2).value)
# print(temp_profile)

# 读取特定湿度数据
file_path_thermodynamics = "C:/Users/steve/python_data/thermodynamics/Thermodynamic/tpe20110802cln.L.Thermodynamic-000" + str(time).zfill(3) + ".nc"
thermodynamics_data = nc.Dataset(file_path_thermodynamics)
qv_profile = []
for i in range(2, 70):
    qv = thermodynamics_data.variables['qv'][0][i][y][x]*1000
    qv_profile.append(qv)
# print(len(qv_profile))

# 读取压力数据
def pressure(path):
    pressure_file_path = path
    delimiter_pattern = re.compile(r'\s+|\n|\t')
    pressure_data = []
    with open(pressure_file_path, 'r') as file:
        for line in file:
            data = re.split(delimiter_pattern, line.strip())
            try:
                pressure_data.append(float(data[1]) / 100)  # 压力数据以 Pa 为单位，转换为 hPa
            except:
                pass
    return pressure_data

P_list = pressure("C:/Users/steve/python_data/thermodynamics/pressure.txt")  # [hPa]
P_list = P_list[2:]

# 计算饱和比湿
def calculate_qvs(temp_profile, P_list):
    qvs_profile = []
    for T, P in zip(temp_profile, P_list):
        e_s = 6.112 * np.exp(17.67 * T / (T + 243.5))  # 计算饱和水汽压
        qvs = 0.622 * e_s / (P - e_s) *1000 # 计算饱和比湿
        qvs_profile.append(qvs)
    return qvs_profile

qvs_profile = calculate_qvs(temp_profile, P_list)
real_time = ws.cell(1, time + 2).value
# 绘图
def plot_profiles(zc, temp_profile, qv_profile, qvs_profile):
    plt.figure(figsize=(15, 5))
    # 温度剖面
    plt.subplot(1, 3, 1)
    plt.plot(temp_profile, zc)
    plt.xlabel('Temperature (K)')
    plt.ylabel('Height (m)')
    plt.title('Temperature Profile')

    # 特定湿度剖面
    plt.subplot(1, 3, 2)
    plt.plot(qv_profile, zc)
    plt.xlabel('Specific Humidity (kg/kg)')
    plt.title('Specific Humidity Profile')

    # 饱和比湿剖面
    plt.subplot(1, 3, 3)
    plt.plot(qvs_profile, zc)
    plt.xlabel('Saturation Mixing Ratio (kg/kg)')
    plt.title('Saturation Mixing Ratio Profile')

    plt.tight_layout()
    plt.show()

# 获取高度数据
zc = thermodynamics_data.variables['zc'][2:]  # 假设高度数据在 'zc' 变量中，从第3个数据开始

# 调用绘图函数
plot_profiles(zc, temp_profile, qv_profile, qvs_profile)
