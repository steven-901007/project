import netCDF4 as nc
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from metpy.plots import SkewT
from metpy.units import units
import metpy.calc as mpcalc
from openpyxl import load_workbook
import re


time = 120
# for i in  range(0,145):
#     time =i
# 讀取數據文件
file_path_thermodynamics = "C:/Users/steve/python_data/thermodynamics/Thermodynamic/tpe20110802cln.L.Thermodynamic-000" + str(time).zfill(3) + ".nc"
file_path_surface = "C:/Users/steve/python_data/thermodynamics/Surface/exp.C.Surface-000" + str(time).zfill(3) + ".nc"
file_path_temp = "C:/Users/steve/python_data/thermodynamics/CA1/Temperature.xlsx"
thermodynamics_data = nc.Dataset(file_path_thermodynamics)
surface_data = nc.Dataset(file_path_surface)
# 列出變量以確認
# print(thermodynamics_data.variables.keys())
# print(surface_data.variables.keys())

# 高度
height = thermodynamics_data.variables['zc'][:]
height = height[2:]
# print(len(height))

# 讀取壓力
def pressure(path):
    pressure_file_path = path
    delimiter_pattern = re.compile(r'\s+|\n|\   ')
    pressure_data = []
    with open(pressure_file_path, 'r') as file:
        for line in file:
            data = re.split(delimiter_pattern, line.strip())
            try:
                pressure_data.append(float(data[1]) / 100)  # 壓力數據以 Pa 為單位，轉換為 hPa
            except:
                pass
    return pressure_data

P_list = pressure("C:/Users/steve/python_data/thermodynamics/pressure.txt")  # [hPa]
P_list = P_list[2:]
# print(P_list)

# 地圖文件
def locate(list, number):
    return np.abs(list - number).argmin()

def map(path, latitude, longitude):
    map_file_path = path
    map = nc.Dataset(map_file_path)
    lat = map.variables["lat"][:]
    lon = map.variables["lon"][:]
    
    # 找到最接近指定經緯度值的索引
    y = locate(lat, latitude)  # south_north
    x = locate(lon, longitude)  # west_east
    return x, y

map_file_path = "C:/Users/steve/python_data/thermodynamics/TOPO.nc"
x, y = map(map_file_path, 24.56457, 120.82458)  # 苗栗座標
# print(x, y)

# 溫度剖面
temp_profile = []

wb = load_workbook(file_path_temp)
ws = wb['Temperature']
t = ws.max_column - 1
h = ws.max_row - 1
for i in range(h - 2):
    temp_profile.append(ws.cell(i + 4, time + 2).value)
# print(temp_profile)
real_time = ws.cell(1, time + 2).value
# 地表溫度
surface_temp = temp_profile[0]
# print(surface_temp)

# 比濕剖面
specific_humidity_profile = []
for i in range(2, 70):
    qv = thermodynamics_data.variables['qv'][0][i][y][x]
    specific_humidity_profile.append(qv / (1 + qv))
# print(len(specific_humidity_profile))

# 地表比濕
surface_qv = specific_humidity_profile[0]

# 確認數據是否正確
# print("Pressure (hPa):", P_list)
# print("Temperature Profile (°C):", temp_profile)
# print("Dewpoint Profile (°C):", specific_humidity_profile)

# 轉換數據單位
p = np.array(P_list, dtype=float) * units.hPa
temperature = np.array(temp_profile, dtype=float) * units.degC
dewpoint = np.array(specific_humidity_profile, dtype=float) * units.degC

# print("Converted Pressure (hPa):", p)
# print("Converted Temperature (°C):", temperature)
# print("Converted Dewpoint (°C):", dewpoint)

# 計算 LCL
def calculate_lcl(t, qv):
    lcl = mpcalc.lcl(1000 * units.hPa, t * units.degC, qv * units.degC)
    return lcl[0].magnitude, lcl[1].magnitude

# 計算 LFC、EL、CAPE 和 CIN
def calculate_lfc_el_cape_cin(p, t, td, parcel_prof):
    lfc_p, lfc_t = mpcalc.lfc(p, t, td)
    el_p, el_t = mpcalc.el(p, t, td)
    cape, cin = mpcalc.cape_cin(p, t, td, parcel_profile=parcel_prof)
    return lfc_p.magnitude, lfc_t.magnitude, el_p.magnitude, el_t.magnitude, cape.magnitude, cin.magnitude

# 讀取數據並計算
parcel_profile = mpcalc.parcel_profile(p, temperature[0], dewpoint[0]).to('degC')

lcl_p, lcl_t = calculate_lcl(surface_temp, surface_qv)
lfc_p, lfc_t, el_p, el_t, cape, cin = calculate_lfc_el_cape_cin(p, temperature, dewpoint, parcel_profile)

# print(f"LCL: {lcl_p} hPa, {lcl_t} °C")
# print(f"LFC: {lfc_p} hPa, {lfc_t} °C")
# print(f"EL: {el_p} hPa, {el_t} °C")
# print(f"CAPE: {cape} J/kg")
# print(f"CIN: {cin} J/kg")

# 計算 COIN
if cape + cin != 0:
    coin = cape / (cape + cin)
else:
    coin = 0

# print(f"COIN: {coin}")

# 繪製斜溫圖
plt.rcParams['font.sans-serif'] = [u'MingLiu'] #細明體
plt.rcParams['axes.unicode_minus'] = False 
fig = plt.figure(figsize=(9, 9))
skew = SkewT(fig, rotation=45)

skew.ax.set_ylim(max(P_list), 100)
skew.ax.set_xlim(-30, 40)
skew.ax.set_ylabel('Pressure (hPa)')
skew.ax.set_xlabel('Temperature (°C)')

skew.plot_dry_adiabats()
skew.plot_moist_adiabats()
skew.plot_mixing_lines()

# 確保數據沒有問題
# print("Plotting Temperature and Dewpoint profiles")

skew.plot(p, temperature, 'blue')
skew.plot(p, dewpoint, 'r')

skew.plot(p, parcel_profile, 'k', linewidth=2)

skew.shade_cape(p, temperature, parcel_profile)
skew.shade_cin(p, temperature, parcel_profile)

text = (f'LCL: {lcl_p:.2f} hPa, {lcl_t:.2f} °C\n'
        f'LFC: {lfc_p:.2f} hPa, {lfc_t:.2f} °C\n'
        f'EL: {el_p:.2f} hPa, {el_t:.2f} °C\n'
        f'CAPE: {cape:.2f} J/kg\n'
        f'CIN: {cin:.2f} J/kg\n'
        f'COIN: {coin:.2f}')

plt.text(0.95, 0.95, text, horizontalalignment='right', verticalalignment='top', backgroundcolor='w', transform=skew.ax.transAxes)
plt.title('tpe20110802cln surver(140.112.66.200)\nSkew-T Log-P Diagram\n苗栗站(24.56457N,120.82458E)\ntime'+real_time)
plt.show()

# 繪製 COIN 的時間演變圖
# plt.figure(figsize=(10, 6))
# plt.plot(time, coin, label='COIN')
# plt.xlabel('Time')
# plt.ylabel('COIN')
# plt.legend()
# plt.title('Time Evolution of COIN')
# plt.tight_layout()
# plt.show()
    # print(i,time,cin,cape,coin)