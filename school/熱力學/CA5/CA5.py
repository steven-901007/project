import netCDF4 as nc
import glob
import re
from openpyxl import load_workbook
import numpy as np
import matplotlib as mpl
import matplotlib.cm as cm
import matplotlib.colors as mcolors
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
# map file
def locate(list,number):
    return np.abs(list-number).argmin()

def map(path,latitude,longitude):

    map_file_path =path
    map = nc.Dataset(map_file_path)
    # print(map)



    # map data
    lat = map.variables["lat"][:]
    lon = map.variables["lon"][:]

    # 指定經緯度的值
    latitude = latitude
    longitude = longitude

    # 找到最接近指定經緯度值的索引
    y = locate(lat, latitude)  # south_north
    x = locate(lon, longitude)  # west_east
    # print(lat[y],lon[x])
    # print(lat)
    return x , y


#density
def ro(path):
    ro_file_path = path
    delimiter_pattern = re.compile(r'\s+|\n|\   ')
    ro_data = []
    with open(ro_file_path, 'r') as file:
        for line in file:
            data = re.split(delimiter_pattern, line.strip())
            # print(data) #以列表顯示
            try:
                ro_data.append(data[1])
            except:pass
    ro_data.remove(ro_data[0])
    # print(len(ro_data))
    # print(ro_data)    #從第一層到第70層
    return ro_data
ro_list = ro("C:/Users/steve/python_data/thermodynamics/density.txt")  #[pa]
# print(ro_list)    


#pressure
def pressure(path):
    pressure_file_path = path

    delimiter_pattern = re.compile(r'\s+|\n|\   ')
    pressure_data = []
    with open(pressure_file_path, 'r') as file:
        for line in file:
            data = re.split(delimiter_pattern, line.strip())
            # print(data) #以列表顯示
            try:
                pressure_data.append(data[1])
            except:pass
    pressure_data.remove(pressure_data[0])
    # print(len(pressure_data))
    # print(pressure_data)    #從第一層到第70層
    return pressure_data
P_list = pressure("C:/Users/steve/python_data/thermodynamics/pressure.txt")  #[pa]
# print(P_list[2])   


map_file_path = "C:/Users/steve/python_data/thermodynamics/TOPO.nc"
# x,y = map(map_file_path,24.56457,120.82458)    #苗栗座標
x,y = map(map_file_path,24.82705,121.01422)    #座標
# print(x,y)


#temp
temp_time_list = []
wb = load_workbook("C:/Users/steve/python_data/thermodynamics/CA1/Temperature.xlsx")
ws = wb['Temperature']
t = ws.max_column-1
h = ws.max_row-1
for i in range(t-1):
    temp_time_list.append(ws.cell(4,i+3).value)
# print(temp_time_list)




# wth(surface flux of potential temperature) and hight

max_zm_hight_time_list = []
time_chack = []
es_time_list = []
wth_folder = "C:/Users/steve/python_data/thermodynamics/Thermodynamic/"
result  =glob.glob(wth_folder+'**')
for file_path in result:
    print(file_path[len(file_path)-6:len(file_path)-3])
    # file_path = "C:/Users/steve/python_data/thermodynamics/Thermodynamic/tpe20110802cln.L.Thermodynamic-000000.nc"
    thrermodynamics_data = nc.Dataset(file_path)
    # print(thrermodynamics_data)
    # pt_list = []
    hight_list = []
    qv_list = []
    add_qv_list = []
    Y = []
    # potential_temps = thrermodynamics_data.variables["th"][0]
    hight = thrermodynamics_data.variables["zc"]
    specific_humidity = thrermodynamics_data.variables["qv"][0]
    for i in range(len(specific_humidity)):
        qv = specific_humidity[i][y][x]
        if qv != 0:
            add_qv_list.append(qv*int(hight[i]))
            hight_list.append(int(hight[i]))
            qv_list.append(qv)
            # print(qv*int(hight[i]))
    # print(add_qv_list)
    # print(sum(add_qv_list))
    # print(qv_list)
    max_zm_hight_time_list.append(sum(add_qv_list))
    es = qv_list[0]*float(P_list[2])/100/(0.622+0.378*qv_list[0])
    es_time_list.append(es)
# print(max_zm_hight_time_list)
print(es_time_list)

#Td
cloud_bass_time_list = []
Td_time_list = []
for i in range(len(es_time_list)):
    Td = (243.5*np.log(es_time_list[i]/6.112))/(17.67-np.log(es_time_list[i]/6.112))
    T = temp_time_list[i]
    Td_time_list.append(Td)
    h = 125*(T-Td)
    cloud_bass_time_list.append(h)
print(Td_time_list)
print(cloud_bass_time_list)

#wqv
wqv_list = []
wth_folder = "C:/Users/steve/python_data/thermodynamics/Surface/"
result  =glob.glob(wth_folder+'**')
ub = 0
for file_path in result:
    print(file_path)
    # file_path = "C:/Users/steve/python_data/thermodynamics/Surface/exp.C.Surface-000144.nc"
    Surface_data = nc.Dataset(file_path)
    # print(Surface_data)

    Y.append(ub)
    ub +=1
    # surface_flux_of_pt = Surface_data.variables["wth"][0][y][x]
    surface_flux_of_wv = Surface_data.variables["wqv"][0][y][x]
    # print(surface_flux_of_wv)
    wqv_list.append(surface_flux_of_wv*600)
add_wqv_list = [0]
for i in range(len(wqv_list)):
    add_wqv_list.append(add_wqv_list[i]+wqv_list[i])
add_wqv_list.pop(0)
# print(wqv_list)
# print(add_wqv_list)

total_max_zm_hight_time_list = []
for i in range(len(max_zm_hight_time_list)):
    total_max_zm_hight_time_list.append(max_zm_hight_time_list[i]+add_wqv_list[i])
print(total_max_zm_hight_time_list)


#time set
Time = []
for i in range(0,24): 
    for j in range(0,51,10):    
        time = str(i).zfill(2)+':'+str(j).zfill(2)
        Time.append(time)
Time.append('24:00')
# print(Time)

time_tick = []
time_count =[]
Y = []
for i in range(0,len(Time)):
    Y.append(i)
    if Time[i][3:] == '00':
        time_tick.append(Time[i])
        time_count.append(i)
        
    # else:
        # print(Time[3:])
        # time_tick.append(' ')
# print(len(time_tick))


#繪圖
plt.rcParams['font.sans-serif'] = [u'MingLiu'] #細明體
plt.rcParams['axes.unicode_minus'] = False 
fig = plt.figure()
a1 = fig.add_subplot()
a1.plot(Y,cloud_bass_time_list,color = 'g',linestyle = '-',label = 'Zc')
a1.plot(Y,total_max_zm_hight_time_list,color =  'r',linestyle = '-',label ='Zm')
a1.set_xticks(time_count,time_tick,fontsize = 15,rotation = 60)

a1.set_xlabel('Time',fontsize = 20)
a1.set_ylabel('hight',fontsize = 20)
plt.legend(fontsize = 15)
plt.yticks(fontsize = 20)
# a1.set_ylim(0)

plt.title('tpe20110802cln surver(140.112.66.200)\ntime :20110802\n苗栗站(24.56457N,120.82458E)',fontsize = 20)
plt.show()