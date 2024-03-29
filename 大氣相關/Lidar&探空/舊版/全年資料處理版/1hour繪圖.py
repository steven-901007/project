import matplotlib.pyplot as plt
import math
from openpyxl import load_workbook
import matplotlib as mpl
import math
import cmath
from math import *
import time as T
import matplotlib.cm as cm
import matplotlib.colors as mcolors
import matplotlib.ticker as ticker
import numpy as np
import os
import matplotlib.dates as dates
from matplotlib.dates import AutoDateLocator
import glob


locate = '松山'
year = '109'

#風標圖


startTime = T.time()


Day =[]
# day = '2019-12-06'
files = glob.glob('C:/Users/steve/Desktop/python相關資料/need data information/' + locate + '/' + year + '/1hour/*')
for file in files:
    day = file[69:79]
    Day.append(day)
for day in Day[100:130 ]: #分區繪圖(數字代表間隔 要改不然會無法繪圖)
    # try:
    wb = load_workbook("C:/Users/steve/Desktop/python相關資料/need data information/" + locate + '/' + year + "/1hour/"+day+".xlsx")
    wsu = wb['U']
    wsv = wb['V']
    colmax = wsu.max_column #24
    rowmax = wsu.max_row #64

    # print (wsu.max_column,wsu.max_row)
    xtik = []
    for i in range(2,colmax+1):
            xtik.append(wsu.cell(1,i).value[:2]+':'+wsu.cell(1,i).value[3:])


    X = [] #time
    Y = [] #hight
    U = []
    V = []
    for col in range(1,colmax):
        X.append(col)
    # print (X)

    for row in range(wsu.cell(2,1).value,wsu.cell(rowmax,1).value+1,25):
        Y.append(row)
    # print (Y)
    U = [[]for i in range(rowmax-1)]
    # print(len(U))
    for j in range(2,rowmax+1):
        
        for i in range(2,colmax+1):
            if wsu.cell(j,i).value == None:
                U[j-2].append(-999)
            else:
                U[j-2].append(wsu.cell(j,i).value)
    # print (U)
    V = [[]for i in range(rowmax-1)]
    for j in range(2,rowmax+1):
        
        for i in range(2,colmax+1):
            if wsv.cell(j,i).value == None:
                V[j-2].append(-999)
            else:
                V[j-2].append(wsv.cell(j,i).value)
    # print (V)

    wbd = load_workbook('C:/Users/steve/Desktop/python相關資料/need data information/'  + locate + '/' + year + '/windcut/' + day + '.xlsx')
    ws = wbd['風切']

    hights =  ws.max_row
    times = ws.max_column

    xb = [] #<0.08
    yb = [] 
    xg = [] #0.08~0.15
    yg = [] 
    xy = [] #0.15~0.2
    yy = [] 
    xr = [] #>0.2
    yr = [] 
    countb = []
    countg = []
    county = []
    countr = []
    for time in range(2,times+1):
        countb.append(0)
        countg.append(0)
        county.append(0)
        countr.append(0)
        for hight in range(2,hights+1):
            if ws.cell(hight,time).value != None:
                if ws.cell(hight,time).value <0.08:
                    xb.append((time-1)*colmax/times)
                    yb.append(ws.cell(hight,1).value)
                    countb[time-2] += 1
                elif 0.08<= ws.cell(hight,time).value <0.15:
                    xg.append((time-1)*colmax/times)
                    yg.append(ws.cell(hight,1).value) 
                    countg[time-2] += 1       
                elif 0.15<= ws.cell(hight,time).value <0.2:
                    xy.append((time-1)*colmax/times)
                    yy.append(ws.cell(hight,1).value)
                    county[time-2] += 1
                elif 0.2<=ws.cell(hight,time).value:
                    xr.append((time-1)*colmax/times)
                    yr.append(ws.cell(hight,1).value)
                    countr[time-2] += 1
    ti = []
    for i in range(2,times+1):
        ti.append(ws.cell(1,i).value[:2]+':'+ws.cell(1,i).value[3:])


    showti = []
    for i in range(1,24):
        i =str(i).zfill(2)
        if i != '23':
            showti.append(i+':00')
            showti.append(i+':30')
        else :
            showti.append(i+':00')
    # print(showti)


    # Z = []

    Colors = []
    for j in range(0,rowmax-1):
        for i in range(0,colmax-1):
            # j = 60-j    

            u = U[j][i]
            v = V[j][i]
            if u != -999 and v != -999:
                z = round(sqrt(pow(u,2) +pow(v,2)),2)
            else:
                z = -999
            # Z.append(z)
            if 0<=z<=1:
                Colors.append('purple')
            elif 1<z<=2:
                Colors.append('blue')
            elif 2<z<=3:
                Colors.append('deepskyblue')
            elif 3<z<=4:
                Colors.append('turquoise')
            elif 4<z<=5:
                Colors.append('mediumturquoise')
            elif 5<z<=6:
                Colors.append('chartreuse')
            elif 6<z<=7:
                Colors.append('greenyellow')
            elif 7<z<=8:
                Colors.append('yellow')
            elif 8<z<=9:
                Colors.append('gold')
            elif 9<z<=10:
                Colors.append('orange')
            elif 10<z<=11:
                Colors.append('orangered')
            elif 11<z:
                Colors.append('fuchsia')
            elif z == -999:
                Colors.append('w')


    for j in range(rowmax-1):
        
        for i in range(colmax-1):
            if U[j][i] == -999:
                U[j][i] = 0
            if V[j][i] == -999:
                V[j][i] = 0


    fig1,axes1 = plt.subplots(1,1,figsize=(19,9))

    axes1.barbs(X,Y,U,V,color = Colors)

    axes1.scatter(xr,yr,color = 'red',s= 5,alpha=0.6,label='>0.2')
    axes1.scatter(xy,yy,color = 'gold',s= 5,alpha=0.6,label='0.15~0.2') 
    axes1.scatter(xg,yg,color = 'green',s= 5,alpha=0.6,label='0.08~0.15')
    axes1.scatter(xb,yb,color = 'blue',s= 5,alpha=0.6,label='<0.08') 

    plt.legend(loc='upper right')
    axes1.set_xticks(X,xtik)

    # ------colorbar-------
    level =[0,1,2,3,4,5,6,7,8,9,10,11]
    nlevel = len(level)
    cmap1 = mpl.colors.ListedColormap(['purple','blue','deepskyblue','turquoise','mediumturquoise','chartreuse','greenyellow','yellow','gold','orange','orangered','fuchsia'],N=12)
    cmap1.set_over('fuchsia')
    norm1 = mcolors.Normalize(vmin=0, vmax=11)
    norm1 = mcolors.BoundaryNorm(level, nlevel,extend='max')
    im = cm.ScalarMappable(norm=norm1, cmap=cmap1)
    cbar1 = fig1.colorbar(im,extend='max', ticks=level)
    # ------colorbar-------

    plt.axvline(x = 8,c = "r" , ls = "--" , lw = 1)
    plt.axvline(x = 20,c = "r" , ls = "--" , lw = 1)
    plt.title (day)
    


    fig2,axes2 = plt.subplots(figsize=(19,9))
    axes2.barbs(X,Y,U,V,color = Colors)

    axes2.set_xticks(X,xtik)

    # ------colorbar-------
    level =[0,1,2,3,4,5,6,7,8,9,10,11]
    nlevel = len(level)
    cmap1 = mpl.colors.ListedColormap(['purple','blue','deepskyblue','turquoise','mediumturquoise','chartreuse','greenyellow','yellow','gold','orange','orangered','fuchsia'],N=12)
    cmap1.set_over('fuchsia')
    norm1 = mcolors.Normalize(vmin=0, vmax=11)
    norm1 = mcolors.BoundaryNorm(level, nlevel,extend='max')
    im = cm.ScalarMappable(norm=norm1, cmap=cmap1)
    cbar2 = fig2.colorbar(im,extend='max', ticks=level)
    # ------colorbar-------

    plt.axvline(x = 8,c = "r" , ls = "--" , lw = 1)
    plt.axvline(x = 20,c = "r" , ls = "--" , lw = 1)    
    plt.title (day)





    # ------風切累計長條圖-----
    fig3,ax3 = plt.subplots(figsize=(19,9))
    plt.bar(ti,countb, color=['blue'],width=0.3,label='<0.08 count')
    plt.bar(ti,countg,bottom=np.array(countb), color=['green'],width=0.3,label='0.08~0.15 count')
    plt.bar(ti,county,bottom=np.array(countg)+np.array(countb), color=['yellow'],width=0.3,label='0.15~0.2 count')
    plt.bar(ti,countr,bottom=np.array(countg)+np.array(countb)+np.array(county), color=['red'],width=0.3,label='>0.2 count')
    ax3.set_xticks(showti) #刻度
    plt.xticks(rotation = 75)
    plt.legend(loc='upper right')


    fig4,ax4 = plt.subplots(figsize=(19,9))
    plt.bar(ti,countg, color=['green'],width=0.3,label='0.08~0.15 count')
    plt.bar(ti,county,bottom=np.array(countg), color=['yellow'],width=0.3,label='0.15~0.2 count')
    plt.bar(ti,countr,bottom=np.array(countg)+np.array(county), color=['red'],width=0.3,label='>0.2 count')
    ax4.set_xticks(showti) #刻度
    plt.xticks(rotation = 75)
    plt.legend(loc='upper right')


    fig5,ax5 = plt.subplots(figsize=(19,9))
    plt.bar(ti,county, color=['yellow'],width=0.3,label='0.15~0.2 count')
    plt.bar(ti,countr,bottom=np.array(county), color=['red'],width=0.3,label='>0.2 count')
    ax5.set_xticks(showti) #刻度
    plt.xticks(rotation = 75)
    plt.legend(loc='upper right')
    # ------風切累計長條圖-----



    try:
        os.makedirs("C:/Users/steve/Desktop/python相關資料/need data information/" + locate + '/' + year + "/picture/"+day)
    except:pass
    fig1.savefig("C:/Users/steve/Desktop/python相關資料/need data information/" + locate + '/' + year + "/picture/"+day+"/風標圖+散布圖.png")
    fig2.savefig("C:/Users/steve/Desktop/python相關資料/need data information/" + locate + '/' + year + "/picture/"+day+"/風標圖.png")
    fig3.savefig("C:/Users/steve/Desktop/python相關資料/need data information/" + locate + '/' + year + "/picture/"+day+"/長條圖.png")
    fig4.savefig("C:/Users/steve/Desktop/python相關資料/need data information/" + locate + '/' + year + "/picture/"+day+"/長條圖缺藍.png")
    fig5.savefig("C:/Users/steve/Desktop/python相關資料/need data information/" + locate + '/' + year + "/picture/"+day+"/長條圖缺蘭缺綠.png")
    print(day)
    # except:
        # print(day+'ERROR')
endTime = T.time()
print('程式執行了%s秒' %(endTime-startTime))

# plt.show()