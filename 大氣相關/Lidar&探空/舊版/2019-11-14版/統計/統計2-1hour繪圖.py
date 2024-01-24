import matplotlib.pyplot as plt
import math
from openpyxl import load_workbook
import matplotlib as mpl

#風標圖

wbu = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計2/u_mean1hour.xlsx")
wsu = wbu['Sheet']
wbv = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計2/v_mean1hour.xlsx")
wsv = wbv['Sheet']
colmax = wsu.max_column #24
rowmax = wsu.max_row #62
# print (wsu.max_column-1,wsu.max_row-1)
Time = []
for i in range(1,24):
    i = str(i)
    i = i.zfill(2)
    Time.append(i)
X = [] #time
Y = [] #hight
U = []
V = []
for col in range(1,colmax):
    X.append(col)
# print (X)
for row in range(100,1601,25):
    Y.append(row)
# print (Y)
U = [[]for i in range(rowmax-1)]
# print(len(U))
for j in range(2,rowmax+1):
    
    for i in range(2,colmax+1):
        U[j-2].append(wsu.cell(j,i).value)
# print (U)
V = [[]for i in range(rowmax-1)]
for j in range(2,rowmax+1):
    
    for i in range(2,colmax+1):
        V[j-2].append(wsv.cell(j,i).value)

ax = plt.subplot(1,1,1)

# Colors =['r','g','b','m','y']
Colors = []
for j in range(0,rowmax-1):
    for i in range(0,colmax-1):
        # j = 60-j    

        u = U[j][i]
        v = V[j][i]
        if u <0 and v<0:
            Colors.append('r')
        elif u<0 and v>=0:
            Colors.append('g')
        elif u>=0 and v<0:
            Colors.append('b')
        elif u>=0 and v>=0:
            Colors.append('c')
# plt.rcParams['font.sans-serif'] = [u'MingLiu'] #細明體
# plt.rcParams['axes.unicode_minus'] = False 
ax.barbs(X,Y,U,V,color = Colors)
# print(len(Colors))
ax.set_xticks(X,Time)
# print (Colors)
plt.legend()
plt.show()