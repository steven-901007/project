from openpyxl import load_workbook
import statistics
wb = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/u_all.xlsx")
ws = wb['Sheet']
wb_mean = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計2/u_mean10min.xlsx")
ws_mean = wb_mean['Mean']
ws_sigma = wb_mean['Sigma']
# sigma_title = [[]for i in range(61)]
for row in range(2,63):
    Mean = []
    sigmanb = []
    for col in range(6,1427,10):
        chunk = []
        chunk_ms = []
        for i in range(col,col+10):
            if ws.cell(row,i).value != None:
                chunk.append(ws.cell(row,i).value)
                chunk_ms.append(ws.cell(row,i).value)
            else :
                chunk.append(ws.cell(row,i).value)
        sigma = statistics.pstdev(chunk_ms)
        mean = round(statistics.mean(chunk_ms),2)
        Mean.append(mean)
        # print (chunk,chunk_ms,mean,sigma)
        n = 0
        for i in range(0,10):
            nb = chunk[i]
            if nb != None:
                if nb <=mean+1.5*sigma and nb >= mean-1.5*sigma:
                    n+=1
        sigmanb.append(n)
    # sigma_title[row-2] = sigmanb
    for i in range(0,143):
        ws_mean.cell(row,i+2).value = Mean[i]
        ws_sigma.cell(row,i+2).value = sigmanb[i]

wb_mean.save("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計2/u_mean10min.xlsx")
# print (sigma_title)

wb = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/v_all.xlsx")
ws = wb['Sheet']
wb_mean = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計2/v_mean10min.xlsx")
ws_mean = wb_mean['Mean']
ws_sigma = wb_mean['Sigma']
# sigma_title = [[]for i in range(61)]
for row in range(2,63):
    Mean = []
    sigmanb = []
    for col in range(6,1427,10):
        chunk = []
        chunk_ms = []
        for i in range(col,col+10):
            if ws.cell(row,i).value != None:
                chunk.append(ws.cell(row,i).value)
                chunk_ms.append(ws.cell(row,i).value)
            else :
                chunk.append(ws.cell(row,i).value)
        sigma = statistics.pstdev(chunk_ms)
        mean = round(statistics.mean(chunk_ms),2)
        Mean.append(mean)
        # print (chunk,chunk_ms,mean,sigma)
        n = 0
        for i in range(0,10):
            nb = chunk[i]
            if nb != None:
                if nb <=mean+1.5*sigma and nb >= mean-1.5*sigma:
                    n+=1
        sigmanb.append(n)
    # sigma_title[row-2] = sigmanb
    for i in range(0,143):
        ws_mean.cell(row,i+2).value = Mean[i]
        ws_sigma.cell(row,i+2).value = sigmanb[i]

wb_mean.save("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計2/v_mean10min.xlsx")
# print (sigma_title)