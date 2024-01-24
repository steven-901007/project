from openpyxl import load_workbook,Workbook
import math
import statistics
import time as Ti



locate = '松山'
year = '108'
day = '2019-01-01'
confidence_level = 100
sigma = 1.5

startTime = Ti.time()

wb = load_workbook("C:/Users/steve/Desktop/python相關資料/need data information/"+day+"/confidence_level="+str(confidence_level)+" need_information_set.xlsx")
wsv = wb['V']
wsu = wb['U']
wsmv = wb['mv']
wsmd = wb['md']
colmax = wsu.max_column #時間
rowmax = wsu.max_row #高度
# # print(colmax,rowmax)

T = [[],[]] #T[0]每10分鐘的起始時間 ， T[1]每10分鐘的結束時間

for hour in range(0,24):
    for minute in range(6,57,10):
        hour = str(hour).zfill(2)
        minute = str(minute).zfill(2)
        t = hour + ':' +minute
        # print(time)
        T[0].append(t)
    for minute in range(5,56,10):
        hour = str(hour).zfill(2)
        minute = str(minute).zfill(2)
        t = hour + ':' +minute
        # print(time)
        T[1].append(t)
T[0].pop()
T[1].pop(0)
# print(T[0])
# print(T[1])
# print(len(T[0]),len(T[1]))

time = []
for i in range(2,colmax+1):
    time.append(wsu.cell(1,i).value[:5])
    # print(wsu.cell(1,i).value[:5])
# print(len(time))

Tlc = [[],[]] #始末時間列表
for i in range(0,len(T[0])):
    # print(T[0][i])
    for j in range(len(time)):
        tg = time[j] 
        if tg == T[0][i]:
            Tlc[0].append(j+2)
            break
    for j in range(len(time)):
        tg = time[j]
        if tg == T[1][i]:
            Tlc[1].append(j+1+time.count(tg))
            break
# print(Tlc[0]) #起始時間座標
# print(Tlc[1]) #結束時間座標

wb_10min = Workbook()
wsu_10min = wb_10min.active
wsu_10min.title = 'U'
wb_10min.create_sheet('V')
wsv_10min = wb_10min['V']
wsmv_10min = wb_10min.create_sheet('mv')
wsmd_10min = wb_10min.create_sheet('md')
t = 2
for h in range(0,24): #時間
    for m in range(0,51,10):
        if h != 0 or m !=0:
            h = str(h).zfill(2)
            m = str(m).zfill(2)
            wsu_10min.cell(1,t).value = h + ":" + m
            wsv_10min.cell(1,t).value = h + ":" + m
            wsmv_10min.cell(1,t).value = h + ":" + m
            wsmd_10min.cell(1,t).value = h + ":" + m 
            t +=1

for j in range(2,rowmax+1): #高度
    wsu_10min.cell(j,1).value = wsu.cell(j,1).value
    wsv_10min.cell(j,1).value = wsv.cell(j,1).value
    wsmv_10min.cell(j,1).value = wsmv.cell(j,1).value
    wsmd_10min.cell(j,1).value = wsmd.cell(j,1).value


time10 = 2
for i in range(len(Tlc[0])):

    for k in range(2,rowmax+1):
        # i = 0
        s = Tlc[0][i]
        e = Tlc[1][i]
        cu = []
        cv = []
        cmv = []
        cmd = []
        try:
            for j in range(s,e+1):
                ui = wsu.cell(k,j).value
                vi = wsv.cell(k,j).value
                mvi = wsmv.cell(k,j).value
                mdi = wsmd.cell(k,j).value
                if ui != None:
                    cu.append(ui)
                if vi != None:
                    cv.append(vi)   
                if mvi !=None:
                    cmv.append(mvi)
                if mdi != None:
                    cmd.append(mdi)  
            # print(c)
            meanu = statistics.mean(cu)
            sigmau = statistics.pstdev(cu)
            meanv = statistics.mean(cv)
            sigmav = statistics.pstdev(cv)
            meanmv = statistics.mean(cmv)
            sigmamv = statistics.pstdev(cmv)
            meanmd = statistics.mean(cmd)
            sigmamd = statistics.pstdev(cmd)
            for l in range(len(cu)):
                try:
                    tgu = cu[l]
                    tgv = cv[l]
                    tgmv = cmv[l]
                    tgmd = cmd[l]
                    if tgu >= meanu+sigma*sigmau or tgu <= meanu-sigma*sigmau:
                        cu.remove(tgu)
                    if tgv >= meanv+sigma*sigmav or tgv <= meanv-sigma*sigmav:
                        cv.remove(tgv)
                    if tgmv >= meanmv+1.5*sigmamv or tgmv <= meanmv-1.5*sigmamv:
                        cmv.remove(tgmv)
                    if tgmd >= meanmd+1.5*sigmamd or tgmd <= meanmd-1.5*sigmamd:
                        cmd.remove(tgmd)
                except: 
                    pass
            wsu_10min.cell(k,time10).value = round(statistics.mean(cu),2)
            wsv_10min.cell(k,time10).value = round(statistics.mean(cv),2)
            wsmv_10min.cell(k,time10).value = round(statistics.mean(cmv),2)
            wsmd_10min.cell(k,time10).value = round(statistics.mean(cmd),2)
        except:
            wsu_10min.cell(k,time10).value = None
            wsv_10min.cell(k,time10).value = None
            wsmv_10min.cell(k,time10).value = None
            wsmd_10min.cell(k,time10).value = None
    time10 += 1


wb_10min.save("C:/Users/steve/Desktop/python相關資料/need data information/"+day+"/confidence_level="+str(confidence_level)+"sigma="+str(sigma)+" 10min.xlsx")
print(day+'\nconfidence '+str(confidence_level)+'\nsigma '+str(sigma))


endTime = Ti.time()

print('程式執行了%s秒' %(endTime-startTime))