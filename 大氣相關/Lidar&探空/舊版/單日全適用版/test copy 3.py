from openpyxl import Workbook ,load_workbook
from glob import glob
import math
import numpy as np
import time as T
import os


day = '2019-01-01' #時間
locate = '松山'
confidence_level = 100


start_time = T.time()

def fileset(path):    #建立資料夾
    if not os.path.exists(path):
        os.makedirs(path)
        print(path + " 已建立") 
    else:
        print(path + " 建立過了")  

# fileset("C:/Users/steve/python_data/lidar&sonding/one_day_version/"+locate+"/"+day+"/need_data")


#檢查資料高度的完整性
hight_error_chack = []
files = glob("C:/Users/steve/python_data/lidar&sonding/one_day_version/"+locate+"/"+day+"/need_data/**")
for file in files:
    # print(file)
    time = os.path.basename(file)
    wb = load_workbook(file)
    sheet_list_name = ['N','E','S','W']
    for name in sheet_list_name:
        ws = wb[name]

        for cell in range(2,ws.max_row):
            cell_a = int(ws.cell(cell,3).value)
            cell_b = int(ws.cell(cell+1,3).value)
            if cell_b-cell_a != 25:
                hight_error_chack.append(time+"between "+cell_a+" and "+cell_b)
                break
if len(hight_error_chack) !=0:
    print(hight_error_chack)
else:
    print(day+' is good')

