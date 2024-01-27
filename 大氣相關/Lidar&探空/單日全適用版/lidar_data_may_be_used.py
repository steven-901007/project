from openpyxl import Workbook ,load_workbook
from glob import glob
import time as T
import os
import pandas as pd
import math
import numpy as np

day = '2019-01-01' #時間
locate = '松山'


start_time = T.time()




def fileset(path):    #建立資料夾
    if not os.path.exists(path):
        os.makedirs(path)
        print(path + " 已建立") 
    else:
        print(path + " 建立過了")  


#建立存放檔案位置
fileset("C:/Users/steve/python_data/lidar&sonding/one_day_version/"+locate+"/"+day+"/need_data")

year = str(int(day[:4])-1911)
month = day[5:7]
# print(month)


#sheet title and 會用到的資料title
direction = ['N','E','S','W']  #0,1,2,3
sheet_list_name = ['Altitude [m]',                  #高度
                'Radial Wind Speed [m/s]',       #radial wind speed
                'Confidence Index [%]']          #confidence level
sheet_title_name = ['Altitude','U','V','confidence']


#讀取每天所有的資料
rawdata_day_path = "C:/Users/steve/python_data/lidar&sonding/lidar_raw_data/"+locate+"/"+year+"/"+year+month+"/"+day+"/wind_reconstruction_data/**"
hours = glob(rawdata_day_path)
for hour in hours:
    # print(hour)


    #建立excel
    new_wb = Workbook()
    new_wb.remove(new_wb['Sheet'])
    data_location = 1
    for list in sheet_title_name:   
        # print(list)
        new_wb.create_sheet(list)


    minutes = glob(hour+'/**')
    for minute in minutes:
        # print(minute)

        file_size = os.path.getsize(minute) #確認檔案大小是否=0(空檔案)
        if file_size !=0:
            #讀取舊資料
            old_data = pd.read_csv(minute,encoding = 'big5',delimiter=';')
            

            time = old_data['Timestamp'][1] #確定當前資料時間
            time = time[11:19].replace(":","_")
            # print(time)
            for list in range(len(sheet_title_name)):             
                ws = new_wb[sheet_title_name[list]]

                ws.cell(1,data_location).value = time
                #data set
                            
                math.cos(75*np.pi/180)
                if sheet_title_name[list] == 'U':   # u = (W-E)/2cos(75度)
                    E = old_data[(round(old_data['Elevation [財'] )== 75) & (round(old_data['Azimuth [財']) == 90)]['Radial Wind Speed [m/s]']
                    W = old_data[(round(old_data['Elevation [財'] )== 75) & (round(old_data['Azimuth [財']) == 270)]['Radial Wind Speed [m/s]']
                    # print(E)
                    for inf in range(len(infs)):    #data的長度(高度)
                        new_wb[sheet_title_name[list]].cell(inf+2,data_location).value = (W-E)/(2*math.cos(75*np.pi/180))
                elif sheet_title_name[list] == 'V':    # v = (S-N)/2cos(75度)
                    S = old_data[(round(old_data['Elevation [財'] )== 75) & (round(old_data['Azimuth [財']) == 180)]['Radial Wind Speed [m/s]']
                    N = old_data[(round(old_data['Elevation [財'] )== 75) & ((round(old_data['Azimuth [財']) == 0)| (round(old_data['Azimuth [財']) == 360))]['Radial Wind Speed [m/s]']
                    for inf in range(len(infs)):    #data的長度(高度)
                        new_wb[sheet_title_name[list]].cell(inf+2,data_location).value = (S-N)/(2*math.cos(75*np.pi/180))           
                else:
                    infs = old_data[round(old_data['Elevation [財'] )== 75 ].groupby(['Altitude [m]'])[sheet_list_name[list]].sum()
                    for inf in range(len(infs)):    #data的長度(高度)
                        new_wb[sheet_title_name[list]].cell(inf+2,data_location).value = infs.iloc[inf]/4



                
            data_location += 1


            wb_name = os.path.basename(hour)    #檔案名稱
            # print(wb_name)                  
    # new_wb.save("C:/Users/steve/python_data/lidar&sonding/one_day_version/"+locate+"/"+day+"/need_data/"+wb_name+".xlsx")
    new_wb.close()
    print(wb_name)

# end_time = T.time()
# print('程式執行了%s秒' %(end_time-start_time))