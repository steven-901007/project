import numpy as np
from openpyxl import Workbook
import glob

day = '2019-02-08_12'

file = "C:/Users/steve/Desktop/python相關資料/觀測rowdata/NAS_B_obs/TW_SOUNDING/2019/46692/"+day[5:7]+"/46692-"+day[0:4]+day[5:7]+day[8:10]+day[11:13]+".shr.txt"
hight = []
v = []
d = []
time = []
with open(file) as f:
    for line in f.readlines():
        s = line.split(',\t')
        try:
            v.append(float(s[7]))
            d.append(float(s[6]))
            hight.append(float(s[2]))
            time.append(str(s[0]))
        except:pass

# print(hight)
# print(v)
# print(d)
# print(len(hight),len(v),len(d))


wb = Workbook()
wsd = wb.active
wsd.title = '風向'
wb.create_sheet('風速')
wsv = wb['風速']

wsd.cell(1,1).value = '高度'
wsd.cell(1,2).value = '風向'
wsd.cell(1,3).value = 'Time'
wsv.cell(1,1).value = '高度'
wsv.cell(1,2).value = '風速'
wsv.cell(1,3).value = 'Time'

lc = 2
for i in range(len(hight)):

    wsd.cell(lc,1).value = float(hight[i])
    wsd.cell(lc,2).value = float(d[i])
    wsd.cell(lc,3).value = str(time[i])
    wsv.cell(lc,1).value = float(hight[i])
    wsv.cell(lc,2).value = float(v[i])
    wsv.cell(lc,3).value = str(time[i])
    lc += 1

wb.save('C:/Users/steve/Desktop/python相關資料/觀測rowdata/need_information/' + day + '.xlsx')

print(day)