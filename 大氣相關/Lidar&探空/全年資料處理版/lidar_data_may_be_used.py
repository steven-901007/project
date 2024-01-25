from openpyxl import Workbook ,load_workbook
from glob import glob
import time as T
import os
import pandas as pd


locate = '松山'
year = '108'


start_time = T.time()


def fileset(path):    #建立資料夾
    if not os.path.exists(path):
        os.makedirs(path)
        print(path + " 已建立") 
    else:
        print(path + " 建立過了")  


direction = ['N','E','S','W']  #0,1,2,3
sheet_list_name = ['Altitude [m]',                  #高度
                'Radial Wind Speed [m/s]',       #radial wind speed
                'Confidence Index [%]']          #confidence level
sheet_title_name = ['Altitude','radial speed','confidence']


year = "C:/Users/steve/python_data/lidar&sonding/lidar_raw_data/"+locate+"/"+year+"/**"
months = glob(year)
for long_name_month in months:
    # print(long_name_month)
    

    #可以分段建立檔案(若不需使用則將start_month設為1)
    month = int(long_name_month[len(long_name_month)-2:])
    # print(month)
    start_month = 3
    if month >= start_month:
        days = glob(long_name_month+'/**')
        for long_name_day in days:
            day = os.path.basename(long_name_day)
            # print(day)
            # print(long_name_day)
            fileset("C:/Users/steve/python_data/lidar&sonding/one_year_version/"+locate+"/"+day+"/need_data/")

            #讀取每天所有的資料
            hours = glob(long_name_day+'/wind_reconstruction_data/**')
            for hour in hours:
                # print(hour)

            

                #建立excel
                new_wb = Workbook()
                new_wb.remove(new_wb['Sheet'])
                data_location = 1
                for list in sheet_title_name:   
                    # print(list)
                    new_wb.create_sheet(list)


                minutes = glob(hour+'/**')
                for minute in minutes:
                    # print(minute)

                    file_size = os.path.getsize(minute) #確認檔案大小是否=0(空檔案)
                    if file_size !=0:
                        #讀取舊資料
                        old_data = pd.read_csv(minute,encoding = 'big5',delimiter=';')
                        

                        time = old_data['Timestamp'][1] #確定當前資料時間
                        time = time[11:19].replace(":","_")
                        # print(time)
                        for list in range(len(sheet_title_name)):             
                            ws = new_wb[sheet_title_name[list]]

                            ws.cell(1,data_location).value = time
                            #data set
                                        
                            infs = old_data[round(old_data['Elevation [財'] )== 75].groupby(['Altitude [m]'])[sheet_list_name[list]].sum()
                        
                            # print(infs)
                            for inf in range(len(infs)):    #data的長度(高度)
                                new_wb[sheet_title_name[list]].cell(inf+2,data_location).value = infs.iloc[inf]/4



                            
                        data_location += 1


                        wb_name = os.path.basename(hour)    #檔案名稱
                        # print(wb_name)                        
                new_wb.save("C:/Users/steve/python_data/lidar&sonding/one_year_version/"+locate+"/lidar/"+day+"/need_data/"+wb_name+".xlsx")
                new_wb.close()
                print(wb_name)


        
end_time = T.time()
print('程式執行了%s秒' %(end_time-start_time))