from openpyxl import Workbook,load_workbook
from glob import glob
import time as T
import os
import pandas as pd


year = '2019'


start_time = T.time()

def fileset(path):    #建立資料夾
    if not os.path.exists(path):
        os.makedirs(path)
        print(path + " 已建立") 
    else:
        print(path + " 建立過了")  

def file_concat(path_a,path_b): #合併檔案(若檔案被誤分成兩分)

    path_a =path_a
    path_b =path_b
    data_a =pd.read_csv(path_a,header=None)
    data_b = pd.read_csv(path_b,header=None)
    data_a_plus_b = pd.concat([data_a,data_b],axis=0,ignore_index=True)
    # print(data_a_plus_b)
    data_a_plus_b.to_csv(path_b)
    os.remove(path_a)

def time_chack(file_a,file_b):  #確認時間差最短的檔案,以及資料的位置cell(1,location)
    wb_a = load_workbook(file_a)
    wb_b = load_workbook(file_b)
    ws_a = wb_a['Altitude']
    ws_b = wb_b['Altitude']
    time_a = ws_a.cell(1,ws_a.max_column).value
    time_a = int(time_a[len(time_a)-2:])
    time_b = ws_b.cell(1,1).value
    time_b = int(time_b[len(time_b)-2:])
    # print(time_a,time_b)
    if 60-time_a>time_b:
        # print(file_b)
        return file_b ,1
    else:
        # print(file_a)
        return file_a ,ws_a.max_column
    

wb = Workbook()
ws_altitude = wb.active
ws_altitude.title = ' Altitude'
ws_sonding = wb.create_sheet('sonding')
ws_lidar = wb.create_sheet('lidar')

# sonding
sonding_year_path = "C:/Users/steve/python_data/lidar&sonding/sonding_raw_data/BANQIAO_Sounding/"+year+"/**"
sonding_months = glob(sonding_year_path)
for sonding_long_name_month in sonding_months:
    month = os.path.basename(sonding_long_name_month)
    # print(sonding_month)
    sonding_days = glob(sonding_long_name_month+'/**.txt')
    for sonding_long_name_day in sonding_days:
        day = os.path.basename(sonding_long_name_day)[12:14]
        time = os.path.basename(sonding_long_name_day)[14:16]
        
        #合併同一時間的資料(若資料被分成兩分)數字會因為檔案位置及檔案名稱而更改
        if len(sonding_long_name_day) != 107:
            tg_path = sonding_long_name_day[:99]+sonding_long_name_day[101:]
            file_concat(sonding_long_name_day,tg_path)

        # print(year,month,day,time)
        print(sonding_long_name_day)
        

        # title = ['時間','氣壓','高度','溫度','濕度','濕球溫度','風向','風速']
        # sonding = pd.read_csv(long_name_day,names = title)
        # print(sonding)


        # lidar
        lidars_path = glob("C:/Users/steve/python_data/lidar&sonding/one_year_version/松山/lidar/**")
        for long_name_lidar_path in lidars_path:
            # print(lidar_path)
            lidars_path = os.path.basename(long_name_lidar_path)
            if lidars_path == year+'-'+month+'-'+day:
                hours = glob(long_name_lidar_path+'/need_data/**')
                for long_name_hour in range(len(hours)):
                    # print(hour)
                    hour = os.path.basename(hours[long_name_hour])
                    long_name_hour_a = hours[long_name_hour-1]
                    if str(int(hour[:2])-8).zfill(2) == time:   #調整國際時間和台灣標準時間的差異
                        long_name_hour_b = hours[long_name_hour]
                        # print(long_name_hour_a)
                        # print(long_name_hour_b)
                        file,cell_location = time_chack(long_name_hour_a,long_name_hour_b)  #確認哪個時間最接近目標時間
                        print(file,cell_location)


                        #所需資料設定
                            #資料位置
                        sonding_file = sonding_long_name_day    #探空資料file path
                        lidar_file = file   #lidar資料file path
                            #wb,ws設定
                        lidar_location = cell_location    #lidar資料的cell位置ws.cell(1,lidar_location)
                        new_file_data_time = 1  #建立data的行位置

                        



                        



# lidar_path = "C:/Users/steve/python_data/lidar&sonding/one_year_version/松山/2019-01-01/need_data/"+time+"-00.xlsx"
        


wb.save("C:/Users/steve/python_data/lidar&sonding/one_year_version/松山/sonding_and_lidar.xlsx")