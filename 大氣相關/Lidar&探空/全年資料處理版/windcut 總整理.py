from openpyxl import load_workbook
import glob
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.ticker as ticker
import matplotlib.dates as mdates


panumber = 0.1
locate = '松山'
year = '108'


plt.rcParams['font.sans-serif'] = [u'MingLiu'] #細明體
plt.rcParams['axes.unicode_minus'] = False #設定中文

windcut = []
timelineday = [[],[]] #[0] = 日期 [1] = 對應的count
files = glob.glob("C:/Users/steve/Desktop/python相關資料/need data information/" + locate + "/" + year + "/windcut/**")
for file in files:
    day = file[71:81]
    timelineday[0].append(day[5:])
    timelineday[1].append(0)
    wb = load_workbook("C:/Users/steve/Desktop/python相關資料/need data information/" + locate + "/" + year + "/windcut/" + day + ".xlsx")
    ws = wb['風切']
    rowmax = ws.max_row
    colmax = ws.max_column

    for row in range(1,rowmax+1):
        for col in range(1,colmax+1):
            if ws.cell(row,col).value != None:
                windcut.append(ws.cell(row,col).value)


# nb = []
# for i in range(len(windcut)):
#     nb.append(i)
# windcut.sort()
# print(windcut)
windcut.reverse()

# plt.plot(nb,windcut,color =  'black',linestyle = '--')
# plt.show()

windcut_count = len(windcut)
max10min = windcut[round(windcut_count*panumber)]

# print(max10min)
# print(timeline[0])
# print(timeline[1])

for t in range(len(timelineday[0])):
    day = timelineday[0][t]
    wb = load_workbook("C:/Users/steve/Desktop/python相關資料/need data information/" + locate + "/" + year + "/windcut/"+str(int(year)+1911)+"-" + day + ".xlsx")
    ws = wb['風切']
    rowmax = ws.max_row
    colmax = ws.max_column

    for row in range(1,rowmax+1):
        for col in range(1,colmax+1):
            if ws.cell(row,col).value != None:
                if ws.cell(row,col).value >= max10min:
                    timelineday[1][t] += 1
# print(timelineday[0])
# print(timelineday[1])
print(max(timelineday[1]))
# print(len(timelineday[1]))
#  ----------------------每日數據繪圖--------------------------

x = np.arange(len(timelineday[0]))
ax = plt.subplot()
ax.plot(x,timelineday[1],color =  'black',linestyle = '--')
ax.set_xticks(x,timelineday[0])
ax.xaxis.set_major_locator(ticker.MaxNLocator(12))

ax.set_title(str(panumber) + '(資料最後一筆為' + timelineday[0][len(timelineday[0])-1] + ')')
plt.show()

# ------------------------------------------------------------



# timelinem = [[],[]]
# for i in range(1,13):
#     i = str(i).zfill(2)
#     timelinem[0].append(i)
#     timelinem[1].append(0)
# # print(timelinem)

# for j in range(len(timelinem[0])):
#     for i in range(len(timelineday[0])):
#         # print(timelineday[0][i][:2])
#         if timelinem[0][j] == timelineday[0][i][:2]:
#             timelinem[1][j] =  timelinem[1][j] + int(timelineday[1][i])

# print(timelinem[0])
# print(timelinem[1])

#  ----------------------每月數據繪圖--------------------------

# x = np.arange(len(timelinem[0]))
# plt.plot(x,timelinem[1],color =  'black',linestyle = '--')
# plt.xticks(x,timelinem[0])
# plt.xticks(rotation = 75) #旋轉字體角度 ('vertical'=垂直)
# plt.title(str(panumber) + '(everymoon statistics)')
# plt.show()

# ------------------------------------------------------------


