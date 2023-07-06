from openpyxl import load_workbook
import statistics
import numpy as np  

#先去除最大值最小值，在做mean以及sigma、把大於1.5標準差的值去掉

wb = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/u_all.xlsx")
ws = wb['Sheet']

for row in range(2,63):

    for col in range(6,1427,10):
        chunk = []
        chunk_ms = []
        for i in range(col,col+10):
            if ws.cell(row,i).value != None:
                chunk.append(ws.cell(row,i).value)
                chunk_ms.append(ws.cell(row,i).value)
            else :
                chunk.append(ws.cell(row,i).value)
        # print(chunk)

        Max = max(chunk_ms)
        Min = min(chunk_ms)
        chunk_ms.remove(Max)
        chunk_ms.remove(Min)
        # print(chunk_ms)
        sigma = statistics.pstdev(chunk_ms)
        mean = statistics.mean(chunk_ms)
        # print (len(chunk))
        # print (len(chunk_ms))

        for i in range(0,10):
            nb = chunk[i]
            if nb != None:
                if nb >=mean+1.5*sigma or nb <= mean-1.5*sigma:
                    chunk[i] = np.nan
        # print (chunk)
        n = 0
        for i in range(col,col+10):
            ws.cell(row,i).value = chunk[n]
            n +=1


    wb.save("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計1/u_all統計1.xlsx")
        # print('mean = '+str(mean))
        # print ('sigma = '+ str(sigma))


wb = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/v_all.xlsx")
ws = wb['Sheet']

for row in range(2,63):

    for col in range(6,1427,10):
        chunk = []
        chunk_ms = []
        for i in range(col,col+10):
            if ws.cell(row,i).value != None:
                chunk.append(ws.cell(row,i).value)
                chunk_ms.append(ws.cell(row,i).value)
            else :
                chunk.append(ws.cell(row,i).value)


        Max = max(chunk_ms)
        Min = min(chunk_ms)
        chunk_ms.remove(Max)
        chunk_ms.remove(Min)

        sigma = statistics.pstdev(chunk_ms)
        mean = statistics.mean(chunk_ms)


        for i in range(0,10):
            nb = chunk[i]
            if nb != None:
                if nb >=mean+1.5*sigma or nb <= mean-1.5*sigma:
                    chunk[i] = np.nan

        n = 0
        for i in range(col,col+10):
            ws.cell(row,i).value = chunk[n]
            n +=1


    wb.save("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計1/v_all統計1.xlsx")