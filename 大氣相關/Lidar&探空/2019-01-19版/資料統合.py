import csv
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import random
import matplotlib
import glob
import numpy as np
import math
import os


time = 1
wball = Workbook()
wsall = wball.active

hight = 64#50~1600

files  =glob.glob("C:/Users/steve/Desktop/python/松山剖風儀資料/2019-01-19/u/*.xlsx")
for f in files:
    print ('u'+f[51:56])
    wb = load_workbook(f)
    ws = wb['Sheet']
# ws['A']
    for i in range(1,hight):
        wsall.cell(i+1,time).value = ws.cell(i,1).value
    wsall.cell(1,time).value = f[51:56]
    time +=1
wsall.insert_cols(1)
for i in range (1,hight):
    wsall.cell(i+1,1).value = 50+(i-1)*25
wball.save('C:/Users/steve/Desktop/python/松山剖風儀資料/2019-01-19/u_all.xlsx')
 

time = 1
wball = Workbook()
wsall = wball.active

files  =glob.glob("C:/Users/steve/Desktop/python/松山剖風儀資料/2019-01-19/v/*.xlsx")
for f in files:
    print ('v'+f[51:56])
    wb = load_workbook(f)
    ws = wb['Sheet']
# ws['A']
    for i in range(1,hight):
        wsall.cell(i+1,time).value = ws.cell(i,1).value
    wsall.cell(1,time).value = f[51:56]
    time +=1
wsall.insert_cols(1)
for i in range (1,hight):
    wsall.cell(i+1,1).value = 50+(i-1)*25
wball.save('C:/Users/steve/Desktop/python/松山剖風儀資料/2019-01-19/v_all.xlsx')

time = 1
wball = Workbook()
wsall = wball.active

files  =glob.glob("C:/Users/steve/Desktop/python/松山剖風儀資料/2019-01-19/u/*.xlsx")
for f in files:
    print ('believe'+f[51:56])
    wb = load_workbook(f)
    ws = wb['Sheet']
# ws['A']
    for i in range(1,hight):
        wsall.cell(i+1,time).value = ws.cell(i,2).value
    wsall.cell(1,time).value = f[51:56]
    time +=1
wsall.insert_cols(1)
for i in range (1,hight):
    wsall.cell(i+1,1).value = 50+(i-1)*25
wball.save('C:/Users/steve/Desktop/python/松山剖風儀資料/2019-01-19/believe_all.xlsx')