import matplotlib.pyplot as plt
import math
from openpyxl import load_workbook
import matplotlib as mpl
import cmath
from math import *
#風標圖

wb = load_workbook('C:/Users/steve/Desktop/python/松山剖風儀資料/2019-01-19/統計/1hour.xlsx')
wsu = wb['u']
wsv = wb['v']
colmax = wsu.max_column #24
rowmax = wsu.max_row #64
# print (wsu.max_column-1,wsu.max_row-1)
Time = []
for i in range(1,24):
    i = str(i)
    i = i.zfill(2)
    Time.append(i)
X = [] #time
Y = [] #hight
U = []
V = []
for col in range(1,colmax):
    X.append(col)
# print (X)
for row in range(50,1601,25):
    Y.append(row)
# print (Y)
U = [[]for i in range(rowmax-1)]
# print(len(U))
for j in range(2,rowmax+1):
    
    for i in range(2,colmax+1):
        if wsu.cell(j,i).value == None:
            U[j-2].append(-999)
        else:
            U[j-2].append(wsu.cell(j,i).value)
# print (U)
V = [[]for i in range(rowmax-1)]
for j in range(2,rowmax+1):
    
    for i in range(2,colmax+1):
        if wsu.cell(j,i).value == None:
            U[j-2].append(-999)
        else:
            V[j-2].append(wsv.cell(j,i).value)

ax = plt.subplot(1,1,1)

# Z = []
Colors = []
for j in range(0,rowmax-1):
    for i in range(0,colmax-1):
        # j = 60-j    

        u = U[j][i]
        v = V[j][i]
        if u != -999 and v != -999:
            z = round(sqrt(pow(u,2) +pow(v,2)),2)
        else:
            z = -999
        # Z.append(z)
        if 0<=z<=1:
            Colors.append('purple')
        elif 1<z<=2:
            Colors.append('blue')
        elif 2<z<=3:
            Colors.append('deepskyblue')
        elif 3<z<=4:
            Colors.append('turquoise')
        elif 4<z<=5:
            Colors.append('mediumturquoise')
        elif 5<z<=6:
            Colors.append('chartreuse')
        elif 6<z<=7:
            Colors.append('greenyellow')
        elif 7<z<=8:
            Colors.append('yellow')
        elif 8<z<=9:
            Colors.append('gold')
        elif 9<z<=10:
            Colors.append('orange')
        elif 10<z<=11:
            Colors.append('orangered')
        elif 11<z:
            Colors.append('fuchsia')
        elif z == -999:
            Colors.append('w')
print(U[9])
for j in range(rowmax-1):
    
    for i in range(23):
        if U[j][i] == -999:
            U[j][i] = 0
        if V[j][i] == -999:
            V[j][i] = 0
print(U[9])

# plt.rcParams['font.sans-serif'] = [u'MingLiu'] #細明體
# plt.rcParams['axes.unicode_minus'] = False 
# print(Z)
ax.barbs(X,Y,U,V,color = Colors)
# print(len(Colors))
ax.set_xticks(X,Time)
# print (max(Colors),min(Colors))
# plt.legend()
plt.title ('2019-01-19')
plt.show()