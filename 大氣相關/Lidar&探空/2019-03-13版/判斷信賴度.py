from openpyxl import load_workbook 


wbblive = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/believe_all.xlsx")
wbu = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/u_all.xlsx")
wbv = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/v_all.xlsx")
wsblive=  wbblive['Sheet']
wsu=  wbu['Sheet']
wsv=  wbv['Sheet']

rowmax = wsblive.max_row #64
colmax = wsblive.max_column #4355
# print(rowmax,colmax)

for col in range(2,colmax+1):
    for row in range(2,rowmax+1):
        if wsblive.cell(row,col).value == 0:
            wsu.cell(row,col).value = None
            wsv.cell(row,col).value = None

wbu.save("C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/u_all.xlsx")
wbv.save("C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/v_all.xlsx")