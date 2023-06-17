from turtle import color
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np  
plt.rcParams['font.sans-serif'] = [u'MingLiu'] #設定字體為'細明體'
plt.rcParams['axes.unicode_minus'] = False #用來正常顯示正負號
f = 30
wb = load_workbook("C:/Users/steve/Desktop/學校/大三下/投資/期末報告.xlsx")
ws = wb['工作表1']

# X = []

# Y = []
# Y2 = []
# for i  in range(2,7):
#     Y.append(ws.cell(i,5).value)
#     X.append(ws.cell(i,1).value)
#     Y2.append(ws.cell(i,7).value)
# print(Y)
# ax = plt.subplots()
# plt.plot(X,Y,label = str(ws.cell(1,5).value),color =  'black',marker = "*",linestyle = '--')
# plt.plot(X,Y2,label = str(ws.cell(1,7).value),color =  'r',marker = "*",linestyle = '--')

# plt.title("原始報酬率VS預期報酬率",fontsize=f)
# plt.xlabel('各台股',labelpad=10,fontsize=f) 
# plt.xticks(fontsize=f)
# plt.ylabel('報酬率%數',fontsize=f)
# plt.yticks(rotation = 0,fontsize=f) 
# plt.legend(fontsize=20) #設定圖例
# plt.show()
l = 2

a = ['買入','賣出']
b = [ws.cell(l,2).value,ws.cell(l,3).value]
print(b)
plt.bar(np.arange(len(a)),b, color=['gold', 'orange'],width=0.2)
plt.xticks(np.arange(len(a)),a,fontsize=f)
plt.ylabel('股價',fontsize=f)
plt.yticks(fontsize=f)
plt.title(str(ws.cell(l,1).value),fontsize=f)
# plt.xticks(rotation = 75) #旋轉字體角度 ('vertical'=垂直)
plt.show() 
