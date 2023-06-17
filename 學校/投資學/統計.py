from openpyxl import load_workbook
import numpy as np
path = "C:/Users/steve/Desktop/學校/大三下/投資/統計.xlsx"

wb = load_workbook(path)
ws = wb['工作表1']

row = 14#高
col = 7 #寬
ws.cell(row+1,1).value = '平均數'
ws.cell(row+2,1).value = '標準差'
for i in range(2,col+1):
    # print(ws.cell(1,i).value)
    file = []
    for j in range(2,row):
        c = float(ws.cell(j,i).value)
        c_1 = float(ws.cell(j+1,i).value)
        # print(c)
        file.append(round((c - c_1)/c_1,5))
    # print(file)
    ws.cell(row+1,i).value = round(np.mean(file),5)
    ws.cell(row+2,i).value = round(np.std(file),5)
file = []   
for j in range(2,row+1):
    
    c = float(ws.cell(j,7).value) #月底
    c_1 = float(ws.cell(j,8).value) #月初
    # print(c)
    file.append(round((c - c_1)/c_1,5))
ws.cell(row+1,7).value = round(np.mean(file),5)
ws.cell(row+2,7).value = round(np.std(file),5)
# print(file)
# print(np.std(file))
wb.save(path)