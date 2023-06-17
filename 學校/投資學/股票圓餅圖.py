import matplotlib.pyplot as plt #繪製視覺化數據用
from openpyxl import load_workbook
wb = load_workbook("C:/Users/steven.LAPTOP-8A1BDJC6/OneDrive/桌面/筆電py/投資.xlsx")
ws = wb['工作表1']
maxrow = ws.max_row
# print(maxrow)
tick = []
x = []

for i in range(3,maxrow):
    if ws.cell(i,13).value != None:
        tick.append(str(ws.cell(i,1).value)+str(ws.cell(i,7).value))
        x.append(ws.cell(i,13).value)
# print(tick)
# print(x)

plt.rcParams['font.sans-serif'] = [u'MingLiu'] #設定字體為'細明體'
plt.rcParams['axes.unicode_minus'] = False #用來正常顯示正負號

#繪製圓餅圖
plt.pie(x, #數據
        labels = tick, #類別
        autopct="%1.0f%%", #數據顯示到小數點後第n位
        textprops = {"fontsize" : 10}, #字體大小
        pctdistance = 0.5, #數值位置
        shadow=True) #陰影

plt.legend() #設定圖例
plt.axis('equal') #使圓餅圖比例相等
plt.title('投資比例',loc = 'left') #設定標題
plt.show() #顯示繪製的圖
