import os
import calendar
from tqdm import tqdm
from openpyxl import load_workbook
import math
import pandas as pd 
from geopy.distance import geodesic


year = '2021' #年分
month = '06' #月份
data_top_path = "C:/Users/steve/python_data"
dis = 36


## 讀取雨量站經緯度資料
def rain_station_location_data():
    data_path = data_top_path+"/研究所/雨量資料/"+year+"測站範圍內測站數.xlsx"
    lon_data_list = []  # 經度
    lat_data_list = []  # 緯度
    name_data_list = []  #測站名稱
    wb = load_workbook(data_path)
    ws = wb[month]
    for i in range(ws.max_column):
        lon_data_list.append(ws.cell(3,i+1).value)
        lat_data_list.append(ws.cell(2,i+1).value)
        name_data_list.append(ws.cell(1,i+1).value)
    wb.close()
    return lon_data_list, lat_data_list ,name_data_list
lon_data_list, lat_data_list ,name_data_list = rain_station_location_data()


## 讀取閃電資料
flash_data_path = data_top_path+'/研究所/閃電資料/raw_data/'+year+'/'+year+month+'.txt'
flash_rawdata = pd.read_csv(flash_data_path,header = 0)
flash_rawdata['simple_time'] = flash_rawdata['日期時間'].str[:4] + flash_rawdata['日期時間'].str[5:7] + flash_rawdata['日期時間'].str[8:10] + flash_rawdata['日期時間'].str[11:13] + flash_rawdata['日期時間'].str[14:16]
# print(flash_rawdata['simple_time'],flash_rawdata['經度'],flash_rawdata['緯度'])






def fileset(path):    #建立資料夾
    if not os.path.exists(path):
        os.makedirs(path)
        print(path + " 已建立") 
    
fileset(data_top_path + "/研究所/閃電資料/依測站分類/"+str(dis)+'km/'+year+ '/' + month)


for station_nb in tqdm(range(len(name_data_list)),desc='寫入資料'):
    station_lon = lon_data_list[station_nb]
    station_lat = lat_data_list[station_nb]
    station_lat_lon = (station_lat,station_lon)
    # print(station_lat_lon)
    csv_file_path = data_top_path + "/研究所/閃電資料/依測站分類/"+str(dis)+'km/'+year+ '/' + month + '/' + year+month+'_'+str(dis)+'km_'+name_data_list[station_nb] + '.csv'

    flash_rawdata['distance_km'] = flash_rawdata.apply(lambda row: geodesic((row['緯度'], row['經度']), station_lat_lon).km, axis=1)
    # print(flash_rawdata['distance_km'])

    target_data = flash_rawdata[flash_rawdata['distance_km'] < dis]['simple_time']
    # print(target_data)
    flash_data_to_save = {
        'yyyymmddHHMM' : target_data,
    }
    pd.DataFrame(flash_data_to_save).to_csv(csv_file_path,index=False)