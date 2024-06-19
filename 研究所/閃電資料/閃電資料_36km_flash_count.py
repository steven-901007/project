from openpyxl import load_workbook,Workbook
from calendar import monthrange #判定一個月有幾天的import
import math
##記錄強降雨發生的地點


year = '2021' #年分
month = '06' #月份
data_top_path = "C:/Users/steve/python_data"

## 讀取雨量站經緯度資料
def rain_station_location_data():
    data_path = data_top_path+"/研究所/雨量資料/"+year+"測站範圍內測站數.xlsx"
    lon_data_list = []  # 經度
    lat_data_list = []  # 緯度
    name_data_list = []  #測站名稱
    wb = load_workbook(data_path)
    ws = wb['6月']
    for i in range(ws.max_column):
        lon_data_list.append(ws.cell(3,i+1).value)
        lat_data_list.append(ws.cell(2,i+1).value)
        name_data_list.append(ws.cell(1,i+1).value)
    wb.close()
    return lon_data_list, lat_data_list ,name_data_list

lon_data_list, lat_data_list ,name_data_list = rain_station_location_data()


## 兩經緯度距離
def haversine(lat1, lon1, lat2, lon2):
    # 地球的半徑（公里）
    R = 6371.0

    # 將經緯度從度轉換為弧度
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)

    # 經緯度差
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad

    # Haversine公式
    a = math.sin(dlat / 2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    # 計算距離
    distance = R * c

    return distance


## 與測站距離<36 km的有那些
dis = 36


## 儲存檔案基本設定(時間、測站)
wb_station_lighting_jump_count_data = Workbook()
ws_station_lighting_jump_count_data = wb_station_lighting_jump_count_data.active
ws_station_lighting_jump_count_data.title = month

max_month_day = monthrange(int(year),int(month))[1]
# print(max_month_day)
#時間序列
col = 2  #第一排是要填測站名稱
for dd in range(1,max_month_day+1):
    dd = str(dd).zfill(2)

    for hh in range(0,24):
        hh = str(hh).zfill(2)

        for mm in range(0,60,10):
            mm = str(mm).zfill(2)
            # print(yy+mm)
            title_time = dd+hh+mm #分隔的時間起始
            ws_station_lighting_jump_count_data.cell(1,col).value = title_time  ##ddHHMM
            col += 1

#測站序列
row = 2
for station in name_data_list:
    ws_station_lighting_jump_count_data.cell(row,1).value = station
    row += 1



#lighting jump count data填入
#light data讀取
wb_light_data_read = load_workbook(data_top_path+"/研究所/閃電資料/need_inf_for_lighting/"+year+"/"+year+"_"+month+"_flash_data.xlsx")
ws_light_data_read = wb_light_data_read[month]
max_light_data_read_col =ws_light_data_read.max_column


for light_data_col in range(1,max_light_data_read_col+1): #時間
    print(ws_station_lighting_jump_count_data.cell(1,light_data_col).value)
    light_data_row = 2
    while ws_light_data_read.cell(light_data_row,light_data_col).value != None:
        light_data = ws_light_data_read.cell(light_data_row,light_data_col).value #每個時間點的資料
        light_data_lon = int(light_data[:5])/100 #經
        light_data_lat = int(light_data[5:])/100 #緯

        for station in range(len(name_data_list)):
            station_lon = lon_data_list[name_data_list.index(name_data_list[station])] #經
            station_lat = lat_data_list[name_data_list.index(name_data_list[station])] #緯
            if haversine(light_data_lat, light_data_lon, station_lat, station_lon) < dis:
                tg_cell = ws_station_lighting_jump_count_data.cell(station+2,light_data_col+1).value
                if tg_cell == None:
                    ws_station_lighting_jump_count_data.cell(station+2,light_data_col+1).value = 1
                else:
                    ws_station_lighting_jump_count_data.cell(station+2,light_data_col+1).value = tg_cell + 1



        light_data_row += 1







wb_station_lighting_jump_count_data.save(data_top_path+"/研究所/閃電資料/lighting_jump事件記數/"+year+'/'+year+'_'+month+'_lighting_jump_count.xlsx')
print('已建立' + data_top_path+"/研究所/閃電資料/lighting_jump事件記數/"+year+'/'+year+'_'+month+'_lighting_jump_count.xlsx')
wb_light_data_read.close()
wb_station_lighting_jump_count_data.close()
