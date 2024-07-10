from openpyxl import Workbook,load_workbook
import math
import re
from datetime import datetime, timedelta
import numpy as np
from tqdm import tqdm ##跑進度條的好玩東西
import pandas as pd
from geopy.distance import geodesic
import time as T


year = '2021' #年分
month = '06' #月份
data_top_path = "C:/Users/steve/python_data"
alpha = 2 #統計檢定

start_time = T.time()

## 讀取雨量站經緯度資料
def rain_station_location_data():
    data_path = data_top_path+"/研究所/雨量資料/"+year+"測站範圍內測站數.xlsx"
    lon_data_list = []  # 經度
    lat_data_list = []  # 緯度
    name_data_list = []  #測站名稱
    wb = load_workbook(data_path)
    ws = wb[month]
    for i in range(ws.max_column):
        lon_data_list.append(ws.cell(3,i+1).value)
        lat_data_list.append(ws.cell(2,i+1).value)
        name_data_list.append(ws.cell(1,i+1).value)
    wb.close()
    return lon_data_list, lat_data_list ,name_data_list
lon_data_list, lat_data_list ,name_data_list = rain_station_location_data()


## 兩經緯度距離
def haversine(lat1, lon1, lat2, lon2):
    # 地球的半徑（公里）
    R = 6371.0

    # 將經緯度從度轉換為弧度
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)

    # 經緯度差
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad

    # Haversine公式
    a = math.sin(dlat / 2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    # 計算距離
    distance = R * c

    return distance


## 與測站距離<36 km的有那些
dis = 36


## 建立存檔位置
wb_lighting_jump = Workbook()
ws_lighting_jump = wb_lighting_jump.active
ws_lighting_jump.title = month


## 讀取閃電資料
flash_data_path = data_top_path+'/研究所/閃電資料/raw_data/'+year+'/'+year+month+'.txt'
flash_rawdata = pd.read_csv(flash_data_path,header = 0)
flash_rawdata['simple_time'] = flash_rawdata['日期時間'].str[:16]
# print(flash_rawdata['simple_time'],flash_rawdata['經度'],flash_rawdata['緯度'])


row = 1
for station_name in name_data_list:

    ws_lighting_jump.cell(row,1).value = station_name
    col = 2

    station_lon = lon_data_list[name_data_list.index(station_name)]
    station_lat = lat_data_list[name_data_list.index(station_name)]
    station_lat_lon = (station_lat,station_lon)
    print(station_name,station_lon,station_lat)

    flash_data_time_list = []

    flash_rawdata['distance_km'] = flash_rawdata.apply(lambda row: geodesic((row['緯度'], row['經度']), station_lat_lon).km, axis=1)
    # print(flash_rawdata)

    target_data = flash_rawdata[flash_rawdata['distance_km'] < dis].reset_index(drop=True)# 篩選距離小於dis的數據並重置索引
    target_data_simple_time = pd.to_datetime(target_data['simple_time'] ,format='%Y-%m-%d %H:%M')
    # print(target_data_simple_time)



    for time in tqdm(target_data_simple_time):

        # print(time)
        SR_list = []


        for start in range(6):
            start_time_dt = time+ timedelta(minutes=start)
            end_time_dt = start_time_dt + timedelta(minutes=5)

            # print(start_time_dt,end_time_dt)
            # print(target_data_simple_time[(start_time_dt <= target_data_simple_time) & (target_data_simple_time< end_time_dt)])
            SR_list.append(target_data_simple_time[(start_time_dt <= target_data_simple_time) & (target_data_simple_time< end_time_dt)].count())



        # print(SR_list)
        SR_1_5_list = SR_list[:5]
        SR_6 = SR_list[5]



        if SR_1_5_list.count(0) == 0:
            SR_1_5_mean = np.mean(SR_1_5_list)
            SR_1_5_std = np.std(SR_1_5_list)

            if SR_6 > SR_1_5_mean + alpha*SR_1_5_std:
                # print(time)
                tg_time = target_data_simple_time[(start_time_dt <= target_data_simple_time) & (target_data_simple_time< end_time_dt)].iloc[-1]
                # print(tg_time)
                ws_lighting_jump.cell(row,col).value = tg_time

                col += 1
    row += 1





    wb_lighting_jump.save(data_top_path+"/研究所/閃電資料/lighting_jump/"+year+'_'+month+'_lighting_jump.xlsx')    
print('已建立' +data_top_path+"/研究所/閃電資料/lighting_jump/"+year+'_'+month+'_lighting_jump.xlsx')

end_time = T.time()
# 計算執行時間
elapsed_time = end_time - start_time

# 將執行時間轉換為小時、分鐘和秒
hours, rem = divmod(elapsed_time, 3600)
minutes, seconds = divmod(rem, 60)

# 打印執行時間
print('程式執行了{}小時{}分{}秒'.format(int(hours), int(minutes), int(seconds)))