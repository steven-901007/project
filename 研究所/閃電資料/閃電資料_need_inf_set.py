from openpyxl import Workbook
from calendar import monthrange #判定一個月有幾天的import
import re



year = '2021'
month = '06'


## 閃電資料讀取
flash_simple_time_list = []
flash_lon_list = [] #*100
flash_lat_list = [] #*100
flash_ic_or_cg = []


data_path = 'C:/Users/steve/python_data/研究所/閃電資料/'+year+'/'+year+month+'.txt'
# result  =glob.glob(path)
# for data_path in result:
#     # f = os.path.basename(f) #單純輸出檔案名稱
#     print(data_path)

delimiter_pattern = re.compile(r',|\n') #當資料分隔符號有"空行"or"空白"or"*"等多個符號時使用
try:
    with open(data_path, 'r',encoding='utf-8') as file:
        next(file)
        for line in file:
            # 使用正則表達式來分割每一行
            elements = re.split(delimiter_pattern, line.strip())
            # print(elements) #以列表顯示
            t = elements[0]
            semple_t = t[0:4] + t[5:7] + t[8:10] + t[11:13] + t[14:16]  #yyyymmddHHMM
            flash_simple_time_list.append(semple_t)
            flash_lon_list.append(round(float(elements[2])*100))
            flash_lat_list.append(round(float(elements[3])*100))
            flash_ic_or_cg.append(elements[5])
except:
    with open(data_path, 'r') as file:
        next(file)
        for line in file:
            # 使用正則表達式來分割每一行
            elements = re.split(delimiter_pattern, line.strip())
            # print(elements) #以列表顯示
            t = elements[0]
            semple_t = t[0:4] + t[5:7] + t[8:10] + t[11:13] + t[14:16]  #yyyymmddHHMM
            flash_simple_time_list.append(semple_t)
            flash_lon_list.append(round(float(elements[2])*100))
            flash_lat_list.append(round(float(elements[3])*100))
            flash_ic_or_cg.append(elements[5])

# print(len(flash_simple_time_list))
# print(flash_lon_list)
# print(flash_lat_list)




## 建立閃電儲存檔案.xlsx
wb_light_data = Workbook()
ws_light_data = wb_light_data.active
ws_light_data.title = month

max_month_day = monthrange(int(year),int(month))[1]
# print(max_month_day)
col = 1
for dd in range(1,max_month_day+1):
    dd = str(dd).zfill(2)
    for hh in range(0,24):
        hh = str(hh).zfill(2)
        for mm in range(0,60,10):
            mm = str(mm).zfill(2)
            # print(yy+mm)
            title_time = dd+hh+mm #分隔的時間起始
            ws_light_data.cell(1,col).value = title_time  ##ddHHMM
            col += 1



## 寫入閃電資料


max_col = ws_light_data.max_column
ws_light_data.cell(1,max_col+1).value = '999999'
for c in range(1,max_col+1):
    ddhhmm = ws_light_data.cell(1,c).value
    simple_title_time = year+month+ddhhmm #yyyymmddHHMM
    # print(simple_title_time)
    end_time = ws_light_data.cell(1,c+1).value
    end_simple_title_time = year+month+end_time
    # print(end_simple_title_time)
    row = 2
    print(ddhhmm)
    for flash_time_lc in range(len(flash_simple_time_list)):
        flash_time = flash_simple_time_list[flash_time_lc]

        if simple_title_time <= flash_time < end_simple_title_time:
            lon_lat = str(flash_lon_list[flash_time_lc]) + str(flash_lat_list[flash_time_lc])
            ws_light_data.cell(row,c).value = lon_lat
            
            row += 1

ws_light_data.cell(1,max_col+1).value = None










wb_light_data.save("C:/Users/steve/python_data/研究所/閃電資料/閃電資料時間線/"+year+'/'+year+"_"+month+'_flash_data.xlsx')
