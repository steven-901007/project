from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta
from tqdm import tqdm
import matplotlib.pyplot as plt
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from cartopy.io.shapereader import Reader
from cartopy.feature import ShapelyFeature
import matplotlib.colors as mcolors
import matplotlib.cm as cm
import matplotlib as mpl



year = '2021' #年分
month = '06' #月份
data_top_path = "C:/Users/steve/python_data"

##測站資料
def rain_station_location_data():
    data_path = data_top_path+"/研究所/雨量資料/"+year+"測站範圍內測站數.xlsx"
    lon_data_list = []  # 經度
    lat_data_list = []  # 緯度
    name_data_list = []  #測站名稱
    wb = load_workbook(data_path)
    ws = wb[month]
    for i in range(ws.max_column):
        lon_data_list.append(ws.cell(3,i+1).value)
        lat_data_list.append(ws.cell(2,i+1).value)
        name_data_list.append(ws.cell(1,i+1).value)
    wb.close()
    return lon_data_list, lat_data_list ,name_data_list
lon_data_list, lat_data_list ,name_data_list = rain_station_location_data()


##雨量資料
rain_data_path = data_top_path+"/研究所/雨量資料/對流性降雨36km統計/"+year+"/"+year+"_"+month+"_36km_rain_data.xlsx"
wb_rain_data = load_workbook(rain_data_path)
ws_rain_data = wb_rain_data[month]
rain_data_max_col = ws_rain_data.max_column

##閃電資料
lighting_jump_path = data_top_path+"/研究所/閃電資料/lighting_jump/"+year+"_"+month+"_lighting_jump.xlsx"
wb_lighting_jump = load_workbook(lighting_jump_path)
ws_lighting_jump = wb_lighting_jump[month]
lighting_jump_data_max_row = ws_lighting_jump.max_row

##降雨資料時間list
rain_time_lc_list = []
for lc in range(1,rain_data_max_col+1):
    rain_time_lc_list.append(ws_rain_data.cell(1,lc).value)
# print(rain_time_lc_list)


##後符
post_agreement_hit_list = [0 for n in name_data_list]

## 後符總量(lighting jump and rain + lighting jump and non_rain)
total_post_agreement_list = [0 for n in name_data_list]

#資料讀取
for lighting_jump_data_row in tqdm(range(1,lighting_jump_data_max_row+1),desc='後符'):

    lighting_jump_station_name = ws_lighting_jump.cell(lighting_jump_data_row,1).value
    # print(lighting_jump_station_name)

    lighting_jump_data_col = 2
    while ws_lighting_jump.cell(lighting_jump_data_row,lighting_jump_data_col).value != None:


        total_post_agreement_list[name_data_list.index(lighting_jump_station_name)] += 1
        lighting_jump_data = ws_lighting_jump.cell(lighting_jump_data_row,lighting_jump_data_col).value
        start_lighting_jump_time = lighting_jump_data[:len(lighting_jump_data)-1] + '0'
        start_lighting_jump_time = datetime.strptime(start_lighting_jump_time, "%Y-%m-%d %H:%M")

        if start_lighting_jump_time.minute % 10 != 0:
            end_lighting_jump_time = start_lighting_jump_time + timedelta(minutes=50)            
        else: 
            start_lighting_jump_time = start_lighting_jump_time + timedelta(minutes=10)
            end_lighting_jump_time = start_lighting_jump_time + timedelta(minutes=40)
        # print(lighting_jump_data)    
        start_lighting_jump_time = start_lighting_jump_time.strftime("%d%H%M")
        end_lighting_jump_time = end_lighting_jump_time.strftime("%d%H%M")

        # print(start_lighting_jump_time,end_lighting_jump_time)
        start_lighting_jump_lc = rain_time_lc_list.index(start_lighting_jump_time)
        end_lighting_jump_lc = rain_time_lc_list.index(end_lighting_jump_time)

        for lc in range(start_lighting_jump_lc+1,end_lighting_jump_lc+2):
            # print(ws_rain_data.cell(1,lc).value)
            rain_data_row = 2
            while ws_rain_data.cell(rain_data_row,lc).value != None:
                rain_data =  ws_rain_data.cell(rain_data_row,lc).value
                rain_style = ws_rain_data.cell(rain_data_row,lc).font.bold
                if lighting_jump_station_name == rain_data and rain_style == False:
                    post_agreement_hit_list[name_data_list.index(lighting_jump_station_name)] += 1
                    # print(lc)

                rain_data_row += 1
        total_post_agreement_list[name_data_list.index(lighting_jump_station_name)] = lighting_jump_data_col-1
        lighting_jump_data_col += 1


#清除資料為0的測站
while post_agreement_hit_list.count(0) != 0:
    lc = post_agreement_hit_list.index(0)
    post_agreement_hit_list.pop(lc)
    lon_data_list.pop(lc)
    lat_data_list.pop(lc)
    total_post_agreement_list.pop(lc)

post_agreement_hit_persent_list = [] # 後符命中率
for i in range(len(total_post_agreement_list)):
    post_agreement_hit_persent_list.append(post_agreement_hit_list[i]/(total_post_agreement_list[i]+post_agreement_hit_list[i])*100)





wb_lighting_jump.close()
wb_rain_data.close()


##後符繪圖

# 設定經緯度範圍
lon_min, lon_max = 120, 122.1
lat_min, lat_max = 21.5, 25.5

plt.figure(figsize=(10, 10))
ax = plt.axes(projection=ccrs.PlateCarree())
ax.set_xlim(lon_min, lon_max)
ax.set_ylim(lat_min, lat_max)

plt.rcParams['font.sans-serif'] = [u'MingLiu']  # 設定字體為'細明體'
plt.rcParams['axes.unicode_minus'] = False  # 用來正常顯示正負號

# 加載台灣的行政邊界
taiwan_shapefile = data_top_path+"/研究所/Taiwan_map_data/COUNTY_MOI_1090820.shp"  # 你需要提供台灣邊界的shapefile文件
shape_feature = ShapelyFeature(Reader(taiwan_shapefile).geometries(),
                               ccrs.PlateCarree(), edgecolor='black', facecolor='white')
ax.add_feature(shape_feature)


# 加入經緯度格線
gridlines = ax.gridlines(draw_labels=True, linestyle='--')
gridlines.top_labels = False
gridlines.right_labels = False

## 計算某個地方達到10mm/10min的次數 + colorbar
color_list = []

level = [0,5,10,20,30,40,50,60,70]
color_box = ['silver','purple','darkviolet','blue','g','y','orange','r']

for nb in post_agreement_hit_list:
    more_then_maxma_or_not = 0
    for j in range(len(level)-1):
        if level[j]<nb<=level[j+1]:
            color_list.append(color_box[j])
            more_then_maxma_or_not = 1
            break
    if more_then_maxma_or_not == 0:
        color_list.append('lime')
        # print(nb)
# print(len(color_list))


# 標記經緯度點
ax.scatter(lon_data_list, lat_data_list, color=color_list, s=3, zorder=5)

# colorbar setting

nlevel = len(level)
cmap1 = mpl.colors.ListedColormap(color_box, N=nlevel)
cmap1.set_over('fuchsia')
cmap1.set_under('black')
norm1 = mcolors.Normalize(vmin=min(level), vmax=max(level))
norm1 = mcolors.BoundaryNorm(level, nlevel, extend='max')
im = cm.ScalarMappable(norm=norm1, cmap=cmap1)
cbar1 = plt.colorbar(im,ax=ax, extend='neither', ticks=level)


# 加入標籤
plt.xlabel('Longitude')
plt.ylabel('Latitude')

ax.set_title(year+"年"+month+"月"+'\n後符 max = '+ str(max(post_agreement_hit_list)))


## 這是用來確認colorbar的配置
fig,ax1 = plt.subplots()
X = [i for i in range(len(post_agreement_hit_list))]
Y = sorted(post_agreement_hit_list)
ax1.plot(X,Y,color =  'black',marker = "*",linestyle = '--') #折線圖
ax1.set_title('這是用來確認colorbar的配置')


## 後符命中率
# 設定經緯度範圍
lon_min, lon_max = 120, 122.1
lat_min, lat_max = 21.5, 25.5

plt.figure(figsize=(10, 10))
ax = plt.axes(projection=ccrs.PlateCarree())
ax.set_xlim(lon_min, lon_max)
ax.set_ylim(lat_min, lat_max)

plt.rcParams['font.sans-serif'] = [u'MingLiu']  # 設定字體為'細明體'
plt.rcParams['axes.unicode_minus'] = False  # 用來正常顯示正負號

# 加載台灣的行政邊界
taiwan_shapefile = data_top_path+"/研究所/Taiwan_map_data/COUNTY_MOI_1090820.shp"  # 你需要提供台灣邊界的shapefile文件
shape_feature = ShapelyFeature(Reader(taiwan_shapefile).geometries(),
                               ccrs.PlateCarree(), edgecolor='black', facecolor='white')
ax.add_feature(shape_feature)


# 加入經緯度格線
gridlines = ax.gridlines(draw_labels=True, linestyle='--')
gridlines.top_labels = False
gridlines.right_labels = False

## 計算某個地方達到10mm/10min的次數 + colorbar
color_list = []

level = [0,5,10,20,30,40,50,60,70]
color_box = ['silver','purple','darkviolet','blue','g','y','orange','r']

for nb in post_agreement_hit_persent_list:
    more_then_maxma_or_not = 0
    for j in range(len(level)-1):
        if level[j]<nb<=level[j+1]:
            color_list.append(color_box[j])
            more_then_maxma_or_not = 1
            break
    if more_then_maxma_or_not == 0:
        color_list.append('lime')
        # print(nb)
# print(len(color_list))


# 標記經緯度點
ax.scatter(lon_data_list, lat_data_list, color=color_list, s=3, zorder=5)

# colorbar setting

nlevel = len(level)
cmap1 = mpl.colors.ListedColormap(color_box, N=nlevel)
cmap1.set_over('fuchsia')
cmap1.set_under('black')
norm1 = mcolors.Normalize(vmin=min(level), vmax=max(level))
norm1 = mcolors.BoundaryNorm(level, nlevel, extend='max')
im = cm.ScalarMappable(norm=norm1, cmap=cmap1)
cbar1 = plt.colorbar(im,ax=ax, extend='neither', ticks=level)


# 加入標籤
plt.xlabel('Longitude')
plt.ylabel('Latitude')

ax.set_title(year+"年"+month+"月"+'\n後符命中率 [%] max = '+ str(max(post_agreement_hit_persent_list)))


## 這是用來確認colorbar的配置
fig,ax1 = plt.subplots()
X = [i for i in range(len(post_agreement_hit_persent_list))]
Y = sorted(post_agreement_hit_persent_list)
ax1.plot(X,Y,color =  'black',marker = "*",linestyle = '--') #折線圖
ax1.set_title('這是用來確認colorbar的配置')


# 顯示地圖
plt.show()