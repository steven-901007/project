import glob
import re
import math
import time
import matplotlib.pyplot as plt
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from cartopy.io.shapereader import Reader
from cartopy.feature import ShapelyFeature
import matplotlib.colors as mcolors
import matplotlib.cm as cm
import matplotlib as mpl
from openpyxl import load_workbook
from mpl_toolkits.axes_grid1 import make_axes_locatable


year = '2021' #年分
month = '06' #月份
data_top_path = "C:/Users/steve/python_data"



## 讀取雨量站經緯度資料
def rain_station_location_data():
    data_path = data_top_path+"/研究所/雨量資料/2021測站範圍內測站數.xlsx"
    lon_data_list = []  # 經度
    lat_data_list = []  # 緯度
    name_data_list = []  #測站名稱
    wb = load_workbook(data_path)
    ws = wb[month]
    for i in range(ws.max_column):
        lon_data_list.append(ws.cell(3,i+1).value)
        lat_data_list.append(ws.cell(2,i+1).value)
        name_data_list.append(ws.cell(1,i+1).value)
    wb.close()
    return lon_data_list, lat_data_list ,name_data_list

lon_data_list, lat_data_list ,name_data_list = rain_station_location_data()


## 測站位置檔案讀取(6月)
wb = load_workbook(data_top_path+"/研究所/雨量資料/2021測站範圍內測站數.xlsx")
ws = wb[month]


count_tg_number_list = [0 for i in range(len(name_data_list))] #計算符合範圍強降雨(36 km)的事件數 對應位置代表name data list
##讀取每日資料
data_name_of_station_list = [] 
data_lon_list = []
data_lat_list = []
data_rain_list = [] #降雨量

## 讀取每月資料
month_path = data_top_path+"/研究所/雨量資料/"+year+"_"+month+"/"+month
result  =glob.glob(month_path+"/*")
for day_path in result:
    # print(day_path)
    day = day_path[53:] #日期
    print('日期:'+day)

    
    ## 讀取每日資料
    result  =glob.glob(day_path+'/*')
    for rain_data_path in result:
        ## 每10分鐘資料處理 rain data >10mm (10min)
        line = 0

        chack_station_not_record_list = []
        with open(rain_data_path, 'r') as files:
            for file in files:
                elements = re.split(re.compile(r'\s+|\n|\*'),file.strip())
                # print(elements)
                # print(len(elements))
                if line >= 3 :
                    station_name = elements[0]
                    
                    rain_data_of_10min = float(elements[7]) #MIN_10
                    if 120 <float(elements[4])< 122.1 and 21.5 <float(elements[3])< 25.5:
                        if rain_data_of_10min >= 0: #排除無資料(data = -998.00)
                            rain_data_of_3_hour = float(elements[8]) #HOUR_3
                            rain_data_of_6_hour = float(elements[9]) #HOUR_6
                            rain_data_of_12_hour = float(elements[10]) #HOUR_12
                            rain_data_of_24_hour = float(elements[11]) #HOUR_24
                            if 10<=rain_data_of_10min <= rain_data_of_3_hour <= rain_data_of_12_hour <= rain_data_of_24_hour: #QC
                                # print(station_name)
                                lc = name_data_list.index(station_name) + 1  #在excel的座標
                                i = 4
                                while ws.cell(i,lc).value != None:
                                    if chack_station_not_record_list.count(ws.cell(i,lc).value) == 0:
                                        chack_station_not_record_list.append(ws.cell(i,lc).value)
                                        count_tg_number_list[name_data_list.index(ws.cell(i,lc).value)] += 1

                                    i += 1
                line += 1

    # break


# print(count_tg_number_list)


# ##dubug區
# print(sum(count_tg_number_list))



## 繪圖

# 設定經緯度範圍
lon_min, lon_max = 120, 122.1
lat_min, lat_max = 21.5, 25.5

plt.figure(figsize=(10, 10))
ax = plt.axes(projection=ccrs.PlateCarree())
ax.set_xlim(lon_min, lon_max)
ax.set_ylim(lat_min, lat_max)

plt.rcParams['font.sans-serif'] = [u'MingLiu']  # 設定字體為'細明體'
plt.rcParams['axes.unicode_minus'] = False  # 用來正常顯示正負號

# 加載台灣的行政邊界
taiwan_shapefile = data_top_path+"/研究所/Taiwan_map_data/COUNTY_MOI_1090820.shp"  # 你需要提供台灣邊界的shapefile文件
shape_feature = ShapelyFeature(Reader(taiwan_shapefile).geometries(),
                               ccrs.PlateCarree(), edgecolor='black', facecolor='white')
ax.add_feature(shape_feature)


# 加入經緯度格線
gridlines = ax.gridlines(draw_labels=True, linestyle='--')
gridlines.top_labels = False
gridlines.right_labels = False

## 計算某個地方達到10mm/10min的次數 + colorbar
color_list = []
level = [0,50,100,150,200,300,350,400,500]
color_box = ['silver','purple','darkviolet','blue','g','y','orange','r']

for i in range(len(count_tg_number_list)):
    more_then_maxma_or_not = 0
    lon_lat_count = count_tg_number_list[i]
    for j in range(len(level)-1):
        if level[j]<=lon_lat_count<level[j+1]:
            color_list.append(color_box[j])
            more_then_maxma_or_not = 1
            break
    if more_then_maxma_or_not == 0:
        color_list.append('black')


# 標記經緯度點
ax.scatter(lon_data_list,lat_data_list, color=color_list, s=3, zorder=5)

# colorbar setting

nlevel = len(level)
cmap1 = mpl.colors.ListedColormap(color_box, N=nlevel)
cmap1.set_over('fuchsia')
cmap1.set_under('black')
norm1 = mcolors.Normalize(vmin=min(level), vmax=max(level))
norm1 = mcolors.BoundaryNorm(level, nlevel, extend='max')
im = cm.ScalarMappable(norm=norm1, cmap=cmap1)
divider = make_axes_locatable(ax)
cax = divider.append_axes("right", size="5%", pad=0.1, axes_class=plt.Axes)
cbar1 = plt.colorbar(im,cax=cax, extend='neither', ticks=level)


# 加入標籤

ax.set_title('雨量>10mm/10min 事件數')

# 顯示地圖
plt.show()
