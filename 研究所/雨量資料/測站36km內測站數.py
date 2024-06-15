import re
import math
from openpyxl import Workbook, load_workbook
import glob

year = '2021' #年分
month = '06' #月份

def rain_station_location_data(path):
    data_path = path
    lon_data_list = []  # 經度
    lat_data_list = []  # 緯度
    name_data_list = []  #測站名稱

    line = 0
    with open(data_path, 'r') as files:
        for file in files:
            if line >=3:
                data = re.split(re.compile(r'\s+|\n|\*'),file.strip())
                # print(data)
                if 120 <float(data[4])< 122.1 and 21.5 <float(data[3])< 25.5:
                    lon_data_list.append(float(data[4]))
                    lat_data_list.append(float(data[3]))
                    name_data_list.append(data[0])
            line +=1
    return lon_data_list, lat_data_list ,name_data_list

lon_data_list, lat_data_list ,station_name_data_list = [],[],[]

##確認所有資料的測站都有被記錄
## 讀取每月資料
month_path = "C:/Users/steve/python_data/研究所/雨量資料/"+year+"_"+month+"/"+month
result  =glob.glob(month_path+"/*")
for day_path in result:
    # print(day_path)
    day = day_path[53:] #日期
    print('日期:'+day)

    ## 讀取每日資料
    result  =glob.glob(day_path+'/*')
    for rain_data_path in result:
        # print(rain_data_path)
        lon_list, lat_list ,station_name_list = rain_station_location_data(rain_data_path)
        for i in range(len(station_name_list)):
            if station_name_list[i] == 'A0W100':
                print(station_name_data_list)
            if station_name_data_list.count(station_name_list[i]) == 0:
                station_name_data_list.append(station_name_list[i])
                lon_data_list.append(lon_list[i])
                lat_data_list.append(lat_list[i])








## 兩經緯度距離
def haversine(lat1, lon1, lat2, lon2):
    # 地球的半徑（公里）
    R = 6371.0

    # 將經緯度從度轉換為弧度
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)

    # 經緯度差
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad

    # Haversine公式
    a = math.sin(dlat / 2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    # 計算距離
    distance = R * c

    return distance


## 與測站距離<36 km的有那些
dis = 36

#儲存位置建立
wb = Workbook() 
ws = wb.active
ws.title = str(int(month))+"月" #創第一個sheet


for i in range(len(station_name_data_list)):
    ws.cell(1,i+1).value = station_name_data_list[i]
    ws.cell(2,i+1).value = lat_data_list[i]
    ws.cell(3,i+1).value = lon_data_list[i]
    lc = 4
    for j in range(len(station_name_data_list)):
        station_dis = haversine(lat_data_list[i],lon_data_list[i],lat_data_list[j],lon_data_list[j])
        if station_dis < 36:
            ws.cell(lc,i+1).value = station_name_data_list[j]
            lc +=1
    print(i)



wb.save("C:/Users/steve/python_data/研究所/雨量資料/"+year+"測站範圍內測站數.xlsx")