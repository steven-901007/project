import glob
import matplotlib.pyplot as plt
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from cartopy.io.shapereader import Reader
from cartopy.feature import ShapelyFeature
import matplotlib.colors as mcolors
import matplotlib.cm as cm
import matplotlib as mpl
from openpyxl import load_workbook



year = '2021' #年分
month = '06' #月份
data_top_path = "C:/Users/steve/python_data"



## 讀取雨量站經緯度資料
def rain_station_location_data():
    data_path = data_top_path+"/研究所/雨量資料/"+year+"測站範圍內測站數.xlsx"
    lon_data_list = []  # 經度
    lat_data_list = []  # 緯度
    name_data_list = []  #測站名稱
    wb = load_workbook(data_path)
    ws = wb[month]
    for i in range(ws.max_column):
        lon_data_list.append(ws.cell(3,i+1).value)
        lat_data_list.append(ws.cell(2,i+1).value)
        name_data_list.append(ws.cell(1,i+1).value)
    wb.close()
    return lon_data_list, lat_data_list ,name_data_list

lon_data_list, lat_data_list ,name_data_list = rain_station_location_data()


##36 km統計雨量資料
rain_data_path = data_top_path+"/研究所/雨量資料/對流性降雨36km統計/"+year+"/"+year+"_"+month+"_36km_rain_data.xlsx"
wb_rain_data = load_workbook(rain_data_path)
ws_rain_data = wb_rain_data[month]
max_col_rain_data = ws_rain_data.max_column



rain_36km_list = [] #站點名稱
rain_36km_count_list = [] #降雨次數
rain_36km_lon_list = []
rain_36km_lat_list = []

##降雨資料讀取


for col in range(1,max_col_rain_data+1):
    row = 2
    print(col)
    while ws_rain_data.cell(row,col).value != None:
        rain_data = ws_rain_data.cell(row,col).value
        rain_data_style = ws_rain_data.cell(row,col).font.bold  #判斷資料是否為粗體
        # print(rain_data_style)
        if rain_data_style == False:
            if rain_36km_list.count(rain_data) == 0:
                rain_36km_list.append(rain_data)
                rain_36km_count_list.append(1)
                rain_36km_lon_list.append(lon_data_list[name_data_list.index(rain_data)])
                rain_36km_lat_list.append(lat_data_list[name_data_list.index(rain_data)])
            else:
                rain_36km_count_list[rain_36km_list.index(rain_data)] += 1
        
        row += 1

print(rain_36km_count_list)
# print(len(rain_36km_list))

# ##debug區
print(sum(rain_36km_count_list))
# # 初步判定資料無誤(2024/06/15)



## 繪圖

# 設定經緯度範圍
lon_min, lon_max = 120, 122.1
lat_min, lat_max = 21.5, 25.5

plt.figure(figsize=(10, 10))
ax = plt.axes(projection=ccrs.PlateCarree())
ax.set_xlim(lon_min, lon_max)
ax.set_ylim(lat_min, lat_max)

plt.rcParams['font.sans-serif'] = [u'MingLiu']  # 設定字體為'細明體'
plt.rcParams['axes.unicode_minus'] = False  # 用來正常顯示正負號

# 加載台灣的行政邊界
taiwan_shapefile = data_top_path+"/研究所/Taiwan_map_data/COUNTY_MOI_1090820.shp"  # 你需要提供台灣邊界的shapefile文件
shape_feature = ShapelyFeature(Reader(taiwan_shapefile).geometries(),
                               ccrs.PlateCarree(), edgecolor='black', facecolor='white')
ax.add_feature(shape_feature)


# 加入經緯度格線
gridlines = ax.gridlines(draw_labels=True, linestyle='--')
gridlines.top_labels = False
gridlines.right_labels = False

## 計算某個地方達到10mm/10min的次數 + colorbar
color_list = []

level = [0,50,100,150,200,300,350,400,500]
color_box = ['silver','purple','darkviolet','blue','g','y','orange','r']

for nb in rain_36km_count_list:
    more_then_maxma_or_not = 0
    for j in range(len(level)-1):
        if level[j]<nb<=level[j+1]:
            color_list.append(color_box[j])
            more_then_maxma_or_not = 1
            break
    if more_then_maxma_or_not == 0:
        color_list.append('lime')
        print(nb)
# print(len(color_list))


# 標記經緯度點
ax.scatter(rain_36km_lon_list, rain_36km_lat_list, color=color_list, s=3, zorder=5)

# colorbar setting

nlevel = len(level)
cmap1 = mpl.colors.ListedColormap(color_box, N=nlevel)
cmap1.set_over('fuchsia')
cmap1.set_under('black')
norm1 = mcolors.Normalize(vmin=min(level), vmax=max(level))
norm1 = mcolors.BoundaryNorm(level, nlevel, extend='max')
im = cm.ScalarMappable(norm=norm1, cmap=cmap1)
cbar1 = plt.colorbar(im, extend='neither', ticks=level)


# 加入標籤
plt.xlabel('Longitude')
plt.ylabel('Latitude')

ax.set_title(year+"年"+month+"月"+'\n雨量>10mm/10min 事件數\nmax = '+ str(max(rain_36km_count_list)))


## 這是用來確認colorbar的配置
fig,ax1 = plt.subplots()
X = [i for i in range(len(rain_36km_count_list))]
Y = sorted(rain_36km_count_list)
ax1.plot(X,Y,color =  'black',marker = "*",linestyle = '--') #折線圖
ax1.set_title('這是用來確認colorbar的配置')


# 顯示地圖
plt.show()


