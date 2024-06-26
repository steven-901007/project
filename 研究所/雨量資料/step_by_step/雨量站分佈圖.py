import matplotlib.pyplot as plt
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from cartopy.io.shapereader import Reader
from cartopy.feature import ShapelyFeature
import re
from openpyxl import load_workbook


year = "2021"
month = '06'
data_top_path = "C:/Users/steve/python_data"


## 讀取雨量站經緯度資料
def rain_station_location_data():
    data_path = data_top_path+"/研究所/雨量資料/"+year+"測站範圍內測站數.xlsx"
    lon_data_list = []  # 經度
    lat_data_list = []  # 緯度
    name_data_list = []  #測站名稱
    wb = load_workbook(data_path)
    ws = wb[month]
    for i in range(ws.max_column):
        lon_data_list.append(ws.cell(3,i+1).value)
        lat_data_list.append(ws.cell(2,i+1).value)
        name_data_list.append(ws.cell(1,i+1).value)
    wb.close()
    return lon_data_list, lat_data_list ,name_data_list

lon_data_list, lat_data_list ,name_data_list = rain_station_location_data()

## 繪圖

# 設定經緯度範圍
lon_min, lon_max = 120, 122.1
lat_min, lat_max = 21.5, 25.5

plt.figure(figsize=(10, 10))
ax = plt.axes(projection=ccrs.PlateCarree())
ax.set_xlim(lon_min, lon_max)
ax.set_ylim(lat_min, lat_max)

plt.rcParams['font.sans-serif'] = [u'MingLiu']  # 設定字體為'細明體'
plt.rcParams['axes.unicode_minus'] = False  # 用來正常顯示正負號

# 加載台灣的行政邊界
taiwan_shapefile = data_top_path+"/研究所/Taiwan_map_data/COUNTY_MOI_1090820.shp"  # 你需要提供台灣邊界的shapefile文件
shape_feature = ShapelyFeature(Reader(taiwan_shapefile).geometries(),
                               ccrs.PlateCarree(), edgecolor='black', facecolor='lightgray')
ax.add_feature(shape_feature)


# 加入經緯度格線
gridlines = ax.gridlines(draw_labels=True, linestyle='--')
gridlines.top_labels = False
gridlines.right_labels = False

# 標記經緯度點
ax.scatter(lon_data_list, lat_data_list, color='red', s=3, zorder=5)

# 加入標籤
plt.xlabel('Longitude')
plt.ylabel('Latitude')
ax.set_title('全台雨量站座標\n雨量站數量：' + str(len(name_data_list)))

# 顯示地圖
plt.show()
