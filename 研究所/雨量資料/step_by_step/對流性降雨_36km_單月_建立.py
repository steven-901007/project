from openpyxl import Workbook,load_workbook
import glob
import re
from openpyxl.styles import Font


year = '2021' #年分
month = '06' #月份


## 測站數file 讀取
station_number_path = "C:/Users/steve/python_data/研究所/雨量資料/"+year+"測站範圍內測站數.xlsx"
wb_station_number = load_workbook(station_number_path)
ws_station_number = wb_station_number[str(int(month))+'月']
station_name_count = ws_station_number.max_column #站點數量


## 建立存檔file
wb_rain_more_10mm_36km_setting = Workbook()
ws_rain_more_10mm_36km_setting = wb_rain_more_10mm_36km_setting.active
ws_rain_more_10mm_36km_setting.title = month



## 強降雨測站file 讀取
rain_more_10mm_data_path = "C:/Users/steve/python_data/研究所/雨量資料/對流性降雨data/"+year+"/"+year+"_"+month+"_rain_data.xlsx"
wb_rain_more_10mm_data = load_workbook(rain_more_10mm_data_path)
sheet_name_list = wb_rain_more_10mm_data.sheetnames #日期
# print(sheet_name_list)


# 讀取強降雨測站資料

ws_rain_more_10mm_data = wb_rain_more_10mm_data[month]
max_col_for_rain_more_10mm_data = ws_rain_more_10mm_data.max_column # time


for col in range(1,max_col_for_rain_more_10mm_data+1): #分鐘
    row = 2
    time = year+month+ws_rain_more_10mm_data.cell(1,col).value #yyyymmddHHMM
    print(time)
    ws_rain_more_10mm_36km_setting.cell(1,col).value = str(ws_rain_more_10mm_data.cell(1,col).value)
    # print(col)

    
    col_for_rain_more_10mm_36km_setting = 2
    while ws_rain_more_10mm_data.cell(row,col).value != None:
        rain_more_10mm_data = ws_rain_more_10mm_data.cell(row,col).value
        # print(rain_more_10mm_data)
        
        for station_count in range(1,station_name_count+1):
            if rain_more_10mm_data == ws_station_number.cell(1,station_count).value:
                col_for_station_number = 4
                while ws_station_number.cell(col_for_station_number,station_count).value != None:
                    ws_rain_more_10mm_36km_setting.cell(col_for_rain_more_10mm_36km_setting,col).value = ws_station_number.cell(col_for_station_number,station_count).value
                    col_for_station_number += 1
                    col_for_rain_more_10mm_36km_setting += 1
        
        row += 1

    ## 確認每個時間點的同一地點資料是否只有一筆(重複的部分設為粗體紅字)
    chack_row = 2
    chack_list = []
    while ws_rain_more_10mm_36km_setting.cell(chack_row,col).value != None:
        tg_data = ws_rain_more_10mm_36km_setting.cell(chack_row,col).value
        if chack_list.count(tg_data) == 0:
            chack_list.append(tg_data)

        else:               
            ws_rain_more_10mm_36km_setting.cell(chack_row,col).font = Font (bold = True , color= "FFFF0000")

        chack_row += 1

                

            
    
    




wb_rain_more_10mm_36km_setting.save("C:/Users/steve/python_data/研究所/雨量資料/對流性降雨36km統計/"+year+"/"+year+"_"+month+"_36km_rain_data.xlsx")
print("已建立\nC:/Users/steve/python_data/研究所/雨量資料/對流性降雨36km統計/"+year+"/"+year+"_"+month+"_36km_rain_data.xlsx")