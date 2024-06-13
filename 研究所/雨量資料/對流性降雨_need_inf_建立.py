from openpyxl import Workbook
import glob
import re

year = '2021' #年分
month = '06' #月份

## 建立存檔file
wb = Workbook()
wb.active
wb.remove(wb['Sheet'])



## 讀取每月資料



month_path = "C:/Users/steve/python_data/研究所/雨量資料/"+year+"_"+month+"/"+month
result  =glob.glob(month_path+"/*")
for day_path in result:
    day = day_path[53:] #日期   
    wb.create_sheet(day) #建立所屬sheet
    print('日期:'+day)
    ws = wb[day]
    time_lc = 1
    ## 讀取每日資料

    result  =glob.glob(day_path+'/*')
    for rain_data_path in result:
        ten_min = rain_data_path[64:68]
        print('時間:'+ten_min)
        ws.cell(1,time_lc).value = ten_min
        data_lc = 2 #資料紀錄起始位置

        # 每10分鐘資料處理 rain data >10mm (10min)
        line = 0
        with open(rain_data_path, 'r') as files:
            for file in files:
                elements = re.split(re.compile(r'\s+|\n|\*'),file.strip()) 

                if line >= 3 :  #移除檔頭
                    
                    if 120 <float(elements[4])< 122.1 and 21.5 <float(elements[3])< 25.5: #確認經緯度範圍
                        station_name = elements[0] #測站名稱
                        rain_data_of_10min = float(elements[7]) #MIN_10

                        if rain_data_of_10min >= 0: #排除無資料(data = -998.00)
                            rain_data_of_3_hour = float(elements[8]) #HOUR_3
                            rain_data_of_6_hour = float(elements[9]) #HOUR_6
                            rain_data_of_12_hour = float(elements[10]) #HOUR_12
                            rain_data_of_24_hour = float(elements[11]) #HOUR_24

                            if 10<=rain_data_of_10min <= rain_data_of_3_hour <= rain_data_of_12_hour <= rain_data_of_24_hour: #QC
                                ws.cell(data_lc,time_lc).value = station_name
                                data_lc += 1
        
                line += 1
        time_lc += 1




wb.save("C:/Users/steve/python_data/研究所/雨量資料/對流性降雨data/"+year+"/"+year+"_"+month+".xlsx")

