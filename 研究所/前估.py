from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta
from tqdm import tqdm
import matplotlib.pyplot as plt
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from cartopy.io.shapereader import Reader
from cartopy.feature import ShapelyFeature
import matplotlib.colors as mcolors
import matplotlib.cm as cm
import matplotlib as mpl



year = '2021' #年分
month = '06' #月份
data_top_path = "C:/Users/steve/python_data"

##測站資料
def rain_station_location_data():
    data_path = data_top_path+"/研究所/雨量資料/"+year+"測站範圍內測站數.xlsx"
    lon_data_list = []  # 經度
    lat_data_list = []  # 緯度
    name_data_list = []  #測站名稱
    wb = load_workbook(data_path)
    ws = wb[month]
    for i in range(ws.max_column):
        lon_data_list.append(ws.cell(3,i+1).value)
        lat_data_list.append(ws.cell(2,i+1).value)
        name_data_list.append(ws.cell(1,i+1).value)
    wb.close()
    return lon_data_list, lat_data_list ,name_data_list
lon_data_list, lat_data_list ,name_data_list = rain_station_location_data()


##雨量資料
rain_data_path = data_top_path+"/研究所/雨量資料/對流性降雨36km統計/"+year+"/"+year+"_"+month+"_36km_rain_data.xlsx"
wb_rain_data = load_workbook(rain_data_path)
ws_rain_data = wb_rain_data[month]
rain_data_max_col = ws_rain_data.max_column

##閃電資料
lighting_jump_path = data_top_path+"/研究所/閃電資料/lighting_jump/"+year+"_"+month+"_lighting_jump.xlsx"
wb_lighting_jump = load_workbook(lighting_jump_path)
ws_lighting_jump = wb_lighting_jump[month]
lighting_jump_data_max_row = ws_lighting_jump.max_row

#建立閃電資料測站在excel的位置
lighting_jump_station_lc_list = [] #位置在list的lc+1
for lc in range(1,lighting_jump_data_max_row+1):
    lighting_jump_station_lc_list.append(ws_lighting_jump.cell(lc,1).value)
# print(lighting_jump_station_lc_list)





##前估
#個測站命中的list
prefigurance_hit_list = [0 for n in name_data_list]

#前估總量(lighting jump and rain + non_lighting jump and rain)
total_prefigurance_list = [0 for n in name_data_list]

#資料讀取
for rain_data_col in tqdm(range(1,rain_data_max_col+1),desc='前估'):

    rain_data_row = 2
    end_rain_time = datetime.strptime(ws_rain_data.cell(1,rain_data_col).value, "%d%H%M")
    start_rain_time = end_rain_time - timedelta(minutes=50)
    end_rain_time = end_rain_time.strftime("%d%H%M")
    start_rain_time = start_rain_time.strftime("%d%H%M")
    # print(start_rain_time,end_rain_time)

    while ws_rain_data.cell(rain_data_row,rain_data_col).value != None:
        rain_data_style = ws_rain_data.cell(rain_data_row,rain_data_col).font.bold 

        if rain_data_style == False:
            rain_data_station = ws_rain_data.cell(rain_data_row,rain_data_col).value
            # print(rain_data_station)
            total_prefigurance_list[name_data_list.index(rain_data_station)] += 1

            lighting_jump_col = 2
            while ws_lighting_jump.cell(lighting_jump_station_lc_list.index(rain_data_station)+1,lighting_jump_col).value != None:
                lighting_jump_data = ws_lighting_jump.cell(lighting_jump_station_lc_list.index(rain_data_station)+1,lighting_jump_col).value
                lighting_jump_data = lighting_jump_data.strftime("%d%H%M")

                if start_rain_time <= lighting_jump_data <= end_rain_time:
                    # print(rain_data_station)
                    prefigurance_hit_list[name_data_list.index(rain_data_station)] += 1
                lighting_jump_col += 1
            # print(rain_data_station)
        rain_data_row += 1
        

# print(prefigurance_hit_list)



#清除資料為0的測站
while prefigurance_hit_list.count(0) != 0:
    lc = prefigurance_hit_list.index(0)
    prefigurance_hit_list.pop(lc)
    lon_data_list.pop(lc)
    lat_data_list.pop(lc)
    total_prefigurance_list.pop(lc)

prefigurance_hit_persent_list = [] # 前估命中率
for i in range(len(total_prefigurance_list)):
    prefigurance_hit_persent_list.append(prefigurance_hit_list[i]/(total_prefigurance_list[i]+prefigurance_hit_list[i])*100)

wb_lighting_jump.close()
wb_rain_data.close()


##前估繪圖

# 設定經緯度範圍
lon_min, lon_max = 120, 122.1
lat_min, lat_max = 21.5, 25.5

plt.figure(figsize=(10, 10))
ax = plt.axes(projection=ccrs.PlateCarree())
ax.set_xlim(lon_min, lon_max)
ax.set_ylim(lat_min, lat_max)

plt.rcParams['font.sans-serif'] = [u'MingLiu']  # 設定字體為'細明體'
plt.rcParams['axes.unicode_minus'] = False  # 用來正常顯示正負號

# 加載台灣的行政邊界
taiwan_shapefile = data_top_path+"/研究所/Taiwan_map_data/COUNTY_MOI_1090820.shp"  # 你需要提供台灣邊界的shapefile文件
shape_feature = ShapelyFeature(Reader(taiwan_shapefile).geometries(),
                               ccrs.PlateCarree(), edgecolor='black', facecolor='white')
ax.add_feature(shape_feature)


# 加入經緯度格線
gridlines = ax.gridlines(draw_labels=True, linestyle='--')
gridlines.top_labels = False
gridlines.right_labels = False

## 計算某個地方達到10mm/10min的次數 + colorbar
color_list = []

level = [0,50,100,150,200,500,700,1000,1500]
color_box = ['silver','purple','darkviolet','blue','g','y','orange','r']

for nb in prefigurance_hit_list:
    more_then_maxma_or_not = 0
    for j in range(len(level)-1):
        if level[j]<nb<=level[j+1]:
            color_list.append(color_box[j])
            more_then_maxma_or_not = 1
            break
    if more_then_maxma_or_not == 0:
        color_list.append('lime')
        # print(nb)
# print(len(color_list))


# 標記經緯度點
ax.scatter(lon_data_list, lat_data_list, color=color_list, s=3, zorder=5)

# colorbar setting

nlevel = len(level)
cmap1 = mpl.colors.ListedColormap(color_box, N=nlevel)
cmap1.set_over('fuchsia')
cmap1.set_under('black')
norm1 = mcolors.Normalize(vmin=min(level), vmax=max(level))
norm1 = mcolors.BoundaryNorm(level, nlevel, extend='max')
im = cm.ScalarMappable(norm=norm1, cmap=cmap1)
cbar1 = plt.colorbar(im,ax=ax, extend='neither', ticks=level)


# 加入標籤
plt.xlabel('Longitude')
plt.ylabel('Latitude')

ax.set_title(year+"年"+month+"月"+'\n前估 max = '+ str(max(prefigurance_hit_list)))


## 這是用來確認colorbar的配置
fig,ax1 = plt.subplots()
X = [i for i in range(len(prefigurance_hit_list))]
Y = sorted(prefigurance_hit_list)
ax1.plot(X,Y,color =  'black',marker = "*",linestyle = '--') #折線圖
ax1.set_title('這是用來確認colorbar的配置')



##前估命中率繪圖

plt.figure(figsize=(10, 10))
ax = plt.axes(projection=ccrs.PlateCarree())
ax.set_xlim(lon_min, lon_max)
ax.set_ylim(lat_min, lat_max)

plt.rcParams['font.sans-serif'] = [u'MingLiu']  # 設定字體為'細明體'
plt.rcParams['axes.unicode_minus'] = False  # 用來正常顯示正負號

# 加載台灣的行政邊界
taiwan_shapefile = data_top_path+"/研究所/Taiwan_map_data/COUNTY_MOI_1090820.shp"  # 你需要提供台灣邊界的shapefile文件
shape_feature = ShapelyFeature(Reader(taiwan_shapefile).geometries(),
                               ccrs.PlateCarree(), edgecolor='black', facecolor='white')
ax.add_feature(shape_feature)


# 加入經緯度格線
gridlines = ax.gridlines(draw_labels=True, linestyle='--')
gridlines.top_labels = False
gridlines.right_labels = False

## 計算某個地方達到10mm/10min的次數 + colorbar
color_list = []

level = [0,10,20,30,40,50,60,70,80]
color_box = ['silver','purple','darkviolet','blue','g','y','orange','r']

for nb in prefigurance_hit_persent_list:
    more_then_maxma_or_not = 0
    for j in range(len(level)-1):
        if level[j]<nb<=level[j+1]:
            color_list.append(color_box[j])
            more_then_maxma_or_not = 1
            break
    if more_then_maxma_or_not == 0:
        color_list.append('lime')
        # print(nb)
# print(len(color_list))


# 標記經緯度點
ax.scatter(lon_data_list, lat_data_list, color=color_list, s=3, zorder=5)

# colorbar setting

nlevel = len(level)
cmap1 = mpl.colors.ListedColormap(color_box, N=nlevel)
cmap1.set_over('fuchsia')
cmap1.set_under('black')
norm1 = mcolors.Normalize(vmin=min(level), vmax=max(level))
norm1 = mcolors.BoundaryNorm(level, nlevel, extend='max')
im = cm.ScalarMappable(norm=norm1, cmap=cmap1)
cbar1 = plt.colorbar(im,ax=ax, extend='neither', ticks=level)


# 加入標籤
plt.xlabel('Longitude')
plt.ylabel('Latitude')

ax.set_title(year+"年"+month+"月"+'\n前估命中率 [%] max = '+ str(max(prefigurance_hit_persent_list)))


## 這是用來確認colorbar的配置
fig,ax1 = plt.subplots()
X = [i for i in range(len(prefigurance_hit_persent_list))]
Y = sorted(prefigurance_hit_persent_list)
ax1.plot(X,Y,color =  'black',marker = "*",linestyle = '--') #折線圖
ax1.set_title('這是用來確認colorbar的配置')


# 顯示地圖
plt.show()