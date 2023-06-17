from openpyxl import load_workbook,Workbook

day = '2019-01-19'

wb = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/"+day+"/10min.xlsx")
wsu = wb['U']
wsv = wb['V']
wb1 = Workbook()
ws1u = wb1.active
ws1u.title = 'U'
ws1v = wb1.create_sheet('V')
lc = 2
for time in range(7,140,6):    
    # print(wsu.cell(1,time).value)
    for hight in range(1,65):
        # print(wsu.cell(hight,time).value)
        ws1u.cell(hight,lc).value = wsu.cell(hight,time).value
    lc+=1
lc = 2
for time in range(7,140,6):    
    # print(wsu.cell(1,time).value)
    for hight in range(1,65):
        # print(wsu.cell(hight,time).value)
        ws1v.cell(hight,lc).value = wsv.cell(hight,time).value
    lc+=1
lc = 2
for i in range(50,1601,25):
    ws1u.cell(lc,1).value = i
    ws1v.cell(lc,1).value = i
    lc +=1

wb1.save("C:/Users/steve/Desktop/python/松山剖風儀資料/"+day+"/1hour.xlsx")

# wb1.create_sheet('v')