import csv
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import random
import matplotlib
import glob
import numpy as np
import math
import os

#起始高度=50m
circle = 63 #一個迴圈資料數量(50~1600)有機筆
files  =glob.glob("C:/Users/steve/Desktop/python/松山剖風儀/2019-03-13/wind_reconstruction_data/*")
for file in files:
        file = file + '/*.csv'

        result  =glob.glob(file)

        for f in result:
            try:
                wb = Workbook()
                ws = wb.active
                wbu = Workbook()    
                wsu = wbu.active
                wbv = Workbook()
                wsv = wbv.active

                # f = "C:/Users/steve/Desktop/python/松山剖風儀/2019-03-13/wind_reconstruction_data/00-00\WLS100s-46_WindReconstructionData_2019-03-13_00-00-20_DBS_74_25m.csv"
                with open(f,newline='') as csvfile:
                    rows = csv.DictReader(csvfile,delimiter=';')
                    for  row in rows :
                #         ws.append([int(row['Altitude [m]']),int(row['LOS ID']),float(row['Radial Wind Speed [m/s]'])]) #高度(整數)、方向代號(整數)、風速(小數)
                # wb.save('123.xlsx')
                        ws.append([float(row['Radial Wind Speed [m/s]']),int(row['Confidence Index [%]'])])
                for i in range(circle+1,2*circle+1):
                    wsu.append([round((float(ws.cell(i+2*circle,1).value) - float(ws.cell(i,1).value))/(2*math.cos(75)),2)])#3-1 (W-E)
                for i in range(1,circle+1):
                    wsv.append([round((float(ws.cell(i,1).value) - float(ws.cell(i+2*circle,1).value))/(2*math.cos(75)),2)])#0-2 (S-N)
                
                for j in range(1,circle+1):
                    if ws.cell(j,2).value >=75 and ws.cell(j+circle,2).value >=75 and ws.cell(j+circle*2,2).value >=75 and ws.cell(j+circle*3,2).value >=75:
                        wsu.cell(j,2).value = 1        
                    else:
                        wsu.cell(j,2).value = 0
                time = f[123:131]
                wbu.save('C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/u/'+time+'.xlsx') #第1行高度 = 100m
                wbv.save('C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/v/'+time+'.xlsx') #第1行高度 = 100m
                print(time)
            except:
                print(f)
