import csv
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import random
import matplotlib
import glob
import numpy as np
import math
import os

#起始高度=100m

files  =glob.glob("C:/Users/steve/Desktop/python/松山剖風儀/2019-03-13/wind_reconstruction_data/*")
for file in files:
        file = file + '/*.csv'

        result  =glob.glob(file)

        for f in result:
            try:
                wb = Workbook()
                ws = wb.active
                wbu = Workbook()    
                wsu = wbu.active
                wbv = Workbook()
                wsv = wbv.active
                # f = "C:/Users/steve/Desktop/python/松山剖風儀/2019-11-24/wind_reconstruction_data/00-00\WLS100s-46_WindReconstructionData_2019-11-24_00-09-15_DBS_103_25m.csv"
                with open(f,newline='') as csvfile:
                    rows = csv.DictReader(csvfile,delimiter=';')
                    for  row in rows :
                        ws.append([float(row['Radial Wind Speed [m/s]'])])

                for i in range(61):
                    ws.delete_rows(1)
                    ws.delete_rows(305-i)


                for i in range(62,123):
                    wsu.append([round((float(ws.cell(i,1).value) - float(ws.cell(i+112,1).value))/(2*math.cos(75)),2)])#1-3 (E-W)
                for i in range(1,62):
                    wsv.append([round((float(ws.cell(i+112,1).value) - float(ws.cell(i,1).value))/(2*math.cos(75)),2)])#2-0 (N-S)
                time = f[112:131]
                wbu.save('C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/u/'+time+'.xlsx') #第1行高度 = 100m
                wbv.save('C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/v/'+time+'.xlsx') #第1行高度 = 100m
            except:
                print(f)
#             # wb.save('2019-11-24-00-09-15.xlsx')
#             # print()
