import matplotlib.pyplot as plt
import numpy as np
from  openpyxl import load_workbook
import matplotlib as mpl
import datetime

#10分鐘等高線圖、每刻個高度資料量等高線圖、個資料量常條圖

wb = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計2/u_mean10min.xlsx")
ws = wb['Mean']
rowmax = ws.max_row #62
colmax = ws.max_column #144

img1 = [[]for i in range(rowmax-1)]
# print(img)
for j in range(2,rowmax+1):
    
    for i in range(2,colmax+1):
        img1[j-2].append(ws.cell(j,i).value)

#     print (len(img[j-2]))
# print (len(img))
# print(img)

x1 = range(colmax-1)
y1 = range(100,1601,25)
style = []
for i in range(0,24):
    for j in range(0,51,10):
        if i != 0 or j != 0:
            i = str(i)
            j = str(j)
            i = i.zfill(2)
            j = j.zfill(2)
            if j =='00':
                style.append(i+'-'+j)
            else:
                style.append('')

        
ws = wb['Sigma']
rowmax = ws.max_row #62
colmax = ws.max_column #144

img2 = [[]for i in range(rowmax-1)]
# print(img)
for j in range(2,rowmax+1):
    
    for i in range(2,colmax+1):
        img2[j-2].append(ws.cell(j,i).value)

#     print (len(img[j-2]))
# print (len(img))
# print(img)
name = ['10','9','8','7','6','5','4']
ten = 0
night = 0
eight = 0
seven = 0
six = 0
five = 0
four = 0
Else = 0
for i in range(2,colmax+1):
    for j in range(2,rowmax+1):
        nb = int(ws.cell(j,i).value)
        if nb ==10:
            ten +=1
        elif nb == 9:
            night +=1
        elif nb == 8:
            eight +=1
        elif nb == 7:
            seven +=1
        elif nb == 6:
            six +=1
        elif nb == 5:
            five +=1
        elif nb == 4:
            four +=1
        else :
            Else += 1
# print ('10=' + str(ten))
# print ('9 ='+str(night)) 
# print ('8 ='+str(eight))
# print ('7 ='+str(seven))
# print ('6 ='+ str(six))
# print ('5 ='+str(five))
# print ('4 ='+str(four))
NB = [ten,night,eight,seven,six,five,four]

x2 = range(colmax-1)
y2 = range(rowmax-1)



fig = plt.subplots()

windv =[-10.7,-7.9,-5.4,-3.3,-1.5,-1,0,1.5,3.3,5.4,7.9,10.7]
plt.contourf(x1,y1,img1,levels = windv,alpha = .75,cmap = mpl.cm.jet)
plt.xlabel('time')
plt.xticks(x1,style,rotation = 75)
plt.ylabel('height')
plt.title('2019/11/24/wind-u')
plt.colorbar()



fig = plt.figure()
drow =fig.add_subplot(1,2,1)
windv =[3,4.5,5.5,6.5,7.5,8.5,9.5,10.5,11]
plt.contourf(x2,y2,img2,levels = windv,alpha = .75,cmap = mpl.cm.jet)
plt.xlabel('time')
plt.ylabel('height')
plt.title('<1.5 singm')

# plt.contour(x2,y2,img2,alpha = .75,)
# meandrow.set_xticks(range(0,144,10))
# C = plt.contour(x1,y1,img1,1,colors = 'black',linewidth=0.5)
# plt.clabel(C,inline=True,fontsize=2)
drow = fig.add_subplot(1,2,2)
x = np.arange(len(name)) 
plt.bar(x,NB, color=['red', 'green', 'blue', 'yellow'])
plt.xticks(x,name)
plt.xlabel('Students')
plt.ylabel('Math')
plt.title('Final Term')

plt.colorbar()



plt.show()