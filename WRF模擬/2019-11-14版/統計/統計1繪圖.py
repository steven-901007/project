import matplotlib.pyplot as plt
import numpy as np
from  openpyxl import load_workbook
import matplotlib as mpl

wb = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計1/u_all統計1.xlsx")
ws = wb['Sheet']
rowmax = ws.max_row #62
colmax = ws.max_column #1440
img = [[]for i in range(61)]
# print(img)
for j in range(2,rowmax+1):
    
    for i in range(2,colmax+1):
        img[j-2].append(ws.cell(j,i).value)

#     print (len(img[j-2]))
# print (len(img))
# print(img)



x = range(colmax-1)
y = range(rowmax-1)

windv =[-10.7,-7.9,-5.4,-3.3,-1.5,-1,0,1.5,3.3,5.4,7.9,10.7]
plt.contourf(x,y,img,levels = windv,alpha = .75,cmap = mpl.cm.jet)

# C = plt.contour(x,y,img,1,colors = 'black',linewidth=0.5)
# plt.clabel(C,inline=True,fontsize=2)
plt.xlabel('time')
plt.ylabel('height')
plt.title('2019/11/24/wind-u')
plt.colorbar()
plt.show()