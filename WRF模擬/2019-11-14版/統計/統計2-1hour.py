from openpyxl import load_workbook ,Workbook


wb = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計2/u_mean10min.xlsx")
ws = wb['Mean']
wbnew = Workbook()
wsnew = wbnew.active
newnb = 2
for i in range (7,140,6):
    for j in range(2,63):
        wsnew.cell(j,newnb).value = ws.cell(j,i).value
    newnb +=1
wbnew.save("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計2/u_mean1hour.xlsx")

wb = load_workbook("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計2/v_mean10min.xlsx")
ws = wb['Mean']
wbnew = Workbook()
wsnew = wbnew.active
newnb = 2
for i in range (7,140,6):
    for j in range(2,63):
        wsnew.cell(j,newnb).value = ws.cell(j,i).value
    newnb +=1
wbnew.save("C:/Users/steve/Desktop/python/松山剖風儀資料/統計/統計2/v_mean1hour.xlsx")