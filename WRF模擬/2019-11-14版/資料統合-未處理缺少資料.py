import csv
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import random
import matplotlib
import glob
import numpy as np
import math
import os




time = 1
wball = Workbook()
wsall = wball.active


# f = "C:/Users/steve/Desktop/python/松山剖風儀資料/u/2019-11-24_00-45-42.xlsx"
files  =glob.glob("C:/Users/steve/Desktop/python/松山剖風儀資料/u/*.xlsx")
for f in files:
    print ('u'+f[51:56])
    wb = load_workbook(f)
    ws = wb['Sheet']
# ws['A']
    for i in range(1,62):
        wsall.cell(i+1,time).value = ws.cell(i,1).value
    wsall.cell(1,time).value = f[51:56]
    time +=1
wsall.insert_cols(1)
for i in range (1,62):
    wsall.cell(i+1,1).value = 100+(i-1)*25
wball.save('C:/Users/steve/Desktop/python/松山剖風儀資料/u_all.xlsx')
 
time = 1
wball = Workbook()
wsall = wball.active


# f = "C:/Users/steve/Desktop/python/松山剖風儀資料/v/2019-11-24_00-45-42.xlsx"
files  =glob.glob("C:/Users/steve/Desktop/python/松山剖風儀資料/v/*.xlsx")
for f in files:
    print ('v'+f[51:56])
    wb = load_workbook(f)
    ws = wb['Sheet']
# ws['A']
    for i in range(1,62):
        wsall.cell(i+1,time).value = ws.cell(i,1).value
    wsall.cell(1,time).value = f[51:56]
    time +=1
wsall.insert_cols(1)
for i in range (1,62):
    wsall.cell(i+1,1).value = 100+(i-1)*25
wball.save('C:/Users/steve/Desktop/python/松山剖風儀資料/v_all.xlsx')
