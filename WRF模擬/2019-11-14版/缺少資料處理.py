from openpyxl import load_workbook
import numpy as np

file = "C:/Users/steve/Desktop/python/松山剖風儀資料/u_all.xlsx"
wb = load_workbook(file)
ws = wb['Sheet']
locate = 2
lose_files = []
lose_files_nb = []
for hour in range(0,24):
    hour  =  str(hour)
    hour = hour.zfill(2)
    # print (hour)
    for min in range(0,60):
        min = str(min)
        min = min.zfill(2)
        # print (min)
        time = hour + '-' + min
        # print (ws.cell(1,locate).value)

        if time != ws.cell(1,locate).value :
            lose_files.append(time)
            lose_files_nb.append(locate)
            locate -=1
        locate += 1     
lose_files_nb.reverse()

for i in lose_files_nb:
    ws.insert_cols(i)

    for j in range(2,63):
        ws.cell(j,i).value = ''
ws.delete_cols(2)
# wb.save("C:/Users/steve/Desktop/python/松山剖風儀資料/newu_all.xlsx")
locate = 1
for hour in range(0,24):
    hour  =  str(hour)
    hour = hour.zfill(2)
    for min in range(0,60):
        min = str(min)
        min = min.zfill(2)
        time = hour + '-' + min
        if time != ws.cell(1,locate).value :
            ws.cell(1,locate).value = time
        locate += 1
ws.cell(1,1).value = np.nan
# print (len(lose_files))
# print (len(lose_files_nb))    
# wb.save("C:/Users/steve/Desktop/python/松山剖風儀資料/new1u_all.xlsx")
wb.save(file)


file = "C:/Users/steve/Desktop/python/松山剖風儀資料/v_all.xlsx"
wb = load_workbook(file)
ws = wb['Sheet']
locate = 2
lose_files = []
lose_files_nb = []
for hour in range(0,24):
    hour  =  str(hour)
    hour = hour.zfill(2)

    for min in range(0,60):
        min = str(min)
        min = min.zfill(2)

        time = hour + '-' + min


        if time != ws.cell(1,locate).value :
            lose_files.append(time)
            lose_files_nb.append(locate)
            locate -=1
        locate += 1     
lose_files_nb.reverse()

for i in lose_files_nb:
    ws.insert_cols(i)

    for j in range(2,63):
        ws.cell(j,i).value = ''
ws.delete_cols(2)

locate = 1
for hour in range(0,24):
    hour  =  str(hour)
    hour = hour.zfill(2)
    for min in range(0,60):
        min = str(min)
        min = min.zfill(2)
        time = hour + '-' + min
        if time != ws.cell(1,locate).value :
            ws.cell(1,locate).value = time
        locate += 1
ws.cell(1,1).value = np.nan
wb.save(file)