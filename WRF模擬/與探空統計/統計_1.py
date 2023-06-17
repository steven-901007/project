from openpyxl import load_workbook,Workbook
import math
from math import *
import numpy as np
import glob

def windv(u,v): #風速
    return round(sqrt(pow(u,2) +pow(v,2)),3)

def windd(u,v): #風向
    if u == 0 and v == 0:
        return None
    else:
        return round(180+math.atan2(u,v)*180/(math.pi),2)

def ins(a,b,t,a1,b1): #內插(高度1,高度2,目標高度,時間1,時間2) 時間可換成風速風向
    return round((b1-a1)*(t-a)/(b-a)+a1,3)

def timecount(Time1,Time2):#type = mm:ss
    m1 = int(str(Time1)[:2])
    m2 = int(str(Time2)[:2])
    s1 = int(str(Time1)[3:])
    s2 = int(str(Time2)[3:])
    all =round(((m1+m2)*60+s1+s2)/2)
    m = str(all//60).zfill(2)
    s = str(all - all//60*60).zfill(2)
    return m +':'+ s
    
def timebigsmall(T): #時間轉換為數字用來比較大小(支援格式 ==> 時:分:秒)
    return int(T[:2] + T[3:5] + T[6:8])


error = []
files = glob.glob("C:/Users/steve/Desktop/python相關資料/觀測rowdata/need_information/*")
for file in files:  
    day = file[61:71]
    t = file[72:74]
    # print(day)
    # print(t)

    # day = '2019-01-01'
    # t = '00'
    try:
        wbset = load_workbook("C:/Users/steve/Desktop/python相關資料/need data information/松山/108/need_information_set/" + day + ".xlsx") #lidar
        wblook = load_workbook('C:/Users/steve/Desktop/python相關資料/觀測rowdata/need_information/' + day + '_'+t+'.xlsx') #探空
        wssetu = wbset['U']
        wssetv = wbset['V']
        wslookv = wblook['風速']
        wslookd = wblook['風向']
        maxhightinf = wssetu.max_row
        maxtimeinf = wssetu.max_column
        maxhightlook = wslookv.max_row
        # print(maxhight1hour)
        wb = Workbook()
        wsv = wb.active
        wsv.title = '風速'
        wsd = wb.create_sheet('風向')
        t = str(int(t)+8).zfill(2)
        # print(t)
        for i in range(2,maxhightinf+1):
            wsv.cell(i,1).value = wssetu.cell(i,1).value
            wsd.cell(i,1).value = wssetu.cell(i,1).value 
        wsv.cell(1,2).value = 'Time'
        wsv.cell(1,3).value = '探空'
        wsv.cell(1,4).value = 'lidar'
        wsd.cell(1,2).value = 'Time'
        wsd.cell(1,3).value = '探空'
        wsd.cell(1,4).value = 'lidar'


        list = [[],[],[]]#高度,時間,風速
        #風速
        for i in range(2,maxhightlook+1):
            list[0].append(wslookv.cell(i,1).value)
            list[1].append(wslookv.cell(i,2).value)
            list[2].append(wslookv.cell(i,3).value)
        for i in range(2,maxhightinf+1):
            for j in range(2,maxhightlook+1):
                look = float(list[0][j])
                tg = float(wsv.cell(i,1).value)
                look_1 = float(list[0][j+1])
                ti = list[2][j]
                ti_1 = list[2][j+1]
                v = float(list[1][j])
                v_1 = float(list[1][j+1])
                if look <= tg <= look_1:
                    # print(look)
                    # print(tg)
                    # print(look_1)
                    tg_ti = timecount(ti,ti_1)
                    tg_v = ins(look,look_1,tg,v,v_1)
                    wsv.cell(i,2).value = t+':'+tg_ti
                    wsv.cell(i,3).value = tg_v
                    break
        list = [[],[],[]]#高度,時間,風向
        #風向
        for i in range(2,maxhightlook+1):
            list[0].append(wslookd.cell(i,1).value)
            list[1].append(wslookd.cell(i,2).value)
            list[2].append(wslookd.cell(i,3).value)
        for i in range(2,maxhightinf+1):
            for j in range(2,maxhightlook+1):
                look = float(list[0][j])
                tg = float(wsd.cell(i,1).value)
                look_1 = float(list[0][j+1])
                ti = list[2][j]
                ti_1 = list[2][j+1]
                d = float(list[1][j])
                d_1 = float(list[1][j+1])
                if look <= tg <= look_1:
                    # print(look)
                    # print(tg)
                    # print(look_1)
                    tg_ti = timecount(ti,ti_1)
                    tg_d = ins(look,look_1,tg,d,d_1)
                    wsd.cell(i,2).value = t+':'+tg_ti
                    wsd.cell(i,3).value = tg_d
                    break
        maxhightfile = wsv.max_row
        hlc = 2

        for i in range(2,maxhightfile+1):
            for j in range(2,maxtimeinf):
                t_1 = timebigsmall(str(wssetv.cell(1,j).value))
                tg = timebigsmall(str(wsv.cell(i,2).value))
                t_2 = timebigsmall(str(wssetv.cell(1,j+1).value))

                if t_1 <= tg < t_2:
                    if (t_1+t_2)/2 >= tg:
                        u = wssetu.cell(hlc,j).value
                        v = wssetv.cell(hlc,j).value
                        if u != None and v != None:
                            wsv.cell(hlc,4).value = windv(u,v)
                            wsd.cell(hlc,4).value = windd(u,v)
                        else:
                            wsv.cell(hlc,4).value = None
                            wsd.cell(hlc,4).value = None                  
                    else:
                        u = wssetu.cell(hlc,j+1).value
                        v = wssetv.cell(hlc,j+1).value
                        if u != None and v != None:
                            wsv.cell(hlc,4).value = windv(u,v)
                            wsd.cell(hlc,4).value = windd(u,v)
                        else:
                            wsv.cell(hlc,4).value = None
                            wsd.cell(hlc,4).value = None 
                    hlc += 1
                    break
        looks = []
        lidars = []
        d_totle = 0
        d_nb = 0
        for i in range(2,maxhightfile+1):
            lookv = wsv.cell(i,3).value
            lidarv = wsv.cell(i,4).value
            if lookv != None and lidarv != None:
                looks.append(lookv)
                lidars.append(lidarv)
            lookd = wsd.cell(i,3).value
            lidard = wsd.cell(i,4).value
            if lookd != None and lidard !=None:
                d = abs(lookd-lidard)
                if 0 < d <= 180:
                    wsd.cell(i,5).value = 1-(d/90)
                elif 180 < d <= 360:
                    wsd.cell(i,5).value = (d/90)-3
                else:
                    wsd.cell(i,5).valeu = 'TypeError'
                d_totle += wsd.cell(i,5).value
                d_nb += 1
        wsv.cell(1,6).value = '相關係數'
        wsv.cell(2,6).value = np.corrcoef(looks,lidars)[0][1]

        wsd.cell(1,6).value = '相關係數'
        wsd.cell(1,5).value = '相關係數'

        wsd.cell(2,6).value = d_totle/d_nb

        wb.save("C:/Users/steve/Desktop/python相關資料/觀測rowdata/統計/"+day+'_'+t+".xlsx")
        print(day+'_'+t)
    except:
        error.append(day+'_'+t)
print(error)