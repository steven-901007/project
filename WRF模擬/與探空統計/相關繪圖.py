from openpyxl import load_workbook
import glob
import matplotlib.pyplot as plt
import numpy as np

Minv = []
Windvc = []
Winddc = []
N = []
# for i in range(0,151,1):
minv = 15 #最小風速
Minv.append(minv)
vx = [minv,30]
dx = [0,350]
wb = load_workbook("C:/Users/steve/Desktop/python相關資料/觀測rowdata/統計.xlsx")
wsvx = wb['風速lidar']
wsvy = wb['風速探空']
wsdx = wb['風向lidar']
wsdy = wb['風向探空']
hi = wsvx.max_row #64
ti = wsvx.max_column #508
# print(hi,ti)
vX = [[]for i in range(hi-1)]
dX = [[]for i in range(hi-1)]
vY = [[]for i in range(hi-1)]
dY = [[]for i in range(hi-1)]
V = [[],[]] #[0] = 風速lidar,[1] = 風速探空 算相關係數用
D = [[],[]]#[0] = 風向lidar,[1] = 風向探空 算相關係數用
for i in range(hi):
    for j in range(ti):
        if wsvx.cell(i+2,j+2).value != None and wsvy.cell(i+2,j+2).value != None and wsdx.cell(i+2,j+2).value != None and wsdy.cell(i+2,j+2).value != None:
            if wsvx.cell(i+2,j+2).value >= minv and wsvy.cell(i+2,j+2).value >= minv:
                vX[i].append(wsvx.cell(i+2,j+2).value)
                V[0].append(wsvx.cell(i+2,j+2).value)
                vY[i].append(wsvy.cell(i+2,j+2).value)
                V[1].append(wsvy.cell(i+2,j+2).value)
                dX[i].append(wsdx.cell(i+2,j+2).value)
                D[0].append(wsdx.cell(i+2,j+2).value)
                dY[i].append(wsdy.cell(i+2,j+2).value)
                D[1].append(wsdy.cell(i+2,j+2).value)

# print(vX)
# print(len(vY))

    x = np.corrcoef(V[0],V[1])
    # print(x[0][1])
    k = 0
    if len(D[0]) != 0:
        for i in range(len(D[0])):
            d = abs(D[0][i]-D[1][i])
            if 0<d<=180:
                k += 1-(d/90)
            elif 180<d<=360:
                k +=(d/90)-3




    windvc = round(x[0][1],4) #風速相關係數
    if len(D[0]) != 0:
        winddc = round(k/len(D[0]),4) #風向相關係數
    else:
        winddc = 0
    n = len(V[0]) #個案數


    Windvc.append(windvc)
    Winddc.append(winddc)
    N.append(n)
# print(Winddc)

plt.rcParams['font.sans-serif'] = [u'MingLiu'] #細明體
plt.rcParams['axes.unicode_minus'] = False #設定中文
# fig ,ax1 = plt.subplots()
# ax2 = ax1.twinx()
# curve1 ,= ax1.plot(Minv,Windvc,color =  'blue',linestyle = '--',label ='風速相關係數')
# curve2 ,= ax1.plot(Minv,Winddc,color =  'r',linestyle = '--',label ='風向相關係數')
# ax1.set_ylabel('風速、風向相關係數')
# ax2.set_ylabel('樣本個數',rotation = 270)
# ax1.set_xlabel('最小風速')
# curve3 ,= ax2.plot(Minv,N,color =  'black',linestyle = '--',label ='樣本個數')
# curves = [curve1, curve2,curve3]
# ax1.legend(curves, [curve.get_label() for curve in curves])



fig = plt.figure()
axv = fig.add_subplot()

color = ['r','orange','y','g','b','c','m']
c = 0
for j in range(2,64,9):
    # print(j)
    hi1 = wsvx.cell(j,1).value
    hi2 = wsvx.cell(j+8,1).value
    # print(hi1,hi2)
    VX = []
    VY = []
    for i in range(j-2,j+7):
        VX.extend(vX[i])
        VY.extend(vY[i])
        # print(len(vY[i]))
    # print(len(VX))
    # print(len(VY))
    axv.scatter(VX,VY,color = color[c],alpha=1-0.1*c,s=5,label = str(hi1) +'~'+str(hi2))
    c +=1
plt.plot(vx,vx,color = 'black',linestyle = '--')
axv.set_title('風速(圖例單位=m),相關係數=' + str(windvc))
axv.set_xlabel('lidar')
axv.set_ylabel('探空')
plt.legend(loc='upper right')


fig2 = plt.figure()
axd = fig2.add_subplot()

color = ['r','orange','y','g','b','c','m']
c = 0
for j in range(2,64,9):
    print(j)
    hi1 = wsdx.cell(j,1).value
    hi2 = wsdx.cell(j+8,1).value
    print(hi1,hi2)
    DX = []
    DY = []
    for i in range(j-2,j+7):
        DX.extend(dX[i])
        DY.extend(dY[i])
    axd.scatter(DX,DY,color = color[c],alpha=1-0.1*c,s=5,label = str(hi1) +'~'+str(hi2))
    c +=1
plt.plot(dx,dx,color = 'black',linestyle = '--')
axd.set_title('風向(圖例單位=m),相關係數='+str(winddc))
axd.set_xlabel('lidar')
axd.set_ylabel('探空')







plt.show()