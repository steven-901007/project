from openpyxl import Workbook,load_workbook
import csv
import glob
import math
import numpy as np
import time as T
import os



day = '2019-01-01' #時間
locate = '松山'
year = '108'

confidence_level = 100

startTime = T.time()



#先將所需要的sheet都建立好
wb_old = load_workbook("C:/Users/steve/Desktop/python相關資料/need data information/"+day+"/confidence_level="+str(confidence_level)+" need_information_read.xlsx")
ws0 = wb_old['0']
ws1 = wb_old['1']   
ws2 = wb_old['2']
ws3 = wb_old['3']
wsb = wb_old['believe'] 
wsmv_old = wb_old['mv']
wsmd_old = wb_old['md']
wb_new = Workbook()
wsu = wb_new.active
wsu.title = 'U'
wsv = wb_new.create_sheet('V')
wsmv = wb_new.create_sheet('mv')
wsmd = wb_new.create_sheet('md')
colmax = wsb.max_column #時間
rowmax = wsb.max_row #高度

for col in range(1,colmax+1):
    wsu.cell(1,col).value = ws3.cell(1,col).value
    wsv.cell(1,col).value = ws2.cell(1,col).value
    wsmv.cell(1,col).value = wsmv_old.cell(1,col).value
    wsmd.cell(1,col).value = wsmd_old.cell(1,col).value
for row in range(1,rowmax+1):
    wsu.cell(row,1).value = ws3.cell(row,1).value
    wsv.cell(row,1).value = ws2.cell(row,1).value
    wsmv.cell(row,1).value = wsmv_old.cell(row,1).value
    wsmd.cell(row,1).value = wsmd_old.cell(row,1).value 
for col in range(2,colmax+1):
    for row in range(2,rowmax+1):
        if wsb.cell(row,col).value == 4: #符合信賴度的才計算
            wsu.cell(row,col).value = round((ws3.cell(row,col).value-ws1.cell(row,col).value)/(2*math.cos(75*np.pi/180)),2) # u = (W-E)/2cos(75度)
            wsv.cell(row,col).value = round((ws2.cell(row,col).value-ws0.cell(row,col).value)/(2*math.cos(75*np.pi/180)),2) # v = (S-N)/2cos(75度)
            wsmv.cell(row,col).value = wsmv_old.cell(row,col).value/4
            wsmd.cell(row,col).value = wsmd_old.cell(row,col).value/4
        else:
            wsmv.cell(row,col).value = None
            wsmd.cell(row,col).value = None      
print('set')

#檔案檢查
chack = []
for i in range(0,24):
    for j in range(00,60):
        i = str(i).zfill(2)
        j = str(j).zfill(2)
        chack.append(i+":"+j)
# print(chack)

wsu.insert_cols(2)
wsv.insert_cols(2)
wsmv.insert_cols(2)
wsmd.insert_cols(2)
wsu.cell(1,2).value = '00:00:00'
wsv.cell(1,2).value = '00:00:00'
wsmv.cell(1,2).value = '00:00:00'
wsmd.cell(1,2).value = '00:00:00'
wsu.cell(1,wsu.max_column+1).value = '23:59:59'
wsv.cell(1,wsv.max_column+1).value = "23:59:59"
wsmv.cell(1,wsu.max_column+1).value = '23:59:59'
wsmd.cell(1,wsv.max_column+1).value = "23:59:59"
timechack = []
for i in range(2,colmax+1):
    timechack.append(wsv.cell(1,i).value[:5])
# print(timechack)
for i in range(len(chack)):

    if timechack.count(chack[i]) != 0:
        chack[i] = 999

while chack.count(999) != 0:
    chack.remove(999) 

for i in range(len(chack)):

    j = 2
    while wsu.cell(1,j).value < chack[i]:
        j +=1
    wsu.insert_cols(j)
    wsv.insert_cols(j)
    wsmv.insert_cols(j)
    wsmd.insert_cols(j)
    wsu.cell(1,j).value = chack[i]
    wsv.cell(1,j).value = chack[i]
    wsmv.cell(1,j).value = chack[i]
    wsmd.cell(1,j).value = chack[i]



if chack == []:
    print("無缺失時間")
else:
    print(chack)
wb_new.save("C:/Users/steve/Desktop/python相關資料/need data information/"+day+"/confidence_level="+str(confidence_level)+" need_information_set.xlsx")
wb_new.close()
print(day+'\n'+str(confidence_level))
endTime = T.time()

print('程式執行了%s秒' %(endTime-startTime))
