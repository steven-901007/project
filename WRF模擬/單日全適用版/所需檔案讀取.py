from openpyxl import Workbook
import csv
import glob
import math
import numpy as np
import time as T
import os



day = '2019-01-01' #時間
locate = '松山'
year = '108'

confidence_level = 100

startTime = T.time()

# 建立存檔位置
try:
    os.makedirs("C:/Users/steve/Desktop/python相關資料/need data information/"+day)
except:
    pass
#先將所需要的sheet都建立好
wb = Workbook()
ws0 = wb.active
ws0.title = '0'
wb.create_sheet('1')
wb.create_sheet('2')
wb.create_sheet('3')
wb.create_sheet('believe')
wb.create_sheet("mv")
wb.create_sheet("md")

ws1 = wb['1']   
ws2 = wb['2']
ws3 = wb['3']
wsb = wb['believe']
wsmv = wb['mv']
wsmd = wb['md']
#確定高度的最大最小值(假設同一天所有資料的高度size都一樣,並且間距為25)

hight = []

files  =glob.glob("C:/Users/steve/Desktop/python相關資料/raw data/" + locate + "/" + year + "/**")
for file in files:
    break
# print(file)
fs = glob.glob(file+'/wind_reconstruction_data/**')
for f in fs:
    break
# print(f)
fh = glob.glob(f + '/*.csv')
for f_hight in fh:
    break
# print(f_hight)

with open(f_hight,newline='') as csvfile:
    rows = csv.DictReader(csvfile,delimiter=';')
    for row in rows:
        hight.append(int(row['Altitude [m]']))
# print(min(hight),max(hight))
gap = int((max(hight)-min(hight))/25+1) #設定(01234)的資料間距
b = 2 #起始高度的cellx
c = min(hight)
# print (gap)
#建立高度座標
while c <=max(hight) :
    ws0.cell(b,1).value = c
    ws1.cell(b,1).value = c
    ws2.cell(b,1).value = c
    ws3.cell(b,1).value = c
    wsb.cell(b,1).value = c
    wsmv.cell(b,1).value = c
    wsmd.cell(b,1).value = c
    # print (c)
    b +=1 
    c += 25 #高度間格

time = 2 #00-00起始X座標
y = 2 #時間起始Y座標
#建立時間座標
files  =glob.glob("C:/Users/steve/Desktop/python相關資料/raw data/" + locate + "/" + year + "/"+day+"/wind_reconstruction_data/*")
for file in files:
        file = file + '/*.csv'

        result  =glob.glob(file)
        
        for f in result: 
            #同一個檔案(時間)
            #設定時間
            handm = f[137:139]+':'+f[140:142]+':'+f[143:145] #時間
            # print(handm)
            ws0.cell(1,y).value = handm
            ws1.cell(1,y).value = handm
            ws2.cell(1,y).value = handm
            ws3.cell(1,y).value = handm
            wsb.cell(1,y).value = handm
            wsmv.cell(1,y).value = handm
            wsmd.cell(1,y).value = handm
            y += 1
            # print(f)
            zero = 2
            one = 2
            two = 2
            three = 2

            with open(f,newline='') as csvfile:
                rows = csv.DictReader(csvfile,delimiter=';')
                for row in rows:
                    believe = int(row['Confidence Index [%]'])
                    singo = int(round(float(row['Azimuth [財'])))
                    
                    #排除 RV
                    if int(round(float(row['Elevation [財'])))== 75:
                    #     #確認所置放的Sheet
                        
                        value = float(row['Radial Wind Speed [m/s]'])
                        v = float(row['Horizontal Wind Speed [m/s]'])
                        d = float(row['Horizontal Wind Direction [財'])
                        if singo == 0 or singo == 360:
                            ws0.cell(zero,time).value = value
                            if wsmv.cell(zero,time).value != None:
                                wsmv.cell(zero,time).value += v
                            else:
                                wsmv.cell(zero,time).value = v

                            if wsmd.cell(zero,time).value != None:
                                wsmd.cell(zero,time).value += d
                            else:
                                wsmd.cell(zero,time).value = d
                            if believe >= confidence_level:
                                if wsb.cell(zero,time).value !=None:
                                    wsb.cell(zero,time).value += 1
                                else:
                                    wsb.cell(zero,time).value =1
                            zero +=1  
                                                    
                        elif singo == 90:
                            ws1.cell(one,time).value = value
                            if wsmv.cell(one,time).value != None:
                                wsmv.cell(one,time).value += v
                            else:
                                wsmv.cell(one,time).value = v
                            if wsmd.cell(one,time).value != None:
                                wsmd.cell(one,time).value += d
                            else:
                                wsmd.cell(one,time).value = d
                            if believe >= confidence_level:
                                if wsb.cell(one,time).value !=None:
                                    wsb.cell(one,time).value += 1
                                else:
                                    wsb.cell(one,time).value =1
                            one +=1

                        elif singo == 180:
                            ws2.cell(two,time).value = value
                            if wsmv.cell(two,time).value != None:
                                wsmv.cell(two,time).value += v
                            else:
                                wsmv.cell(two,time).value = v
                            if wsmd.cell(two,time).value != None:
                                wsmd.cell(two,time).value += d
                            else:
                                wsmd.cell(two,time).value = d
                            if believe >= confidence_level:
                                if wsb.cell(two,time).value !=None:
                                    wsb.cell(two,time).value += 1
                                else:
                                    wsb.cell(two,time).value =1
                            two +=1
                            
                        elif singo == 270:
                            ws3.cell(three,time).value = value
                            if wsmv.cell(three,time).value != None:
                                wsmv.cell(three,time).value += v
                            else:
                                wsmv.cell(three,time).value = v
                            if wsmd.cell(three,time).value != None:
                                wsmd.cell(three,time).value += d
                            else:
                                wsmd.cell(three,time).value = d
                            if believe >= confidence_level:
                                if wsb.cell(three,time).value !=None:
                                    wsb.cell(three,time).value += 1
                                else:
                                    wsb.cell(three,time).value =1
                            three +=1

            time +=1 #部分資料缺失時(例如2019-11-24_01-39)缺失的部分會呈現空白
colmax = wsb.max_column #時間
rowmax = wsb.max_row #高度
# print(colmax,rowmax)
for row in range(2,rowmax+1):
    for col in range(2,colmax+1):
        if wsb.cell(row,col).value == None:
            wsb.cell(row,col).value = 0
            
wb.save("C:/Users/steve/Desktop/python相關資料/need data information/"+day+"/confidence_level="+str(confidence_level)+" need_information_read.xlsx")
wb.close()
print(day+'\n'+str(confidence_level))
endTime = T.time()

print('程式執行了%s秒' %(endTime-startTime))
