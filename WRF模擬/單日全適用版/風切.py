from openpyxl import load_workbook,Workbook
import math
import glob
import time as Ti

def windcut(u1,v1,h1,u2,v2,h2):
    return pow((pow(((u1-u2)/(h1-h2)),2) + pow(((v1-v2)/(h1-h2)),2)),(1/2)) #  風切 = ( (du/dx)^2 + (dv/dx)^2 )^(1/2)


locate = '松山'
year = '108'
day = "2019-04-20"
confidence_level = 100
sigma = 1
startTime = Ti.time()


try:
    wb10 = load_workbook("C:/Users/steve/Desktop/python相關資料/need data information/"+day+"/confidence_level="+str(confidence_level)+"sigma="+str(sigma)+" 10min.xlsx")
    wsu = wb10['U']
    wsv = wb10['V']
    hightmax10 = wsu.max_row
    timemax = wsu.max_column
    # print(hightmax10,timemax)
    wb = Workbook()
    ws = wb.active
    ws.title = '風切'
    for i in range(2,hightmax10+1):
        ws.cell(2*i-2,1).value = wsu.cell(i,1).value

    for i in range(2,timemax+1):
        ws.cell(1,i).value = wsu.cell(1,i).value

    for i in range(3,hightmax10*2-2,2):
        ws.cell(i,1).value = (ws.cell(i-1,1).value + ws.cell(i+1,1).value)/2    


    t = 2
    hightmax = ws.max_row

    for time in range(2,timemax+1):
        sametime = [[],[],[]] #高度,U,V
        for h in range(2,hightmax10+1):
            tight = wsu.cell(h,1).value
            u     = wsu.cell(h,time).value
            v     = wsv.cell(h,time).value
            if u != None and v != None:
                sametime[0].append(tight)
                sametime[1].append(u)
                sametime[2].append(v)
        # print(sametime[0])
        # print(sametime[1])
        # print(sametime[2])
        cut = [[],[]] #高度,風切
        # try:
        if len(sametime[0]) > 1:
            for i in range(len(sametime[0])-1):
                cut[1].append(round(windcut(sametime[1][i],sametime[2][i],sametime[0][i],sametime[1][i+1],sametime[2][i+1],sametime[0][i+1]),3))
                cut[0].append((sametime[0][i+1]+sametime[0][i])/2)
            # print(cut)
            # print(cut[0])
            for i in range(len(cut[0])):
                for j in range(2,hightmax+1):
                    if cut[0][i] == ws.cell(j,1).value :
                        ws.cell(j,t).value = cut[1][i]
                        # print(ws.cell(j,1).value)
            
            t +=1
        else:
            t +=1
    wb.save("C:/Users/steve/Desktop/python相關資料/need data information/"+day+"/confidence_level="+str(confidence_level)+"sigma="+str(sigma)+" windcut.xlsx")
    print(day+'\nconfidence '+str(confidence_level)+'\nsigma '+str(sigma))
except:print(day + 'ERROR')
endTime = Ti.time()

print('程式執行了%s秒' %(endTime-startTime))