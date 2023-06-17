from openpyxl import load_workbook,Workbook
import time as T

locate = '松山'
year = '108'
day = '2019-03-05'
confidence_level = 100
sigma = 1

startTime = T.time()


wb = load_workbook("C:/Users/steve/Desktop/python相關資料/need data information/"+day+"/confidence_level="+str(confidence_level)+"sigma="+str(sigma)+" 10min.xlsx")
wsu = wb['U']
wsv = wb['V']
wb1 = Workbook()
ws1u = wb1.active
ws1u.title = 'U'
ws1v = wb1.create_sheet('V')

lc = 2
for time in range(7,140,6):    
    # print(wsu.cell(1,time).value)
    for hight in range(1,65):
        # print(wsu.cell(hight,time).value)
        ws1u.cell(hight,lc).value = wsu.cell(hight,time).value
    lc+=1
lc = 2
for time in range(7,140,6):    
    # print(wsu.cell(1,time).value)
    for hight in range(1,65):
        # print(wsu.cell(hight,time).value)
        ws1v.cell(hight,lc).value = wsv.cell(hight,time).value
    lc+=1

for hi in range(2,wsu.max_row+1):
    ws1u.cell(hi,1).value = wsu.cell(hi,1).value
    ws1v.cell(hi,1).value = wsv.cell(hi,1).value

wb1.save("C:/Users/steve/Desktop/python相關資料/need data information/"+day+"/confidence_level="+str(confidence_level)+"sigma="+str(sigma)+" 1hour.xlsx")
print(day+'\nconfidence '+str(confidence_level)+'\nsigma '+str(sigma))
# wb1.create_sheet('v')

endTime = T.time()

print('程式執行了%s秒' %(endTime-startTime))