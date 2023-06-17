from openpyxl import load_workbook,Workbook
import time as T
import glob


locate = '松山'
year = '108'


startTime = T.time()

files = glob.glob('C:/Users/steve/Desktop/python相關資料/need data information/' + locate + '/' + year + '/10min/*')
for file in files:

    day = file[69:79]
    print(day)
    wb = load_workbook("C:/Users/steve/Desktop/python相關資料/need data information/" + locate + '/' + year + "/10min/"+day+".xlsx")
    wsu = wb['U']
    wsv = wb['V']
    wsmv = wb['mv']
    wsmd = wb['md']
    wb1 = Workbook()
    ws1u = wb1.active
    ws1u.title = 'U'
    ws1v = wb1.create_sheet('V')
    ws1mv = wb1.create_sheet('mv')
    ws1md = wb1.create_sheet('md')
    lc = 2
    for time in range(7,140,6):    
        # print(wsu.cell(1,time).value)
        for hight in range(1,65):
            # print(wsu.cell(hight,time).value)
            ws1u.cell(hight,lc).value = wsu.cell(hight,time).value
            ws1v.cell(hight,lc).value = wsv.cell(hight,time).value
            ws1mv.cell(hight,lc).value = wsmv.cell(hight,time).value
            ws1md.cell(hight,lc).value = wsmd.cell(hight,time).value
        lc+=1

    for hi in range(2,wsu.max_row+1):
        ws1u.cell(hi,1).value = wsu.cell(hi,1).value
        ws1v.cell(hi,1).value = wsv.cell(hi,1).value
        ws1mv.cell(hi,1).value = wsmv.cell(hi,1).value
        ws1md.cell(hi,1).value = wsmd.cell(hi,1).value

    wb1.save("C:/Users/steve/Desktop/python相關資料/need data information/" + locate + '/' + year + "/1hour/"+day+".xlsx")

# wb1.create_sheet('v')

endTime = T.time()

print('程式執行了%s秒' %(endTime-startTime))