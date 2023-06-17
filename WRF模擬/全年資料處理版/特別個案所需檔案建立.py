from openpyxl import load_workbook
import math
import numpy as np




wb = load_workbook ("C:/Users/steve/Desktop/python相關資料/need data information/need_information/2019-06-06.xlsx")
ws0 = wb['0']
ws1 = wb['1']   
ws2 = wb['2']
ws3 = wb['3']
wsb = wb['believe'] 
wsu = wb["U"]
wsv = wb["V"]


# timemax =  2637#ws1.max_column #2637
# hightmax  = 151#ws1.max_row #151
# # print(timemax,hightmax)

# for i in range(2,timemax+1):
#     print(timemax+2-i)
#     ws0.insert_cols(timemax+2-i)
#     ws1.insert_cols(timemax+2-i)
#     ws2.insert_cols(timemax+2-i)
#     ws3.insert_cols(timemax+2-i)
#     wsb.insert_cols(timemax+2-i)
# for i in range(2,ws0.max_column,2):
#     for j in range(2,77):
#         ws0.cell(j,i).value = ws0.cell(j+75,i+1).value
#         ws0.cell(j+75,i+1).value = None
#         ws1.cell(j,i).value = ws1.cell(j+75,i+1).value
#         ws1.cell(j+75,i+1).value = None
#         ws2.cell(j,i).value = ws2.cell(j+75,i+1).value
#         ws2.cell(j+75,i+1).value = None
#         ws3.cell(j,i).value = ws3.cell(j+75,i+1).value
#         ws3.cell(j+75,i+1).value = None
#         wsb.cell(j,i).value = wsb.cell(j+75,i+1).value
#         wsb.cell(j+75,i+1).value = None
# for i in range(2,ws0.max_column,2):
#     ws0.cell(1,i).value = ws0.cell(1,i+1).value
#     ws1.cell(1,i).value = ws1.cell(1,i+1).value
#     ws2.cell(1,i).value = ws2.cell(1,i+1).value
#     ws3.cell(1,i).value = ws3.cell(1,i+1).value
#     wsb.cell(1,i).value = wsb.cell(1,i+1).value

colmax = wsb.max_column #時間
rowmax = wsb.max_row #高度



# for col in range(2,colmax+1):
#     for row in range(2,rowmax+1):
#         if wsb.cell(row,col).value == 4: #符合信賴度的才計算
#             wsu.cell(row,col).value = round((ws3.cell(row,col).value-ws1.cell(row,col).value)/(2*math.cos(75*np.pi/180)),2) # u = (W-E)/2cos(75度)
#             wsv.cell(row,col).value = round((ws2.cell(row,col).value-ws0.cell(row,col).value)/(2*math.cos(75*np.pi/180)),2) # v = (S-N)/2cos(75度)
chack = []
for i in range(0,24):
    for j in range(00,60):
        i = str(i).zfill(2)
        j = str(j).zfill(2)
        chack.append(i+"-"+j)
# print(chack)

wsu.insert_cols(2)
wsv.insert_cols(2)
wsu.cell(1,2).value = '00-00'
wsv.cell(1,2).value = '00-00'
wsu.cell(1,wsu.max_column+1).value = '23-59'
wsv.cell(1,wsv.max_column+1).value = "23-59"
timechack = []
for i in range(2,colmax+1):
    timechack.append(wsv.cell(1,i).value)
# print(timechack)
for i in range(len(chack)):

    if timechack.count(chack[i]) != 0:
        chack[i] = 999

while chack.count(999) != 0:
    chack.remove(999) 

for i in range(len(chack)):

    j = 2
    while wsu.cell(1,j).value < chack[i]:
        j +=1
    wsu.insert_cols(j)
    wsv.insert_cols(j)
    wsu.cell(1,j).value = chack[i]
    wsv.cell(1,j).value = chack[i]




if chack == []:
    print("無缺失時間")
else:
    print(chack)
wb.save("C:/Users/steve/Desktop/python相關資料/2019-06-06.xlsx")