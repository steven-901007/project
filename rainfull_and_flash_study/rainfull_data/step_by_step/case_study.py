import pandas as pd
from openpyxl import Workbook
from glob import glob
from tqdm import tqdm
year = '2021' #年分
month = '06' #月份
data_top_path = "C:/Users/steve/python_data"
dis = 36

station = 'C0S910'

##save data set up
save_data_wb = Workbook()
save_data_ws = save_data_wb.active

print(station)
##測站名稱建立
around_station_data_path = data_top_path + "/研究所/雨量資料/"+year+"測站範圍內測站數/"+station+".csv"
around_station_datas = pd.read_csv(around_station_data_path)
around_station_datas['station name'] = around_station_datas['station name'].astype(str)
# print(around_station_datas)
row = 2
for around_station in around_station_datas['station name']:
    save_data_ws.cell(row,1).value = around_station
    row += 1
del row

row_for_around_station_datas_lsit = around_station_datas['station name'].to_list()


##測站真名
real_name_path = data_top_path + "/研究所/雨量資料/對流性降雨次數.csv"
datas = pd.read_csv(real_name_path)
real_name = str(datas[datas['station name'] == station]['real name'].iloc[0])
print(real_name)

##資料時間建立
col = 2
data_paths = data_top_path + "/研究所/雨量資料/對流性降雨data/"+year+"/"+month+"/**.csv"
result = glob(data_paths)
for data_path in tqdm(result,'資料寫入中....'):
# data_path = result[1]
    # print(data_path)
    time = data_path.split('/')[-1].split('\\')[-1].split('.')[0][6:]
    # print(time)
    
    data = pd.read_csv(data_path)
    data['station name'] = data['station name'].astype(str)
    intersection = pd.merge(around_station_datas, data, on='station name')
    if not intersection.empty:
        save_data_ws.cell(1,col).value = time
        # print(time)
        # print(intersection)
        
        for i in range(intersection['station name'].count()):
            # print(i)
            save_data_ws.cell(row_for_around_station_datas_lsit.index(intersection['station name'][i])+2,col).value = intersection['rain data'][i]
        

        col += 1


save_data_ws.title = real_name
save_data_path = data_top_path + "/研究所/雨量資料/"+str(dis)+"km個案分析/" + month + '/' + station + '.xlsx'
save_data_wb.save(save_data_path)