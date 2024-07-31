from openpyxl import load_workbook
import glob
import re
import pandas as pd
import os
from tqdm import tqdm

year = '2021' #年分
month = '06' #月份
data_top_path = "C:/Users/steve/python_data"
dis = 36

##建立資料夾
def file_set():
    file_path = data_top_path + "/研究所/雨量資料/對流性降雨"+str(dis)+"km統計/"+year +'/' + month
    if not os.path.exists(file_path):
            os.makedirs(file_path)
            print(file_path + " 已建立")
file_set()

## 測站數file 讀取
station_number_path = data_top_path+"/研究所/雨量資料/"+year+"測站範圍內測站數.xlsx"
wb_station_number = load_workbook(station_number_path)
ws_station_number = wb_station_number[month]
station_name_count = ws_station_number.max_column #站點數量
#建立測站在excel位置的list
station_lc_in_excel_list = [] #list的位置跟excel上的位置差1 list+1 = excel
for lc in range(station_name_count):
    station_lc_in_excel_list.append(ws_station_number.cell(1,lc+1).value)



#儲存list
rain_data_36km_sta_list_list = [[] for i in range(station_name_count)] ##二階list
# print(rain_data_36km_sta_list_list)


##處理雨量資料(未做36km統計)
rain_data_paths = data_top_path + "/研究所/雨量資料/對流性降雨data/"+year+"/"+month+"/**.csv"

result  =glob.glob(rain_data_paths)

for rain_data_path in tqdm(result,desc='資料讀取+紀錄'):

    time = rain_data_path.split('\\')[-1].split('.')[0]
    # print(time)
    rain_datas = pd.read_csv(rain_data_path, dtype=str)
    # print(rain_datas)

    for index, row in rain_datas.iterrows():
        rain_data_station_name = row['station name']
        # print(row['station name'])
        data_raw = station_lc_in_excel_list.index(rain_data_station_name) + 1
        data_col = 4
        while ws_station_number.cell(data_col,data_raw).value != None:
            # print(ws_station_number.cell(data_col,data_raw).value)
            data = ws_station_number.cell(data_col,data_raw).value
            save_list_lc = rain_data_36km_sta_list_list[station_lc_in_excel_list.index(data)]
            if save_list_lc.count(time) == 0:
                save_list_lc.append(time)
            


            data_col += 1
# print(rain_data_36km_sta_list_list)
wb_station_number.close()

##資料建立
for station in tqdm(range(len(rain_data_36km_sta_list_list)),desc='資料建立'):
    if rain_data_36km_sta_list_list[station] != []:
        save_data = {
            'time data':rain_data_36km_sta_list_list[station]
        }
        save_data['time data'] = pd.to_datetime(save_data['time data']).strftime('%Y/%m/%d %H:%M')

        save_path = data_top_path + "/研究所/雨量資料/對流性降雨"+str(dis)+"km統計/"+year +'/' + month +'/' + station_lc_in_excel_list[station] + '.csv'
        pd.DataFrame(save_data, dtype=str).to_csv(save_path,index=False)