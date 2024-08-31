from importset import fileset
import calendar
from tqdm import tqdm
from openpyxl import load_workbook
import math
import pandas as pd 
from geopy.distance import geodesic


year = '2021' #年分
month = '06' #月份
data_top_path = "C:/Users/steve/python_data"
dis = 36


## 讀取雨量站經緯度資料
def rain_station_location_data():
    data_path = data_top_path+"/研究所/雨量資料/"+year+"測站範圍內測站數.xlsx"
    lon_data_list = []  # 經度
    lat_data_list = []  # 緯度
    name_data_list = []  #測站名稱
    wb = load_workbook(data_path)
    ws = wb[month]
    for i in range(ws.max_column):
        lon_data_list.append(ws.cell(4,i+1).value)
        lat_data_list.append(ws.cell(3,i+1).value)
        name_data_list.append(ws.cell(1,i+1).value)
    wb.close()
    return lon_data_list, lat_data_list ,name_data_list
lon_data_list, lat_data_list ,name_data_list = rain_station_location_data()


## 讀取閃電資料
flash_data_path = data_top_path+'/研究所/閃電資料/raw_data/'+year+'/'+year+month+'.txt'
flash_rawdata = pd.read_csv(flash_data_path,header = 0)

# print(flash_rawdata['日期時間'],flash_rawdata['經度'],flash_rawdata['緯度'])

    
fileset(data_top_path + "/研究所/閃電資料/依測站分類/"+str(dis)+'km/'+year+ '/' + month)


for station_nb in tqdm(range(len(name_data_list)),desc='寫入資料'):
# station_nb = 0
    station_name = name_data_list[station_nb]
    station_lon = lon_data_list[station_nb]
    station_lat = lat_data_list[station_nb]
    station_lat_lon = (station_lat,station_lon)
    # print(station_name)


    flash_rawdata['distance_km'] = flash_rawdata.apply(lambda row: geodesic((row['緯度'], row['經度']), station_lat_lon).km, axis=1)
    # print(flash_rawdata)

    target_data = pd.to_datetime(flash_rawdata[flash_rawdata['distance_km'] < dis]['日期時間'])#確認符合範圍的時間
    target_data = target_data.dt.floor('min') #時間精度到分
    target_data = target_data + pd.Timedelta(minutes=1) #時間分+1
    target_data = pd.DataFrame(target_data)
    # target_data_count = target_data.value_counts().reset_index()

    target_data['count'] = target_data.groupby(['日期時間'])['日期時間'].transform('size')
    target_data =target_data.drop_duplicates(subset=['日期時間']) #捨棄重複值
    # print(target_data)

    flash_data_to_save = {
        'data time' : target_data['日期時間'],
        'count' : target_data['count']
    }
    flash_data_to_path = data_top_path + "/研究所/閃電資料/依測站分類/"+str(dis)+'km/'+year+ '/' + month + '/' + station_name + '.csv'
    pd.DataFrame(flash_data_to_save).to_csv(flash_data_to_path,index=False)