##前估
from openpyxl import load_workbook
import glob
import pandas as pd
from datetime import datetime, timedelta
from tqdm import tqdm
import matplotlib.pyplot as plt
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from cartopy.io.shapereader import Reader
from cartopy.feature import ShapelyFeature
import matplotlib.colors as mcolors
import matplotlib.cm as cm
import matplotlib as mpl



year = '2021' #年分
month = '06' #月份
dis = 36
data_top_path = "C:/Users/steve/python_data"

##測站資料
def rain_station_location_data():
    data_path = data_top_path+"/研究所/雨量資料/"+year+"測站範圍內測站數.xlsx"
    lon_data_list = []  # 經度
    lat_data_list = []  # 緯度
    name_data_list = []  #測站名稱
    wb = load_workbook(data_path)
    ws = wb[month]
    for i in range(ws.max_column):
        lon_data_list.append(ws.cell(4,i+1).value)
        lat_data_list.append(ws.cell(3,i+1).value)
        name_data_list.append(ws.cell(1,i+1).value)
    wb.close()
    return lon_data_list, lat_data_list ,name_data_list
lon_data_list, lat_data_list ,name_data_list = rain_station_location_data()

def check_in_time_range(row, lj_times):
    return int(any((lj_times >= row['start time']) & (lj_times <= row['end time'])))

##強降雨發生但沒有lighting jump


prefigurance_station_name_list = []#前估測站名稱
prefigurance_hit_list = []#個測站命中的list
total_prefigurance_list = []#前估總量(lighting jump and rain + non_lighting jump and rain)
prefigurance_lon_data_list = []
prefigurance_lat_data_list = []

month_path = data_top_path + "/研究所/雨量資料/對流性降雨"+str(dis)+"km統計/"+year+"/"+month+"/**.csv"
result  =glob.glob(month_path)

for rain_station_path in tqdm(result,desc='資料處理中....'):
# rain_station_path = "C:/Users/steve/python_data/研究所/雨量資料/對流性降雨36km統計/2021/06/C0V730.csv"
    rain_station_name = rain_station_path[56:62]
    # print(rain_station_name)

    #flash
    try:
        flash_station_path = data_top_path + "/研究所/閃電資料/lighting_jump/半徑"+str(dis)+"km/"+year+"/"+month+"/"+rain_station_name+".csv"
    except:
        flash_station_path = None



    if flash_station_path != None:
        rain_data = pd.read_csv(rain_station_path)
        flash_data = pd.read_csv(flash_station_path)
        #end time< lighting jump <= time data
        rain_data['time data'] = pd.to_datetime(rain_data['time data'])
        rain_data['start time'] = rain_data["time data"] - pd.Timedelta(minutes= 50)
        rain_data['end time'] = rain_data['time data'] + pd.Timedelta(minutes=10)
        flash_data['LJ_time'] = pd.to_datetime(flash_data['LJ_time'])
        # print(rain_data)

        rain_data['LJ_in_time_range'] = rain_data.apply(lambda row: check_in_time_range(row, flash_data['LJ_time']), axis=1)

        # print(rain_data[rain_data['LJ_in_time_range'] == 1])
        pd.set_option('display.max_rows', None)

        # print(rain_data)
        # print(rain_data['LJ_in_time_range'].sum())
        if rain_data['LJ_in_time_range'].sum() != 0:
            prefigurance_station_name_list.append(rain_station_name)
            prefigurance_hit_list.append(rain_data['LJ_in_time_range'].sum())
            total_prefigurance_list.append(len(rain_data))
            prefigurance_lon_data_list.append(lon_data_list[name_data_list.index(rain_station_name)])
            prefigurance_lat_data_list.append(lat_data_list[name_data_list.index(rain_station_name)])
# print(rain_station_name,rain_data['LJ_in_time_range'].sum(),len(rain_data))
# print(prefigurance_hit_list)
# print(total_prefigurance_list)

prefigurance_hit_persent_list = [] # 前估命中率
for i in range(len(total_prefigurance_list)):
    prefigurance_hit_persent_list.append(prefigurance_hit_list[i]/(total_prefigurance_list[i]+prefigurance_hit_list[i])*100)

##前估繪圖

# 設定經緯度範圍
lon_min, lon_max = 120, 122.1
lat_min, lat_max = 21.5, 25.5

plt.figure(figsize=(10, 10))
ax = plt.axes(projection=ccrs.PlateCarree())
ax.set_xlim(lon_min, lon_max)
ax.set_ylim(lat_min, lat_max)

plt.rcParams['font.sans-serif'] = [u'MingLiu']  # 設定字體為'細明體'
plt.rcParams['axes.unicode_minus'] = False  # 用來正常顯示正負號

# 加載台灣的行政邊界
taiwan_shapefile = data_top_path+"/研究所/Taiwan_map_data/COUNTY_MOI_1090820.shp"  # 你需要提供台灣邊界的shapefile文件
shape_feature = ShapelyFeature(Reader(taiwan_shapefile).geometries(),
                               ccrs.PlateCarree(), edgecolor='black', facecolor='white')
ax.add_feature(shape_feature)


# 加入經緯度格線
gridlines = ax.gridlines(draw_labels=True, linestyle='--')
gridlines.top_labels = False
gridlines.right_labels = False

## 計算某個地方達到10mm/10min的次數 + colorbar
color_list = []

level = [0,3,5,7,10,15,20,25,30]
color_box = ['silver','purple','darkviolet','blue','g','y','orange','r']

for nb in prefigurance_hit_list:
    more_then_maxma_or_not = 0
    for j in range(len(level)-1):
        if level[j]<nb<=level[j+1]:
            color_list.append(color_box[j])
            more_then_maxma_or_not = 1
            break
    if more_then_maxma_or_not == 0:
        color_list.append('lime')
        # print(nb)
# print(len(color_list))


# 標記經緯度點
ax.scatter(prefigurance_lon_data_list, prefigurance_lat_data_list, color=color_list, s=3, zorder=5)

# colorbar setting

nlevel = len(level)
cmap1 = mpl.colors.ListedColormap(color_box, N=nlevel)
cmap1.set_over('fuchsia')
cmap1.set_under('black')
norm1 = mcolors.Normalize(vmin=min(level), vmax=max(level))
norm1 = mcolors.BoundaryNorm(level, nlevel, extend='max')
im = cm.ScalarMappable(norm=norm1, cmap=cmap1)
cbar1 = plt.colorbar(im,ax=ax, extend='neither', ticks=level)


# 加入標籤
plt.xlabel('Longitude')
plt.ylabel('Latitude')

ax.set_title(year+"年"+month+"月"+'\n前估 max = '+ str(max(prefigurance_hit_list)))


## 這是用來確認colorbar的配置
fig,ax1 = plt.subplots()
X = [i for i in range(len(prefigurance_hit_list))]
Y = sorted(prefigurance_hit_list)
ax1.plot(X,Y,color =  'black',marker = "*",linestyle = '--') #折線圖
ax1.set_title('這是用來確認colorbar的配置')



##前估命中率繪圖

plt.figure(figsize=(10, 10))
ax = plt.axes(projection=ccrs.PlateCarree())
ax.set_xlim(lon_min, lon_max)
ax.set_ylim(lat_min, lat_max)

plt.rcParams['font.sans-serif'] = [u'MingLiu']  # 設定字體為'細明體'
plt.rcParams['axes.unicode_minus'] = False  # 用來正常顯示正負號

# 加載台灣的行政邊界
taiwan_shapefile = data_top_path+"/研究所/Taiwan_map_data/COUNTY_MOI_1090820.shp"  # 你需要提供台灣邊界的shapefile文件
shape_feature = ShapelyFeature(Reader(taiwan_shapefile).geometries(),
                               ccrs.PlateCarree(), edgecolor='black', facecolor='white')
ax.add_feature(shape_feature)


# 加入經緯度格線
gridlines = ax.gridlines(draw_labels=True, linestyle='--')
gridlines.top_labels = False
gridlines.right_labels = False

## 計算某個地方達到10mm/10min的次數 + colorbar
color_list = []

level = [0,0.5,1,2,3,4,5,6,7]
color_box = ['silver','purple','darkviolet','blue','g','y','orange','r']

for nb in prefigurance_hit_persent_list:
    more_then_maxma_or_not = 0
    for j in range(len(level)-1):
        if level[j]<nb<=level[j+1]:
            color_list.append(color_box[j])
            more_then_maxma_or_not = 1
            break
    if more_then_maxma_or_not == 0:
        color_list.append('lime')
        # print(nb)
# print(len(color_list))


# 標記經緯度點
ax.scatter(prefigurance_lon_data_list, prefigurance_lat_data_list, color=color_list, s=3, zorder=5)

# colorbar setting

nlevel = len(level)
cmap1 = mpl.colors.ListedColormap(color_box, N=nlevel)
cmap1.set_over('fuchsia')
cmap1.set_under('black')
norm1 = mcolors.Normalize(vmin=min(level), vmax=max(level))
norm1 = mcolors.BoundaryNorm(level, nlevel, extend='max')
im = cm.ScalarMappable(norm=norm1, cmap=cmap1)
cbar1 = plt.colorbar(im,ax=ax, extend='neither', ticks=level)


# 加入標籤
plt.xlabel('Longitude')
plt.ylabel('Latitude')

ax.set_title(year+"年"+month+"月"+'\n前估命中率 [%] max = '+ str(max(prefigurance_hit_persent_list)))


## 這是用來確認colorbar的配置
fig,ax1 = plt.subplots()
X = [i for i in range(len(prefigurance_hit_persent_list))]
Y = sorted(prefigurance_hit_persent_list)
ax1.plot(X,Y,color =  'black',marker = "*",linestyle = '--') #折線圖
ax1.set_title('這是用來確認colorbar的配置')


# 顯示地圖
plt.show()