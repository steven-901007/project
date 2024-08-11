##前估
from openpyxl import load_workbook
import glob
import pandas as pd
from datetime import datetime, timedelta
from tqdm import tqdm
import matplotlib.pyplot as plt
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from cartopy.io.shapereader import Reader
from cartopy.feature import ShapelyFeature
import matplotlib.colors as mcolors
import matplotlib.cm as cm
import matplotlib as mpl



year = '2021' #年分
month = '06' #月份
dis = 36
data_top_path = "C:/Users/steve/python_data"



##測站資料
def rain_station_location_data():
    data_path = data_top_path+"/研究所/雨量資料/"+year+"測站範圍內測站數.xlsx"
    lon_data_list = []  # 經度
    lat_data_list = []  # 緯度
    name_data_list = []  #測站名稱
    wb = load_workbook(data_path)
    ws = wb[month]
    for i in range(ws.max_column):
        lon_data_list.append(ws.cell(4,i+1).value)
        lat_data_list.append(ws.cell(3,i+1).value)
        name_data_list.append(ws.cell(1,i+1).value)
    wb.close()
    return lon_data_list, lat_data_list ,name_data_list
lon_data_list, lat_data_list ,name_data_list = rain_station_location_data()

def check_in_time_range(row, lj_times):
    return int(any((lj_times >= row['start time']) & (lj_times <= row['end time'])))

##強降雨發生但沒有lighting jump


month_path = data_top_path + "/研究所/雨量資料/對流性降雨"+str(dis)+"km統計/"+year+"/"+month+"/**.csv"
result  =glob.glob(month_path)

for rain_station_path in tqdm(result,desc='資料處理中....'):
# rain_station_path = "C:/Users/steve/python_data/研究所/雨量資料/對流性降雨36km統計/2021/06/C0R490.csv"
    rain_station_name = rain_station_path.split('/')[-1].split('\\')[-1].split('.')[0]


    #flash
    try:
        flash_station_path = f"{data_top_path}/研究所/閃電資料/lighting_jump/"+str(dis)+f"km/{year}/{month}/{rain_station_name}.csv"
    except:
        flash_station_path = None

    if flash_station_path != None:
        rain_data = pd.read_csv(rain_station_path)
        flash_data = pd.read_csv(flash_station_path)
        #end time< lighting jump <= time data
        rain_data['time data'] = pd.to_datetime(rain_data['time data'])
        rain_data['start time'] = rain_data["time data"] - pd.Timedelta(minutes= 50)
        rain_data['end time'] = rain_data['time data'] + pd.Timedelta(minutes=10)
        flash_data['LJ_time'] = pd.to_datetime(flash_data['LJ_time'])
        # print(rain_data)

        rain_data['LJ_in_time_range'] = rain_data.apply(lambda row: check_in_time_range(row, flash_data['LJ_time']), axis=1)

        # print(rain_data[rain_data['LJ_in_time_range'] == 1])
        # pd.set_option('display.max_rows', None)

        # print(flash_data)
        # print(rain_data['LJ_in_time_range'].sum())
        if rain_data['LJ_in_time_range'].sum() != 0:
            time_data = pd.to_datetime(rain_data[rain_data['LJ_in_time_range'] == 1]['time data'])
            LJ_data_save_path = data_top_path + "/研究所/個案分析/前估命中個案/" + rain_station_name + '_' + str(rain_data['LJ_in_time_range'].sum()) + '.csv'
            LJ_data_save = {
                'time data' : time_data
            }
            pd.DataFrame(LJ_data_save).to_csv(LJ_data_save_path,index=False)


# prefigurance_hit_persent_list = [] # 前估命中率
# for i in range(len(total_prefigurance_list)):
#     prefigurance_hit_persent_list.append(prefigurance_hit_list[i]/(total_prefigurance_list[i]+prefigurance_hit_list[i])*100)

