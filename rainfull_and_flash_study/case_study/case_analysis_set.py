##前估
import os
from openpyxl import load_workbook
import glob
import datetime
import pandas as pd
import re
import math
from tqdm import tqdm





def haversine(lat1, lon1, lat2, lon2): #經緯度距離
    # 地球的半徑（公里）
    R = 6371.0

    # 將經緯度從度轉換為弧度
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)

    # 經緯度差
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad

    # Haversine公式
    a = math.sin(dlat / 2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    # 計算距離
    distance = R * c

    return distance

def fileset(path):    #建立資料夾
    if not os.path.exists(path):
        os.makedirs(path)
        print(path + " 已建立") 

def case_data_set(year,month,day,time_start,time_end,dis,station_name,data_top_path):

    #變數設定(測站經緯度and36km的測站有哪些)
    data_path = f"{data_top_path}/研究所/雨量資料/{year}測站範圍內測站數.xlsx"
    wb = load_workbook(data_path)
    ws = wb[month]
    stations_name_for_36km_list = []
    for i in range(ws.max_column):

        if ws.cell(1,i+1).value == station_name:
            station_lon_data = ws.cell(3,i+1).value
            station_lat_data = ws.cell(2,i+1).value
            lc = 4

            while ws.cell(lc,i+1).value != None:
                stations_name_for_36km_list.append(ws.cell(lc,i+1).value)

                lc += 1
    wb.close()


    ## 建立資料夾
    case_root_path = f"{data_top_path}/研究所/個案分析/{station_name}_{dis}_{year}{month}{day}_{str(time_start).zfill(2)}00to{str(time_end)}00"
    fileset(case_root_path)

    # time_start = datetime.datetime(time_start,0)
    # time_end = datetime.datetime(time_end,0)
    #將時間的type改成時間型態
    time_start = datetime.datetime(int(year),int(month),int(day),time_start)
    time_end = datetime.datetime(int(year),int(month),int(day),time_end)
    print(time_start,time_end)


    ##雨量raw data建立
    rain_data_time_list = []
    rain_data_of_10min_list = []
    rain_data_station_name_list = []
    rain_paths = f"{data_top_path}/研究所/雨量資料/{year}_{month}/{month}/{year}{month}{day}/**"
    rain_file_paths  =glob.glob(rain_paths)
    # print(rain_file_paths)

    for rain_path in tqdm(rain_file_paths,desc='雨量raw data讀取'):
        time_str = rain_path.split('\\')[-1].split('.')[0][8:]  # 提取時間部分
        time_str = year + month + day + time_str
        # print(time_str)
        file_time = datetime.datetime.strptime(time_str, '%Y%m%d%H%M')
        # print(file_time)
        if time_start <= file_time < time_end:
            # print(rain_path)
    # rain_path = 'C:/Users/steve/python_data/研究所/雨量資料/2021_06/06/20210604/202106041250.QPESUMS_GAUGE.10M.mdf'      
            time = rain_path.split('/')[-1].split('\\')[-1].split('.')[0]
            # print(time)
            line = 0
            with open(rain_path, 'r') as files:
                for file in files:
                    elements = re.split(re.compile(r'\s+|\n|\*'),file.strip()) 

                    if line >= 3 :  #移除檔頭
                        rain_data_of_10min = float(elements[7]) #MIN_10
                        rain_data_of_3_hour = float(elements[8]) #HOUR_3
                        rain_data_of_6_hour = float(elements[9]) #HOUR_6
                        rain_data_of_12_hour = float(elements[10]) #HOUR_12
                        rain_data_of_24_hour = float(elements[11]) #HOUR_24

                        if  10<=rain_data_of_10min <= rain_data_of_3_hour <= rain_data_of_12_hour <= rain_data_of_24_hour: #QC
                            rain_station_name = elements[0]

                            if stations_name_for_36km_list.count(rain_station_name) != 0:
                                # print(elements)
                                rain_data_time_list.append(time)
                                rain_data_station_name_list.append(rain_station_name)
                                rain_data_of_10min_list.append(rain_data_of_10min)
                                
                    line += 1

    # print(rain_data_time_list)
    rain_data_time_list = pd.to_datetime(rain_data_time_list) #改變成時間格式

    rain_data_save_path = case_root_path + '/rain_raw_data.csv'
    rain_data_save = {
        'data time' :rain_data_time_list,
        'station name' :rain_data_station_name_list,
        '10min rain' :rain_data_of_10min_list
    }
    pd.DataFrame(rain_data_save).to_csv(rain_data_save_path,index=False)
    print('雨量資料已建立')


    ##閃電資料建立(讀取依測站分類)
    flash_path = f"{data_top_path}/研究所/閃電資料/依測站分類/{dis}km/{year}/{month}/{station_name}.csv"
    flash_data = pd.read_csv(flash_path)
    flash_data['data time'] = pd.to_datetime(flash_data['data time'])
    # print(flash_data['data time'])
    save_data = flash_data[(flash_data['data time'] >= time_start) & 
                        (flash_data['data time'] < time_end)]
    # print(save_data)
    flash_data_save_path = case_root_path + '/flash_data.csv'
    save_data.to_csv(flash_data_save_path,index=False)
    print('閃電資料已建立')



# case_data_set('2021','06','04',9,13,36,'01P190',"C:/Users/steve/python_data")

