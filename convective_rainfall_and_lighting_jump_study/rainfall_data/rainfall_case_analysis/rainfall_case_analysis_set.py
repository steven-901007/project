import pandas as pd
from openpyxl import Workbook
from glob import glob
from tqdm import tqdm
from importset import fileset
import re
from openpyxl.styles import Font
from datetime import datetime, timedelta

year = '2021' #年分
month = '06' #月份
data_top_path = "C:/Users/steve/python_data"
dis = 36

station = 'V2C250'

print(station)
file_path = f"{data_top_path}/研究所/雨量資料/{str(dis)}km個案分析/{month}/{station}"
fileset(file_path)

##save data set up
save_data_wb = Workbook()
save_data_ws = save_data_wb.active


##周圍測站名稱建立
around_station_data_path = f"{data_top_path}/研究所/雨量資料/測站範圍內測站數/{year}_{month}/{station}.csv"
around_station_datas = pd.read_csv(around_station_data_path)
around_station_datas['station name'] = around_station_datas['station name'].astype(str)
# print(around_station_datas)

#測站經緯度
around_station_lon_lat_path = f"{data_top_path}/研究所/雨量資料/測站資料/{year}_{month}.csv"
around_station_lon_lat_datas = pd.read_csv(around_station_lon_lat_path)
real_name = around_station_lon_lat_datas[around_station_lon_lat_datas['station name'] == station]['station real name'].iloc[0]
print(real_name)

save_data_ws.cell(1,1).value = 'station name'
save_data_ws.cell(1,2).value = 'station lon'
save_data_ws.cell(1,3).value = 'station lat'
row = 2
for around_station in around_station_datas['station name']:
    save_data_ws.cell(row,1).value = around_station
    save_data_ws.cell(row,2).value = around_station_lon_lat_datas[around_station_lon_lat_datas['station name'] == around_station]['lon'].iloc[0]
    save_data_ws.cell(row,3).value = around_station_lon_lat_datas[around_station_lon_lat_datas['station name'] == around_station]['lat'].iloc[0]

    row += 1
del row

row_for_around_station_datas_lsit = around_station_datas['station name'].to_list() #跟excel的位置差2


##資料時間建立

col = 4

month_path = f"{data_top_path}/研究所/雨量資料/{year}_{month}/{month}"
result  =glob(month_path+"/*")

for day_path in tqdm(result,desc='資料建立'):
    day = day_path.split('\\')[-1][-2:] #日期   
    # print('日期:'+day)
    
    # ## 讀取每日資料

    result  =glob(day_path+'/*')
    for rain_data_path in result:
        time = rain_data_path.split('\\')[-1].split('.')[0]
        time_obj = datetime.strptime(time, '%Y%m%d%H%M')+ timedelta(hours=8) #將UTC轉成LT
        time = time_obj.strftime('%Y%m%d%H%M')
        save_data_ws.cell(1,col).value = time[6:]
        # print('時間:'+time)
        rain_data_list = []
        rainfall_list = []
        # 每10分鐘資料處理 rain data >10mm (10min)
        line = 0
        with open(rain_data_path, 'r') as files:
            for file in files:
                elements = re.split(re.compile(r'\s+|\n|\*'),file.strip()) 

                if line >= 3 :  #移除檔頭
                    station_name = elements[0] #測站名稱
                    rain_data_of_10min = float(elements[7]) #MIN_10
                    rain_data_of_3_hour = float(elements[8]) #HOUR_3
                    rain_data_of_6_hour = float(elements[9]) #HOUR_6
                    rain_data_of_12_hour = float(elements[10]) #HOUR_12
                    rain_data_of_24_hour = float(elements[11]) #HOUR_24

                    if 0<rain_data_of_10min <= rain_data_of_3_hour <=rain_data_of_6_hour<= rain_data_of_12_hour <= rain_data_of_24_hour: #QC and data !=0 or -999
                        try:
                            lc = row_for_around_station_datas_lsit.index(station_name)+2
                            save_data_ws.cell(lc,col).value = rain_data_of_10min
                            # print(lc,rain_data_of_10min)
                            if rain_data_of_10min >= 10:
                                save_data_ws.cell(lc,col).font = Font(bold = True , color= 'FF0000')

                        except:pass

                line += 1
        ##移除全空白的時間
        is_empty = True
        for row in range(2,len(row_for_around_station_datas_lsit)+2):
            if save_data_ws.cell(row,col).value != None:
                is_empty = False
        if is_empty == False:
            col += 1
            save_data_ws.cell(1,col).value = time[6:]


save_data_ws.title = real_name
save_data_path = f"{data_top_path}/研究所/雨量資料/{dis}km個案分析/{month}/{station}/{station}_rain_data.xlsx"
save_data_wb.save(save_data_path)
