from openpyxl import load_workbook,Workbook
import math
import statistics
import time as Ti



locate = '松山'
year = '108'
day = '2019-04-20'
confidence_level = 100
sigma = 1

startTime = Ti.time()

wb = load_workbook("C:/Users/steve/Desktop/python相關資料/need data information/"+day+"/confidence_level="+str(confidence_level)+" 10min_rowdata.xlsx")
wsv = wb['V']
wsu = wb['U']



cmax = wsu.max_column
remove = []
for i in range(2,cmax+1,3):
    remove.append(i)

for i in range(2,cmax+1,2):
    if remove.count(i) == 0:
        remove.append(i)

remove.sort()
remove.reverse()
# print(remove)
for i in remove:
    wsu.delete_cols(i)
    wsv.delete_cols(i)




wb.save("C:/Users/steve/Desktop/python相關資料/need data information/"+day+"/confidence_level="+str(confidence_level)+"sigma="+str(sigma)+" 10min.xlsx")
print(day+'\nconfidence '+str(confidence_level)+'\nsigma '+str(sigma))


endTime = Ti.time()

print('程式執行了%s秒' %(endTime-startTime))