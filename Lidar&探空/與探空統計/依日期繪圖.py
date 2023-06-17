from openpyxl import load_workbook
import glob
import matplotlib.pyplot as plt



days = []
vs = []
ds = []
file = glob.glob("C:/Users/steve/Desktop/python相關資料/觀測rowdata/統計/*")
for f in file:
    # print(f)
    day = f[47:57]
    print(day)
    wb =load_workbook(f)
    wsv = wb['風速']
    wsd = wb['風向']
    if wsv.cell(3,5).value != None and wsd.cell(3,5).value != None:
        vs.append(round(wsv.cell(3,5).value,3))
        ds.append(round(wsd.cell(3,5).value,3))
        days.append(day)

v = 0
for i in range(len(vs)):
    v += vs[i]    
avev = v/len(vs)

d = 0
for i in range(len(ds)):
    d += ds[i]    

aved = d/len(ds)

xtick = []
for i in range(0,len(days),20):
    xtick.append(days[i])
if xtick[len(xtick)-1] != days[len(days)-1]:
    xtick.append(days[len(days)-1])

# print(days)
# print(vs)
# print(ds)
plt.rcParams['font.sans-serif'] = [u'MingLiu'] #細明體
plt.rcParams['axes.unicode_minus'] = False #設定中文

fig = plt.figure()
ax = fig.add_subplot()
ax.scatter(days,ds,color = 'red',s=5,label= '風向')
ax.scatter(days,vs,color = 'blue',s=5,label = '風速')
ax.axhline(y = avev ,c = "skyblue" , ls = "--" , lw = 1 ,label = '風速平均')
ax.axhline(y = aved ,c = "purple" , ls = "--" , lw = 1 ,label = '風向平均')
ax.set_xticks(xtick)
plt.xticks(rotation = 75)
plt.legend(loc='upper right')
plt.title('自己計算風速風向相關係數')
plt.show()