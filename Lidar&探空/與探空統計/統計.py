from openpyxl import load_workbook,Workbook
import math
from math import *
import numpy as np
import glob
def windv(u,v): #風速
    return round(sqrt(pow(u,2) +pow(v,2)),3)

def windd(u,v): #風向
    if u == 0 and v == 0:
        return None
    else:
        return round(180+math.atan2(u,v)*180/(math.pi),2)

def ins(a,b,t,a1,b1): #內插
    return round((b1-a1)*(t-a)/(b-a)+a1,3)


# day = '2019-01-01'

files = glob.glob("C:/Users/steve/Desktop/python相關資料/觀測rowdata/need_information/*")
for file in files:  
    day = file[61:71]
    try:
        wb1hour = load_workbook("C:/Users/steve/Desktop/python相關資料/need data information/松山/108/1hour/" + day + ".xlsx") #lidar
        wbneed = load_workbook('C:/Users/steve/Desktop/python相關資料/觀測rowdata/need_information/' + day + '_00.xlsx') #探空
        ws1houru = wb1hour['U']
        ws1hourv = wb1hour['V']
        ws1hourmv = wb1hour['mv']
        ws1hourmd = wb1hour['md']
        wsneedv = wbneed['風速']
        wsneedd = wbneed['風向']
        maxhight1hour = ws1houru.max_row
        maxhightneed = wsneedv.max_row
        # print(maxhight1hour)
        lc = 0
        for i in range(2,ws1houru.max_column+1):
            if ws1houru.cell(1,i).value == '08-00':
                lc = i
                break

        wb = Workbook()
        wsv = wb.active
        wsv.title = '風速'
        wb.create_sheet('風向')
        wsd = wb['風向']
        for i in range(2,maxhight1hour+1):
            wsv.cell(i,1).value = ws1houru.cell(i,1).value
            wsd.cell(i,1).value = ws1houru.cell(i,1).value
            wsv.cell(i,6).value = ws1hourmv.cell(i,lc).value
            wsd.cell(i,6).value = ws1hourmd.cell(i,lc).value   
        wsv.cell(1,2).value = 'lidar自算'
        wsv.cell(1,3).value = '探空'
        wsv.cell(2,5).value = '相關係數'
        wsv.cell(1,6).value = 'lidar電腦算'
        wsv.cell(2,8).value = '相關係數'

        wsd.cell(1,2).value = 'lidar自算'
        wsd.cell(1,3).value = '探空'
        wsd.cell(1,4).value = '相關係數'
        wsd.cell(2,5).value = '相關係數'
        wsd.cell(1,6).value = 'lidar電腦算'
        wsd.cell(1,7).value = '相關係數'
        wsd.cell(2,8).value = '相關係數'
        
        time1hour = [[],[],[]] #[0] = 目標高度,[1] = 目標風速,[2] = 目標風向
        timeneed = [[],[],[]] #[0] = 內插高度,[1] = 內插風速,[2] = 內插風向

        for i in range(2,maxhight1hour+1):
            if ws1houru.cell(i,lc).value != None:
                time1hour[0].append(ws1houru.cell(i,1).value)
                time1hour[1].append(windv(ws1houru.cell(i,lc).value,ws1hourv.cell(i,lc).value))
                time1hour[2].append(windd(ws1houru.cell(i,lc).value,ws1hourv.cell(i,lc).value))
            else:
                time1hour[0].append(ws1houru.cell(i,1).value)
                time1hour[1].append(None)
                time1hour[2].append(None)
        # print(time1hour[0])
        # print(time1hour[1])
        # print(time1hour[2])
        for i in time1hour[0]: 
            # print(i)
            for j in range(2,maxhightneed+1):
                if wsneedv.cell(j,1).value<=i<wsneedv.cell(j+1,1).value:
                    # print(i) #t
                    # print(wsneedv.cell(j,1).value) #a
                    # print(wsneedv.cell(j+1,1).value) #b
                    # print(wsneedv.cell(j,2).value) #a1
                    # print(wsneedv.cell(j+1,2).value) #b1
                    insv = ins(wsneedv.cell(j,1).value,wsneedv.cell(j+1,1).value,i,wsneedv.cell(j,2).value,wsneedv.cell(j+1,2).value) #內插好的風速
                    # print(insv)
                    timeneed[1].append(insv)
                    insd = ins(wsneedd.cell(j,1).value,wsneedd.cell(j+1,1).value,i,wsneedd.cell(j,2).value,wsneedd.cell(j+1,2).value) #內插好的風向
                    # print(insd)
                    timeneed[2].append(insd)
                    break
            timeneed[0].append(i)
        # print(timeneed[0])
        # print(timeneed[1])
        # print(timeneed[2])
        l = 2
        for i in range(len(timeneed[0])):
            wsv.cell(l,2).value = time1hour[1][i]
            wsv.cell(l,3).value = timeneed[1][i]    
            wsd.cell(l,2).value = time1hour[2][i]
            wsd.cell(l,3).value = timeneed[2][i]
            if wsd.cell(l,2).value != None and wsd.cell(l,3).value !=None:
                d = abs(wsd.cell(l,2).value - wsd.cell(l,3).value)
            else:
                d = None

            if wsd.cell(l,6).value != None and wsd.cell(l,3).value !=None:
                D = abs(wsd.cell(l,6).value - wsd.cell(l,3).value)
            else:
                D = None
                
            if d == None:
                wsd.cell(l,4).value = None
            elif 0<d<=180:
                wsd.cell(l,4).value = 1-(d/90)
            elif 180<d<=360:
                wsd.cell(l,4).value = (d/90)-3
            else:
                wsd.cell(l,4).value = 'ERROR'
                print(day+'error')

            if D == None:
                wsd.cell(l,7).value = None
            elif 0<D<=180:
                wsd.cell(l,7).value = 1-(D/90)
            elif 180<D<=360:
                wsd.cell(l,7).value = (D/90)-3
            else:
                wsd.cell(l,7).value = 'ERROR'
                print(day+'error')
            l +=1

        for i in range(len(timeneed[0])):
            if timeneed[1][i] == None or time1hour[1][i] == None:
                time1hour[1][i] = None
                timeneed[1][i] = None
        for i in range(len(time1hour[1])):
            try:
                time1hour[1].remove(None)
                timeneed[1].remove(None)
            except:pass

        # print(time1hour[1])
        # print(timeneed[1])
        # print(len(time1hour[1]))
        # print(len(timeneed[1]))

        x = np.corrcoef(time1hour[1],timeneed[1])
        # print(x[0][1])
        wsv.cell(3,5).value = x[0][1]
        # print(len(timeneed[1]))

        lidarmv = []
        for i in range(2,maxhight1hour+1):
            if wsv.cell(i,6).value != None and wsv.cell(i,3).value != None:
                lidarmv.append(wsv.cell(i,6).value)

        x = np.corrcoef(lidarmv,timeneed[1])
        wsv.cell(3,8).value = x[0][1]
        
        a = 0
        d = 0
        for i in range(2,wsd.max_row+1):
            if wsd.cell(i,4).value != None:
                a += wsd.cell(i,4).value
                d +=1
        try:
            wsd.cell(3,5).value = a/d
        except:pass

        for i in range(2,wsd.max_row+1):
            if wsd.cell(i,7).value != None:
                a += wsd.cell(i,7).value
                d +=1
        try:
            wsd.cell(3,8).value = a/d
        except:pass
        wb.save("C:/Users/steve/Desktop/python相關資料/觀測rowdata/統計/"+ day +'_00.xlsx')   
        print(day+"_00")

        wbneed = load_workbook('C:/Users/steve/Desktop/python相關資料/觀測rowdata/need_information/' + day + '_12.xlsx') #探空
        ws1houru = wb1hour['U']
        ws1hourv = wb1hour['V']
        ws1hourmv = wb1hour['mv']
        ws1hourmd = wb1hour['md']
        wsneedv = wbneed['風速']
        wsneedd = wbneed['風向']
        maxhight1hour = ws1houru.max_row
        maxhightneed = wsneedv.max_row
        # print(maxhight1hour)
        lc = 0
        for i in range(2,ws1houru.max_column+1):
            if ws1houru.cell(1,i).value == '20-00':
                lc = i
                break

        wb = Workbook()
        wsv = wb.active
        wsv.title = '風速'
        wb.create_sheet('風向')
        wsd = wb['風向']
        for i in range(2,maxhight1hour+1):
            wsv.cell(i,1).value = ws1houru.cell(i,1).value
            wsd.cell(i,1).value = ws1houru.cell(i,1).value
            wsv.cell(i,6).value = ws1hourmv.cell(i,lc).value
            wsd.cell(i,6).value = ws1hourmd.cell(i,lc).value   
        wsv.cell(1,2).value = 'lidar自算'
        wsv.cell(1,3).value = '探空'
        wsv.cell(2,5).value = '相關係數'
        wsv.cell(1,6).value = 'lidar電腦算'
        wsv.cell(2,8).value = '相關係數'

        wsd.cell(1,2).value = 'lidar自算'
        wsd.cell(1,3).value = '探空'
        wsd.cell(1,4).value = '相關係數'
        wsd.cell(2,5).value = '相關係數'
        wsd.cell(1,6).value = 'lidar電腦算'
        wsd.cell(1,7).value = '相關係數'
        wsd.cell(2,8).value = '相關係數'
        
        time1hour = [[],[],[]] #[0] = 目標高度,[1] = 目標風速,[2] = 目標風向
        timeneed = [[],[],[]] #[0] = 內插高度,[1] = 內插風速,[2] = 內插風向

        for i in range(2,maxhight1hour+1):
            if ws1houru.cell(i,lc).value != None:
                time1hour[0].append(ws1houru.cell(i,1).value)
                time1hour[1].append(windv(ws1houru.cell(i,lc).value,ws1hourv.cell(i,lc).value))
                time1hour[2].append(windd(ws1houru.cell(i,lc).value,ws1hourv.cell(i,lc).value))
            else:
                time1hour[0].append(ws1houru.cell(i,1).value)
                time1hour[1].append(None)
                time1hour[2].append(None)
        # print(time1hour[0])
        # print(time1hour[1])
        # print(time1hour[2])
        for i in time1hour[0]: 
            # print(i)
            for j in range(2,maxhightneed+1):
                if wsneedv.cell(j,1).value<=i<wsneedv.cell(j+1,1).value:
                    # print(i) #t
                    # print(wsneedv.cell(j,1).value) #a
                    # print(wsneedv.cell(j+1,1).value) #b
                    # print(wsneedv.cell(j,2).value) #a1
                    # print(wsneedv.cell(j+1,2).value) #b1
                    insv = ins(wsneedv.cell(j,1).value,wsneedv.cell(j+1,1).value,i,wsneedv.cell(j,2).value,wsneedv.cell(j+1,2).value) #內插好的風速
                    # print(insv)
                    timeneed[1].append(insv)
                    insd = ins(wsneedd.cell(j,1).value,wsneedd.cell(j+1,1).value,i,wsneedd.cell(j,2).value,wsneedd.cell(j+1,2).value) #內插好的風向
                    # print(insd)
                    timeneed[2].append(insd)
                    break
            timeneed[0].append(i)
        # print(timeneed[0])
        # print(timeneed[1])
        # print(timeneed[2])
        l = 2
        for i in range(len(timeneed[0])):
            wsv.cell(l,2).value = time1hour[1][i]
            wsv.cell(l,3).value = timeneed[1][i]    
            wsd.cell(l,2).value = time1hour[2][i]
            wsd.cell(l,3).value = timeneed[2][i]
            if wsd.cell(l,2).value != None and wsd.cell(l,3).value !=None:
                d = abs(wsd.cell(l,2).value - wsd.cell(l,3).value)
            else:
                d = None

            if wsd.cell(l,6).value != None and wsd.cell(l,3).value !=None:
                D = abs(wsd.cell(l,6).value - wsd.cell(l,3).value)
            else:
                D = None
                
            if d == None:
                wsd.cell(l,4).value = None
            elif 0<d<=180:
                wsd.cell(l,4).value = 1-(d/90)
            elif 180<d<=360:
                wsd.cell(l,4).value = (d/90)-3
            else:
                wsd.cell(l,4).value = 'ERROR'
                print(day+'error')

            if D == None:
                wsd.cell(l,7).value = None
            elif 0<D<=180:
                wsd.cell(l,7).value = 1-(D/90)
            elif 180<D<=360:
                wsd.cell(l,7).value = (D/90)-3
            else:
                wsd.cell(l,7).value = 'ERROR'
                print(day+'error')
            l +=1

        for i in range(len(timeneed[0])):
            if timeneed[1][i] == None or time1hour[1][i] == None:
                time1hour[1][i] = None
                timeneed[1][i] = None
        for i in range(len(time1hour[1])):
            try:
                time1hour[1].remove(None)
                timeneed[1].remove(None)
            except:pass

        # print(time1hour[1])
        # print(timeneed[1])
        # print(len(time1hour[1]))
        # print(len(timeneed[1]))

        x = np.corrcoef(time1hour[1],timeneed[1])
        # print(x[0][1])
        wsv.cell(3,5).value = x[0][1]
        # print(len(timeneed[1]))

        lidarmv = []
        for i in range(2,maxhight1hour+1):
            if wsv.cell(i,6).value != None and wsv.cell(i,3).value != None:
                lidarmv.append(wsv.cell(i,6).value)

        x = np.corrcoef(lidarmv,timeneed[1])
        wsv.cell(3,8).value = x[0][1]
        
        a = 0
        d = 0
        for i in range(2,wsd.max_row+1):
            if wsd.cell(i,4).value != None:
                a += wsd.cell(i,4).value
                d +=1
        try:
            wsd.cell(3,5).value = a/d
        except:pass

        for i in range(2,wsd.max_row+1):
            if wsd.cell(i,7).value != None:
                a += wsd.cell(i,7).value
                d +=1
        try:
            wsd.cell(3,8).value = a/d
        except:pass
        wb.save("C:/Users/steve/Desktop/python相關資料/觀測rowdata/統計/"+ day +'_12.xlsx')   
        print(day+"_12")
 
    except:pass