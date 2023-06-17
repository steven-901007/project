from openpyxl import load_workbook,Workbook
import statistics
import numpy as np


file10 = "C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/統計/10min.xlsx"
wb = load_workbook ("C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/u_all.xlsx")
ws = wb['Sheet']
wb10 = load_workbook(file10)
ws10 = wb10['u']
colmax = ws.max_column #4355
rowmax = ws.max_row #64
# print(colmax,rowmax)

group = [] #len = 1440
for i in range(2,colmax+1):

    if ws.cell(1,i).value !=ws.cell(1,i+1).value:
        # group.append(ws.cell(1,i).value)
        group.append(i-1)
# print (group[1435])
location = 2
for time in range(6,1427,10):
    
    col = []
    Col = []
    for j in range(2,rowmax+1):
        for i in range(group[time]-1,group[time+9]+2):
            if ws.cell(j,i).value != None:
                col.append(ws.cell(j,i).value)
                Col.append(ws.cell(j,i).value)
        try:
            col.remove(max(col))
            col.remove(min(col))
            sigma = statistics.pstdev(col)
            mean = statistics.mean(col)
            for i in range(len(Col)):
                if mean-1.5*sigma>=Col[i] or Col[i]>=mean+1.5*sigma:
                    Col[i] = -999

            nb =0

            for a in range(len(Col)):
                if Col[a] == -999:
                    nb +=1
            

            for b in range(nb):
                # print (i)
                Col.remove(-999)

            ws10.cell(j,location).value= round(statistics.mean(Col),2)
        except:
            ws10.cell(j,location).value= -999            
    location +=1
lc = 2
for hour in range(0,24):
    for men in range(0,51,10):
        if hour !=0 or men !=0:
            hour = str(hour).zfill(2)
            men = str(men).zfill(2)            
            ws10.cell(1,lc).value = hour + '-' + men
            lc +=1
lc = 2
for h in range(50,1601,25):
    ws10.cell(lc,1).value = h 
    lc +=1
wb10.save(file10)


wb = load_workbook ("C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/v_all.xlsx")
ws = wb['Sheet']
wb10 = load_workbook(file10)
ws10 = wb10['v']
colmax = ws.max_column #4355
rowmax = ws.max_row #64
# print(colmax,rowmax)

group = [] #len = 1440
for i in range(2,colmax+1):

    if ws.cell(1,i).value !=ws.cell(1,i+1).value:
        # group.append(ws.cell(1,i).value)
        group.append(i-1)
# print (group[1435])
location = 2
for time in range(6,1427,10):
    
    col = []
    Col = []
    for j in range(2,rowmax+1):
        for i in range(group[time]-1,group[time+9]+2):
            if ws.cell(j,i).value != None:
                col.append(ws.cell(j,i).value)
                Col.append(ws.cell(j,i).value)
        try:
            col.remove(max(col))
            col.remove(min(col))
            sigma = statistics.pstdev(col)
            mean = statistics.mean(col)
            for i in range(len(Col)):
                if mean-1.5*sigma>=Col[i] or Col[i]>=mean+1.5*sigma:
                    Col[i] = -999

            nb =0

            for a in range(len(Col)):
                if Col[a] == -999:
                    nb +=1
            

            for b in range(nb):
                # print (i)
                Col.remove(-999)

            ws10.cell(j,location).value= round(statistics.mean(Col),2)
        except:
            ws10.cell(j,location).value= -999 # 
    location +=1
lc = 2
for hour in range(0,24):
    for men in range(0,51,10):
        if hour !=0 or men !=0:
            hour = str(hour).zfill(2)
            men = str(men).zfill(2)            
            ws10.cell(1,lc).value = hour + '-' + men
            lc +=1
lc = 2
for h in range(50,1601,25):
    ws10.cell(lc,1).value = h 
    lc +=1
wb10.save(file10)