from openpyxl import load_workbook,Workbook

wb = load_workbook('C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/統計/10min.xlsx')
wsu = wb['u']
wsv = wb['v']
wb1 = load_workbook('C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/統計/1hour.xlsx')
ws1u = wb1['u']
ws1v = wb1['v']
lc = 2
for time in range(7,140,6):    
    # print(wsu.cell(1,time).value)
    for hight in range(1,65):
        # print(wsu.cell(hight,time).value)
        ws1u.cell(hight,lc).value = wsu.cell(hight,time).value
    lc+=1
lc = 2
for time in range(7,140,6):    
    # print(wsu.cell(1,time).value)
    for hight in range(1,65):
        # print(wsu.cell(hight,time).value)
        ws1v.cell(hight,lc).value = wsv.cell(hight,time).value
    lc+=1
lc = 2
for i in range(50,1601,25):
    ws1u.cell(lc,1).value = i
    ws1v.cell(lc,1).value = i
    lc +=1

wb1.save('C:/Users/steve/Desktop/python/松山剖風儀資料/2019-03-13/統計/1hour.xlsx')

# wb1.create_sheet('v')