from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime
xlsxlc = "C:/Users/steve/Desktop/股票_媽/股票收盤價.xlsx"
wb = load_workbook(xlsxlc)
ws = wb['收盤價']
col = ws.max_column
row =ws.max_row #高


path = "C:/Users/steve/Downloads/chromedriver_win32/chromedriver.exe" #chromedriver位置

for i in range(2,col):
    try:
        stock = ws.cell(1,i).value
        print(stock)
        driver = webdriver.Chrome(path)
        driver.get("https://www.cnyes.com/twstock/"+str(stock)) #目標頁面
        # print(driver.title) #網頁title
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[1]/div[4]/div[2]/div[1]/div[1]/div[1]/div/div[2]/div[2]/div/h3"))) #系統搜尋到某些東西時才繼續進行

        tg = driver.find_element(By.XPATH,"/html/body/div[1]/div[1]/div[4]/div[2]/div[1]/div[1]/div[1]/div/div[2]/div[2]/div/h3")
        # print(tg.text)
        ws.cell(row+1,i).value = float(tg.text)*int(ws.cell(2,i).value)
        driver.quit() #關閉網頁 
    except:
        ws.cell(row+1,i).value = '資料輸入錯誤'



t = datetime.now()
now = t.strftime("%Y/%m/%d")
# print(now)
ws.cell(row+1,1).value = now

ws.cell(row+1,10).value = t
wb.save(xlsxlc)
wb.close()