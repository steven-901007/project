from openpyxl import Workbook
import random 
wb = Workbook()
ws = wb.active


ws.title = 'random'
for a in range(1,10000):
    ws.cell(a,1).value = random.randint(-10,100)
wb.save('123.xlsx')

ws.title = '工作表1'
for i in range(1,100):
    ws.cell(i,1).value = random.randint(1,100)

for j in range(1,100):
    ws.cell(j,2).value = random.randint(1,10)
wb.save('randomnb.xlsx')